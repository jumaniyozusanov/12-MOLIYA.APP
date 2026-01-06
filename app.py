import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import os
from functools import reduce
from openpyxl.utils import get_column_letter
import string
from openpyxl.cell.cell import MergedCell
from io import BytesIO
import tempfile
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import cm
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
from reportlab.lib.pagesizes import A4, landscape
import openpyxl




# ===== SOZLAMALAR =====
st.set_page_config(page_title="12-Moliya", layout="centered")

# ===== LOGIN MA'LUMOTLARI =====
USERNAME = "12-moliya"
PASSWORD = "2026"

# ===== SESSION =====
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

# ===== LOGIN SAHIFA =====
if not st.session_state.logged_in:
    st.title("🔐 Kirish")

    username = st.text_input("Login")
    password = st.text_input("Parol", type="password")

    if st.button("Kirish"):
        if username == USERNAME and password == PASSWORD:
            st.session_state.logged_in = True
            st.success("Kirish muvaffaqiyatli!")
            st.rerun()
        else:
            st.error("Login yoki parol xato")







else:


# ==========================
# EXCEL → PDF FUNKSIYA
# ==========================



# def excel_sheets_to_pdf(excel_file):
#     """
#     Excel fayldagi barcha sheetlarni PDFga aylantiradi.
#     Excel fayl BytesIO yoki UploadedFile bo'lishi mumkin.
#     """
#     # Excelni o'qish
#     xls = pd.ExcelFile(excel_file)
    
#     # PDF buffer
#     pdf_buffer = BytesIO()
#     doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
#     elements = []

#     for sheet_name in xls.sheet_names:
#         df = xls.parse(sheet_name)
#         if df.empty:
#             continue

#         # DataFrame ni list of lists ko'rinishiga o'tkazish
#         data = [df.columns.tolist()] + df.values.tolist()

#         # Table yaratish
#         table = Table(data, repeatRows=1)
#         table.setStyle(
#             TableStyle([
#                 ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#D3D3D3")), # Header fon
#                 ('TEXTCOLOR', (0,0), (-1,0), colors.black),
#                 ('ALIGN', (0,0), (-1,-1), 'CENTER'),
#                 ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
#                 ('FONTSIZE', (0,0), (-1,-1), 9),
#                 ('BOTTOMPADDING', (0,0), (-1,0), 6),
#                 ('GRID', (0,0), (-1,-1), 0.5, colors.black),
#             ])
#         )
#         elements.append(table)
#         elements.append(PageBreak())  # Har sheet alohida sahifa

#     # PDF yaratish
#     doc.build(elements)
#     pdf_buffer.seek(0)
#     return pdf_buffer






# ==========================
# STREAMLIT ILOVA
# ==========================
    st.title("💰 Moliyaviy Hisobot Dasturi")
    menu = ["12-Moliya", "Moliya Natija", "Hisobot"]
    choice = st.sidebar.selectbox("Sahifa tanlang", menu)

    # ==========================
    # 12-MOLIYA
    # ==========================
    if choice == "12-Moliya":
        st.header("📁 12-Moliya uchun fayllarni tanlang")
        katalog1 = st.file_uploader("1-oy katalog", type=["xlsx"])
        baza1 = st.file_uploader("1-oy baza", type=["xlsx"])
        katalog2 = st.file_uploader("2-oy katalog", type=["xlsx"])
        baza2 = st.file_uploader("2-oy baza", type=["xlsx"])
        
        template = st.file_uploader("Shablon", type=["xlsx"])
        output_name = st.text_input("Natija nomi")
        # output_folder = st.text_input("Papka yo'li (masalan C:/Users/User/Desktop)")

        def numeric_like_kivy(df, cols):
            for col in cols:
                if col in df.columns:
                    df[col] = (
                        df[col]
                        .astype(str)
                        .str.replace(",", "", regex=False)
                        .str.strip()
                    )
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
                return df
            # --- Excelga yozish funksiyasi 
        from openpyxl.styles import Alignment
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils.dataframe import dataframe_to_rows

        def yoz_sheetga(df, wb, sheet_name, start_row=4, start_col=1):
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            df = df.reset_index(drop=True)

            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=False),
                start=start_row
            ):
                for c_idx, value in enumerate(row, start=start_col):
                    cell = ws.cell(row=r_idx, column=c_idx)

                    if isinstance(cell, MergedCell):
                        continue

                    cell.value = value
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.0'



        





        if st.button("START"):
            if katalog1 is None:
                st.error("❌ Katalog1 fayli tanlanmadi")
            elif baza1 is None:
                st.error("❌ Baza1 fayli tanlanmadi")
            elif katalog2 is None:
                st.error("❌ katalog2 fayli tanlanmadi")
            elif baza2 is None:
                st.error("❌ baza2 fayli tanlanmadi")
            elif template is None:
                st.error("❌ Shablon fayli tanlanmadi")
            elif not output_name.strip():
                st.error("❌ Natija nomini kiriting")
            # elif not output_folder.strip():
            #     st.error("❌ Papka yo‘lini kiriting")
            else:
                st.success("✅ Hamma narsa joyida, ishni boshlaymiz")


        if st.button("🚪 Chiqish"):
            st.session_state.logged_in = False
            st.rerun()


            # 1️⃣ Excel’larni TO‘G‘RI o‘qish
            katalog08 = pd.read_excel(
                katalog1,
                dtype={"OKPO": str, "SOOGU": str, "OKED": str, "ADRES": str}
            )

            baza08 = pd.read_excel(baza1, dtype={"OKPO": str})

            katalog09 = pd.read_excel(
                katalog2,
                dtype={"OKPO": str, "SOOGU": str, "OKED": str, "ADRES": str}
            )

            baza09 = pd.read_excel(baza2, dtype={"OKPO": str})


            # 2️⃣ FAQAT numeric ustunlar
            baza08 = numeric_like_kivy(
                baza08, ["G1","G2","G3","G4","G5","G6","G7"]
            )
            baza09 = numeric_like_kivy(
                baza09, ["G1","G2","G3","G4","G5","G6","G7"]
            )



            KATALOH08=katalog08[["OKPO","ADRES","SOOGU"]]
            KATALOH08.columns = KATALOH08.columns.str.strip()
            BAZA08=baza08[["OKPO","G1","G2","G3","G4","G5","G6","G7","SATR"]]
            BAZA08.columns=BAZA08.columns.str.strip()
            BAZA08=BAZA08[BAZA08["SATR"]==201]




            JAMI08=pd.merge(BAZA08,KATALOH08,on="OKPO",how="left")
            JAMI08=JAMI08[["SOOGU","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format

            soogu_data = [
                ["00000", "Республика бўйича жами"],
                ["00001", "Айрим вазирлик ва идоралар бўйича"],
                ["04403", "Ўзбекистон Республикаси Қурилиш ва уй-жой коммунал хўжалиги вазирлиги"],
                ["01354", "«Ўзавтосаноат» АЖ"],
                ["08114", "«Ўздонмаҳсулот» АК"],
                ["08654", "«Ўзкимёсаноат» АЖ"],
                ["06264", "«Ўзагротехсаноатхолдинг» АЖ"],
                ["03504", "«Ўзбекистон темир йўллари» АЖ"],
                ["01024", "«Ўзбекнефтгаз» АЖ"],
                ["01124", "«Ҳудудгазтаъминот» АЖ"],
                ["01104", "«Ўзтрансгаз» АЖ"],
                ["06224", "«Ўзпахтасаноат» АЖ"],
                ["01014", "«Ўзбекистон миллий электр тармоқлари» АЖ"],
                ["01074", "«Ҳудудий электр тармоқлари» АЖ"],
                ["01094", "«Иссиқлик электр станциялари» АЖ"],
                ["08524", "Ўзбекистон Республикаси Тоғ-кон саноати ва геология вазирлиги"],
                ["00002", "шу жумладан:"],
                ["06213", "«Олмалиқ кон-металлургия комбинати» АЖ"],
                ["01164", "«Навоий кон-металлургия комбинати» АЖ"],
                ["99999", "Бошқалар"]
            ]

            # --- DataFrame ---
            soogu_df = pd.DataFrame(soogu_data, columns=["SOOGU", "NAIMUZ"])

            # === 1. Asosiy baza (JMAI09) ===
            # JMAI09 = pd.read_excel("JMAI09.xlsx")  # agar fayl bo‘lsa shu yo‘l bilan o‘qisan

            # soogu kodlarini matn sifatida (boshidagi 0 saqlanib) olish
            JAMI08["SOOGU"] = JAMI08["SOOGU"].astype(str).str.zfill(5)

            # === 2. Tartib jadvali (asosiy ro‘yxat) ===
            tartib = [
                ["00000", "Республика бўйича жами"],
                ["00001", "Айрим вазирлик ва идоралар бўйича"],
                ["04403", "Ўзбекистон Республикаси Қурилиш ва уй-жой коммунал хўжалиги вазирлиги"],
                ["01354", "«Ўзавтосаноат» АЖ"],
                ["08114", "«Ўздонмаҳсулот» АК"],
                ["08654", "«Ўзкимёсаноат» АЖ"],
                ["06264", "«Ўзагротехсаноатхолдинг» АЖ"],
                ["03504", "«Ўзбекистон темир йўллари» АЖ"],
                ["01024", "«Ўзбекнефтгаз» АЖ"],
                ["01124", "«Ҳудудгазтаъминот» АЖ"],
                ["01104", "«Ўзтрансгаз» АЖ"],
                ["06224", "«Ўзпахтасаноат» АЖ"],
                ["01014", "«Ўзбекистон миллий электр тармоқлари» АЖ"],
                ["01074", "«Ҳудудий электр тармоқлари» АЖ"],
                ["01094", "«Иссиқлик электр станциялари» АЖ"],
                ["08524", "Ўзбекистон Республикаси Тоғ-кон саноати ва геология вазирлиги"],
                ["00002", "шу жумладан:"],
                ["06213", "«Олмалиқ кон-металлургия комбинати» АЖ"],
                ["01164", "«Навоий кон-металлургия комбинати» АЖ"],
                ["99999", "Бошқалар"]
            ]
            tartib_df = pd.DataFrame(tartib, columns=["SOOGU", "NAIMUZ"])

            # === 3. JMAI09 da soogu bo‘yicha yig‘ish ===
            agg = JAMI08.groupby("SOOGU", as_index=False)[["G1", "G2"]].sum()

            # === 4. Tartib bilan birlashtirish ===
            merged = pd.merge(tartib_df, agg, on="SOOGU", how="left")

            ayrm = ["04403","01354","08114","08654","06264","03504","01024","01124","01104","06224","01014","01074","01094","08524","06213","01164"
                    ]

            # === 5. “Айрим вазирлик ва идоралар бўйича” uchun sum ===
            # Bu kodlar “asosiy” ro‘yxatda yo‘q, lekin “idora” sifatida kiritiladi
            boshqamin = ["01144", "04043", "04413", "06184", "07254"]
            ayrim_sum = JAMI08.loc[JAMI08["SOOGU"].isin(ayrm), ["G1", "G2"]].sum()
            merged.loc[merged["SOOGU"] == "00001", ["G1", "G2"]] = ayrim_sum.values

            # === 6. “Республика бўйича жами” uchun barcha yig‘indi ===
            total_sum = JAMI08[["G1", "G2"]].sum()
            merged.loc[merged["SOOGU"] == "00000", ["G1", "G2"]] = total_sum.values

            # === 7. “Бошқалар” uchun qolgan kodlar sum ===
            asosiy_kodlar = tartib_df["SOOGU"].tolist() + boshqamin
            boshqalar_sum = JAMI08.loc[~JAMI08["SOOGU"].isin(asosiy_kodlar), ["G1", "G2"]].sum()
            merged.loc[merged["SOOGU"] == "99999", ["G1", "G2"]] = boshqalar_sum.values

            # === 8. To‘ldirish va yaxlitlash ===
            merged[["G1", "G2"]] = merged[["G1", "G2"]].fillna(0).round(1) / 1000

            # === 9. Yakuniy natija ===
            HISOB08 = merged[["NAIMUZ", "G1", "G2"]]


            HISOB08=HISOB08.replace("0","-")
            wb = load_workbook(template)
            #---------------------------------------------------------------------------------------------------------------------
            ##---------------------------------------------------------------------------------------------------------------------






            KATALOH09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]
            BAZA09=BAZA09[BAZA09["SATR"]==201]

            JAMI09=pd.merge(BAZA09,KATALOH09,on="OKPO",how="left")
            JAMI09=JAMI09[["SOOGU","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format

            # ==============================
            # 1️⃣ Bazani tayyorlash
            # ==============================
            # Masalan: JMAI09 = pd.read_excel("JMAI09.xlsx")
            JAMI09["SOOGU"] = JAMI09["SOOGU"].astype(str).str.zfill(5)

            # ==============================
            # 2️⃣ Asosiy SOOGU ro‘yxati
            # ==============================
            asosiy_kodlar = [
                "01014","01024","01074","01094","01104","01124","01144","01164",
                "01354","03504","04043","04403","04413","06184","06213","06224",
                "06264","07254","08114","08524","08654"
            ]

            # ==============================
            # 3️⃣ Asosiy kodlar bo‘yicha yig‘indi
            # ==============================
            asosiy_sum = (
                JAMI09[JAMI09["SOOGU"].isin(asosiy_kodlar)]
                .groupby("SOOGU", as_index=False)[["G1","G2"]]
                .sum()
            )

            # ==============================
            # 4️⃣ Qolganlarini "Бошқалар"ga yig‘ish
            # ==============================
            boshqalar_sum = (
                JAMI09[~JAMI09["SOOGU"].isin(asosiy_kodlar)][["G1","G2"]]
                .sum()
                .to_frame()
                .T
            )
            boshqalar_sum.insert(0, "SOOGU", "99999")

            # ==============================
            # 5️⃣ Barchasini birlashtirish
            # ==============================
            yakuniy = pd.concat([asosiy_sum, boshqalar_sum], ignore_index=True)

            # Respublika bo‘yicha jami
            res_sum = yakuniy[["G1","G2"]].sum().to_frame().T
            res_sum.insert(0, "SOOGU", "00000")
        

            # Айрим вазирлик ва идоралар бўйича (asosiylar yig‘indisi)
            ayrim_sum = asosiy_sum[["G1","G2"]].sum().to_frame().T
            ayrim_sum.insert(0, "SOOGU", "00001")

            yakuniy = pd.concat([res_sum, ayrim_sum, yakuniy], ignore_index=True)

            # ==============================
            # 6️⃣ Mapping jadvalini yaratamiz
            # ==============================
            mapping_list = [
                ["00000", "Республика бўйича жами"],
                ["00001", "Айрим вазирлик ва идоралар бўйича"],
                ["04403", "Ўзбекистон Республикаси Қурилиш ва уй-жой коммунал хўжалиги вазирлиги"],
                ["01354", "«Ўзавтосаноат» АЖ"],
                ["08114", "«Ўздонмаҳсулот» АК"],
                ["08654", "«Ўзкимёсаноат» АЖ"],
                ["06264", "«Ўзагротехсаноатхолдинг» АЖ"],
                ["03504", "«Ўзбекистон темир йўллари» АЖ"],
                ["01024", "«Ўзбекнефтгаз» АЖ"],
                ["01124", "«Ҳудудгазтаъминот» АЖ"],
                ["01104", "«Ўзтрансгаз» АЖ"],
                ["06224", "«Ўзпахтасаноат» АЖ"],
                ["01014", "«Ўзбекистон миллий электр тармоқлари» АЖ"],
                ["01074", "«Ҳудудий электр тармоқлари» АЖ"],
                ["01094", "«Иссиқлик электр станциялари» АЖ"],
                ["08524", "Ўзбекистон Республикаси Тоғ-кон саноати ва геология вазирлиги"],
                ["00002", "шу жумладан:"],
                ["06213", "«Олмалиқ кон-металлургия комбинати» АЖ"],
                ["01164", "«Навоий кон-металлургия комбинати» АЖ"],
                ["99999", "Бошқалар"]
            ]
            mapping_df = pd.DataFrame(mapping_list, columns=["SOOGU", "NAIMUZ"])

            # ==============================
            # 7️⃣ Mapping bilan birlashtiramiz
            # ==============================
            yakuniy = mapping_df.merge(yakuniy, on="SOOGU", how="left").fillna(0)

            # ==============================
            # 8️⃣ Tartibni mapping_list tartibida saqlaymiz
            # ==============================
            yakuniy["order"] = yakuniy.index
            yakuniy = yakuniy.sort_values("order").drop(columns="order")

            # ==============================
            # 9️⃣ Formatlash va natija
            # ==============================
            yakuniy["G1"] = yakuniy["G1"].round(1)/1000
            yakuniy["G2"] = yakuniy["G2"].round(1)/1000

            pd.set_option('display.float_format', '{:,.1f}'.format)
            HISOB09=yakuniy[[ "NAIMUZ", "G1", "G2"]]


            HISOB=pd.merge(HISOB08,HISOB09,on="NAIMUZ")


            df = HISOB.copy()
            df["G1_Δ"] =  df["G1_y"] - df["G1_x"]
            df["G1_%"] = ((df["G1_x"] / df["G1_y"]) * 100  ).replace([float('inf'), -float('inf')], 0).fillna(0)
            df["G2_Δ"] = df["G2_x"] - df["G2_y"]
            df["G2_%"] = ((df["G2_x"] / df["G2_y"]) * 100 ).replace([float('inf'), -float('inf')], 0).fillna(0)
                    # --- 3. Yaxlitlash ---
            # --- 3. Yaxlitlash ---
            for col in ["G1_y", "G1_x", "G1_Δ", "G2_y", "G2_x", "G2_Δ"]:
                df[col] = df[col].round(1)
            for col in ["G1_%", "G2_%"]:
                df[col] = df[col].round(1)

            # --- 4. Ustunlarni tartiblash va nomlash ---
                df_final = df[[
            "NAIMUZ",
            "G1_x", "G1_y", "G1_Δ", "G1_%",
            "G2_x", "G2_y", "G2_Δ", "G2_%"]]
            
            colm = ["G1_x", "G1_y", "G1_Δ", "G1_%",
            "G2_x", "G2_y", "G2_Δ", "G2_%"]
            # df_final = df_final[~(df_final[colm] == 0).all(axis=1)]
            # --- 5. Yakuniy natija ---

            df_final=df_final.set_index("NAIMUZ")
            df_final=df_final.replace("0","-")

            #-------------------------------------------------------------------------------------------------------------------#
            # --------------------------------------------------------------------------------------------------------------------


            #----------------------------------------------------------------------------------------------------------
            #  EXCELGA YOZISHNI BOSHLANISHI 


            # HHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHH


            KATALOH08=katalog08[["SOOGU","OKPO"]]
            BAZA08=baza08[["OKPO","G1","G2","SATR"]]
            BAZA08_m2=BAZA08[BAZA08["SATR"]==210]

            JAMI08_m2=pd.merge(BAZA08_m2,KATALOH08,on="OKPO",how="left")
            JAMI08_m2=JAMI08_m2[["SOOGU","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format

            # --- DataFrame ---
            soogu_df = pd.DataFrame(soogu_data, columns=["SOOGU", "NAIMUZ"])

            # === 1. Asosiy baza (JMAI09) ===
            # JMAI09 = pd.read_excel("JMAI09.xlsx")  # agar fayl bo‘lsa shu yo‘l bilan o‘qisan

            # soogu kodlarini matn sifatida (boshidagi 0 saqlanib) olish
            JAMI08_m2["SOOGU"] = JAMI08_m2["SOOGU"].astype(str).str.zfill(5)


            tartib_df = pd.DataFrame(tartib, columns=["SOOGU", "NAIMUZ"])

            # === 3. JMAI09 da soogu bo‘yicha yig‘ish ===
            agg_m2 = JAMI08_m2.groupby("SOOGU", as_index=False)[["G1", "G2"]].sum()

            # === 4. Tartib bilan birlashtirish ===
            merged_m2 = pd.merge(tartib_df, agg_m2, on="SOOGU", how="left")

            # === 5. “Айрим вазирлик ва идоралар бўйича” uchun sum ===
            # Bu kodlar “asosiy” ro‘yxatda yo‘q, lekin “idora” sifatida kiritiladi
            boshqamin = ["01144", "04043", "04413", "06184", "07254"]
            ayrim_sum_m2 = JAMI08_m2.loc[JAMI08_m2["SOOGU"].isin(boshqamin), ["G1", "G2"]].sum()
            merged_m2.loc[merged_m2["SOOGU"] == "00001", ["G1", "G2"]] = ayrim_sum_m2.values

            # === 6. “Республика бўйича жами” uchun barcha yig‘indi ===
            total_sum_m2 = JAMI08_m2[["G1", "G2"]].sum()
            merged_m2.loc[merged_m2["SOOGU"] == "00000", ["G1", "G2"]] = total_sum_m2.values

            # === 7. “Бошқалар” uchun qolgan kodlar sum ===
            asosiy_kodlar = tartib_df["SOOGU"].tolist() + boshqamin
            boshqalar_sum_m2 = JAMI08_m2.loc[~JAMI08_m2["SOOGU"].isin(asosiy_kodlar), ["G1", "G2"]].sum()
            merged_m2.loc[merged_m2["SOOGU"] == "99999", ["G1", "G2"]] = boshqalar_sum_m2.values

            # === 8. To‘ldirish va yaxlitlash ===
            merged_m2[["G1", "G2"]] = merged_m2[["G1", "G2"]].fillna(0).round(1) / 1000

            # === 9. Yakuniy natija ===
            HISOB08_m2 = merged_m2[["NAIMUZ", "G1", "G2"]]

            #---------------------------------------------------------------------------------------------------------------------
            ##---------------------------------------------------------------------------------------------------------------------






            KATALOH09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]
            BAZA09_m2=BAZA09[BAZA09["SATR"]==210]

            JAMI09_m2=pd.merge(BAZA09_m2,KATALOH09,on="OKPO",how="left")
            JAMI09_m2=JAMI09_m2[["SOOGU","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format

            # ==============================
            # 1️⃣ Bazani tayyorlash
            # ==============================
            # Masalan: JMAI09 = pd.read_excel("JMAI09.xlsx")
            JAMI09_m2["SOOGU"] = JAMI09_m2["SOOGU"].astype(str).str.zfill(5)




            # ==============================
            # 3️⃣ Asosiy kodlar bo‘yicha yig‘indi
            # ==============================
            asosiy_sum_m2 = (
                JAMI09_m2[JAMI09_m2["SOOGU"].isin(asosiy_kodlar)]
                .groupby("SOOGU", as_index=False)[["G1","G2"]]
                .sum()
            )

            # ==============================
            # 4️⃣ Qolganlarini "Бошқалар"ga yig‘ish
            # ==============================
            boshqalar_sum_m2 = (
                JAMI09_m2[~JAMI09_m2["SOOGU"].isin(asosiy_kodlar)][["G1","G2"]]
                .sum()
                .to_frame()
                .T
            )
            boshqalar_sum_m2.insert(0, "SOOGU", "99999")

            # ==============================
            # 5️⃣ Barchasini birlashtirish
            # ==============================
            yakuniy_t = pd.concat([asosiy_sum_m2, boshqalar_sum_m2], ignore_index=True)

            # Respublika bo‘yicha jami
            res_sum_ = yakuniy_t[["G1","G2"]].sum().to_frame().T
            res_sum_.insert(0, "SOOGU", "00000")

            # Айрим вазирлик ва идоралар бўйича (asosiylar yig‘indisi)
            ayrim_sum_m2 = asosiy_sum_m2[["G1","G2"]].sum().to_frame().T
            ayrim_sum_m2.insert(0, "SOOGU", "00001")

            yakuniy_t = pd.concat([res_sum_, ayrim_sum_m2, yakuniy_t], ignore_index=True)


            mapping_df_m2 = pd.DataFrame(mapping_list, columns=["SOOGU", "NAIMUZ"])

            # ==============================
            # 7️⃣ Mapping bilan birlashtiramiz
            # ==============================
            yakuniy_t = mapping_df_m2.merge(yakuniy_t, on="SOOGU", how="left").fillna(0)

            # ==============================
            # 8️⃣ Tartibni mapping_list tartibida saqlaymiz
            # ==============================
            yakuniy_t["order"] = yakuniy_t.index
            yakuniy_t = yakuniy_t.sort_values("order").drop(columns="order")

            # ==============================
            # 9️⃣ Formatlash va natija
            # ==============================
            yakuniy_t["G1"] = yakuniy_t["G1"].round(1)/1000
            yakuniy_t["G2"] = yakuniy_t["G2"].round(1)/1000

            pd.set_option('display.float_format', '{:,.1f}'.format)
            HISOB09_m2=yakuniy_t[[ "NAIMUZ", "G1", "G2"]]


            HISOB_m2=pd.merge(HISOB08_m2,HISOB09_m2,on="NAIMUZ")


            df = HISOB_m2.copy()

            df["G1_Δ"] = df["G1_x"] - df["G1_y"]
            df["G1_%"] = ((df["G1_x"] / df["G1_y"]) * 100  ).replace([float('inf'), -float('inf')], 0).fillna(0)
            df["G2_Δ"] = df["G2_x"] - df["G2_y"]
            df["G2_%"] = ((df["G2_x"] / df["G2_y"]) * 100 ).replace([float('inf'), -float('inf')], 0).fillna(0)

            # --- 3. Yaxlitlash ---
            for col in ["G1_y", "G1_x", "G1_Δ", "G2_y", "G2_x", "G2_Δ"]:
                df[col] = df[col].round(1)
            for col in ["G1_%", "G2_%"]:
                df[col] = df[col].round(1)

            # --- 4. Ustunlarni tartiblash va nomlash ---
            df_final_2 = df[[
            "NAIMUZ",
            "G1_x", "G1_y", "G1_Δ", "G1_%",
            "G2_x", "G2_y", "G2_Δ", "G2_%"]]
        
            colm = ["G1_x", "G1_y", "G1_Δ", "G1_%",
            "G2_x", "G2_y", "G2_Δ", "G2_%"]
            df_final_2 = df_final_2[~(df_final_2[colm] == 0).all(axis=1)]
            # --- 5. Yakuniy natija ---

            df_final_2=df_final_2.set_index("NAIMUZ")
            df_final_2=df_final_2.replace("0","-")



            # yoz_sheetga(df_final, "кредитор-по министр", start_row=9, start_col=2)

            # 4️⃣ Faylni faqat 1 MARTA saqlaymiz
        
            print("✅ кредитор-по министр sheetga yozildi!")




            # Viloyat boyisha debitor===================================================================================================




            # katalog07=katalog08
            # baza07=baza08

            KAT08=katalog08[["ADRES","OKPO"]]
            KAT08["ADRES"]=KAT08["ADRES"].str.split(",").str[0]
            KAT08.columns=KAT08.columns.str.strip().str.upper()
            KAT08["ADRES"] = KAT08["ADRES"].str.replace("`", "'", regex=False).str.lower().str.strip()

            KAT08.columns=KAT08.columns.str.strip().str.upper()




            BAZ08=baza08[["OKPO","SATR","G1","G2"]]
            BAZ08=BAZ08[BAZ08["SATR"]==201]
            pd.set_option('display.float_format', '{:,.0f}'.format)
            #BAZA07.groupby("soato_4")[["g1","g2"]].sum()
            BAZ08.columns=BAZ08.columns.str.strip().str.upper()

            mer=pd.merge(BAZ08,KAT08,on="OKPO",how="left")



            pd.set_option('display.float_format', '{:,.0f}'.format)
            HAM08=mer.groupby("ADRES")[["G1","G2"]].sum()/ 1_000
            pd.options.display.float_format = '{:,.1f}'.format
            HAM08
    #       BU YERDA 08 UN GRUP BY QILINDI TAYYOR


            KAT09=katalog09[["OKPO","ADRES"]]
            BAZ09=baza09[["OKPO","G1","G2","SATR"]]

            BAZ09=BAZ09[BAZ09["SATR"]==201]

            KAT09["ADRES"]=KAT09["ADRES"].str.split(",").str[0]
            KAT09.columns=KAT09.columns.str.strip().str.upper()
            KAT09["ADRES"] = KAT09["ADRES"].str.replace("`", "'", regex=False).str.lower().str.strip()


            JAM=pd.merge(BAZ09,KAT09,on="OKPO",how="left")
            JAM=JAM[["ADRES","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format


            HAM09=JAM.groupby("ADRES")[["G1","G2"]].sum()/1000
            pd.options.display.float_format = '{:,.1f}'.format
            HAM09

            HAMMASI=pd.merge(HAM08,HAM09,on="ADRES")
            
            
            HAMMASI
            # BU YERGACHA HAMMASI  ADRES KESMIDA BIRIKDI 

            pd.set_option('display.float_format', '{:,.1f}'.format)

            # 1️⃣ Farq va foizlarni hisoblash
            HAMMASI = HAMMASI.copy()
            HAMMASI['G1_diff'] = HAMMASI['G1_y'] - HAMMASI['G1_x']
            HAMMASI['G1_pct'] = (HAMMASI['G1_diff'] / HAMMASI['G1_x'] * 100).round(1)+100
            HAMMASI['G2_diff'] = HAMMASI['G2_y'] - HAMMASI['G2_x']
            HAMMASI['G2_pct'] = (HAMMASI['G2_diff'] / HAMMASI['G2_x'] * 100).round(1)+100

            # 2️⃣ Ustunlar tartibi
            HAMMASI = HAMMASI[['G1_x','G1_y','G1_diff','G1_pct','G2_x','G2_y','G2_diff','G2_pct']]

            # 3️⃣ Ўзбекистон Республикаси yig‘indisi
            uz_sum = pd.DataFrame(HAMMASI.sum()).T
            uz_sum.index = ["O'zbekiston Respublikasi"]
            uz_sum['G1_pct'] = (uz_sum['G1_y'] / uz_sum['G1_x'] * 100).round(1)
            uz_sum['G2_pct'] = (uz_sum['G2_y'] / uz_sum['G2_x'] * 100).round(1)

            # 4️⃣ Barchasini birlashtirish
            DEB_VILOYAT = pd.concat([uz_sum, HAMMASI])

            # 5️⃣ Tartib bo‘yicha viloyatlar
            viloyat_tartib = [
                "O'zbekiston Respublikasi",
                "qoraqalpog'iston respublikasi",
                "viloyatlar",
                "andijon viloyati",
                "buxoro viloyati",
                "jizzax viloyati",
                "qashqadaryo viloyati",
                "navoiy viloyati",
                "namangan viloyati",
                "samarqand viloyati",
                "surxondaryo viloyati",
                "sirdaryo viloyati",
                "toshkent viloyati",
                "farg'ona viloyati",
                "xorazm viloyati",
                "toshkent shahri"
            ]

            DEB_VILOYAT = DEB_VILOYAT.reindex(viloyat_tartib)

            # 6️⃣ Ustun nomlari
        
            colm = ['G1_x','G1_y','G1_diff','G1_pct','G2_x','G2_y','G2_diff','G2_pct']
            # DEB_VILOYAT = DEB_VILOYAT[~(DEB_VILOYAT[colm] == 0).all(axis=1)]
            # 7️⃣ Natija
            #HISOB_FINAL.to_excel("hisob0809_krit.xlsx")
            DEB_VILOYAT=DEB_VILOYAT.replace("0","-")

            # yoz_sheetga(HISOB_FINAL,"дебитор-по обл",start_row=8,start_col=2)
            # print("✅ дебитор-по обл sheetga yozildi!")


            # Viloyat boyisha kreditor ===================================================================================================






            KAT08=katalog08[["ADRES","OKPO"]]
            KAT08["ADRES"]=KAT08["ADRES"].str.split(",").str[0]
            KAT08.columns=KAT08.columns.str.strip().str.upper()
            KAT08["ADRES"] = KAT08["ADRES"].str.replace("`", "'", regex=False).str.lower().str.strip()

            KAT08.columns=KAT08.columns.str.strip().str.upper()




            BAZ08_m=baza08[["OKPO","SATR","G1","G2"]]
            BAZ08_m=BAZ08_m[BAZ08_m["SATR"]==210]
            pd.set_option('display.float_format', '{:,.0f}'.format)
            #BAZA07.groupby("soato_4")[["g1","g2"]].sum()
            BAZ08_m.columns=BAZ08_m.columns.str.strip().str.upper()

            mer_m=pd.merge(BAZ08_m,KAT08,on="OKPO",how="left")



            pd.set_option('display.float_format', '{:,.0f}'.format)
            HAM08_m=mer_m.groupby("ADRES")[["G1","G2"]].sum()/ 1_000
            pd.options.display.float_format = '{:,.1f}'.format
            HAM08_m
    #       BU YERDA 08 UN GRUP BY QILINDI TAYYOR


            KAT09=katalog09[["OKPO","ADRES"]]
            BAZ09_m=baza09[["OKPO","G1","G2","SATR"]]

            BAZ09_m=BAZ09_m[BAZ09_m["SATR"]==210]

            KAT09["ADRES"]=KAT09["ADRES"].str.split(",").str[0]
            KAT09.columns=KAT09.columns.str.strip().str.upper()
            KAT09["ADRES"] = KAT09["ADRES"].str.replace("`", "'", regex=False).str.lower().str.strip()


            JAM_m=pd.merge(BAZ09_m,KAT09,on="OKPO",how="left")
            JAM_m=JAM_m[["ADRES","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format


            HAM09_m=JAM_m.groupby("ADRES")[["G1","G2"]].sum()/1000
            pd.options.display.float_format = '{:,.1f}'.format
            HAM09_m

            HAMMASI_m=pd.merge(HAM08_m,HAM09_m,on="ADRES")
            
            
            HAMMASI_m
            # BU YERGACHA HAMMASI  ADRES KESMIDA BIRIKDI 

            pd.set_option('display.float_format', '{:,.1f}'.format)

            # 1️⃣ Farq va foizlarni hisoblash
            HAMMASI_m = HAMMASI_m.copy()
            HAMMASI_m['G1_diff'] = HAMMASI_m['G1_y'] - HAMMASI_m['G1_x']
            HAMMASI_m['G1_pct'] = (HAMMASI_m['G1_diff'] / HAMMASI_m['G1_x'] * 100).round(1)+100
            HAMMASI_m['G2_diff'] = HAMMASI_m['G2_y'] - HAMMASI_m['G2_x']
            HAMMASI_m['G2_pct'] = (HAMMASI_m['G2_diff'] / HAMMASI_m['G2_x'] * 100).round(1)+100

            # 2️⃣ Ustunlar tartibi
            HAMMASI_m = HAMMASI_m[['G1_x','G1_y','G1_diff','G1_pct','G2_x','G2_y','G2_diff','G2_pct']]

            # 3️⃣ Ўзбекистон Республикаси yig‘indisi
            uz_sum_m = pd.DataFrame(HAMMASI_m.sum()).T
            uz_sum_m.index = ["O'zbekiston Respublikasi"]
            uz_sum_m['G1_pct'] = (uz_sum_m['G1_y'] / uz_sum_m['G1_x'] * 100).round(1)
            uz_sum_m['G2_pct'] = (uz_sum_m['G2_y'] / uz_sum_m['G2_x'] * 100).round(1)

            # 4️⃣ Barchasini birlashtirish
            KIR_VILOYAT = pd.concat([uz_sum_m, HAMMASI_m])

            # 5️⃣ Tartib bo‘yicha viloyatlar
            viloyat_tartib = [
                "O'zbekiston Respublikasi",
                "qoraqalpog'iston respublikasi",
                "viloyatlar",
                "andijon viloyati",
                "buxoro viloyati",
                "jizzax viloyati",
                "qashqadaryo viloyati",
                "navoiy viloyati",
                "namangan viloyati",
                "samarqand viloyati",
                "surxondaryo viloyati",
                "sirdaryo viloyati",
                "toshkent viloyati",
                "farg'ona viloyati",
                "xorazm viloyati",
                "toshkent shahri"
            ]

            KIR_VILOYAT = KIR_VILOYAT.reindex(viloyat_tartib)

            # 6️⃣ Ustun nomlari
        
            colm = ['G1_x','G1_y','G1_diff','G1_pct','G2_x','G2_y','G2_diff','G2_pct']
            # KIR_VILOYAT = KIR_VILOYAT[~(KIR_VILOYAT[colm] == 0).all(axis=1)]
            # 7️⃣ Natija
            #HISOB_FINAL.to_excel("hisob0809_krit.xlsx")
            KIR_VILOYAT=KIR_VILOYAT.replace("0","-")

            # yoz_sheetga(HISOB_FINAL,"дебитор-по обл",start_row=8,start_col=2)
            # print("✅ дебитор-по обл sheetga yozildi!")

            # Zarar viloyat boyicha ============================================================================================




            KATALOG09=katalog09[["OKPO","ADRES"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.replace("`", "'", regex=False).str.lower().str.strip()


            #-----------------------------------------------------------------------------------------------------------
            import pandas as pd

            # 1️⃣ 102 va 103 satrlarni ajratamiz
            BAZA09_102 = BAZA09[BAZA09["SATR"] == 102]
            BAZA09_103 = BAZA09[BAZA09["SATR"] == 103]

            # 2️⃣ Har ikkalasini KATALOH bilan bog‘laymiz
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.split(",").str[0]

            JAMI_102 = pd.merge(BAZA09_102, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_103 = pd.merge(BAZA09_103, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]



            JAMI102 = (
                JAMI_102.groupby("ADRES", as_index=False)
                    .agg(
                        G2_SUM_1000_102=("G2", lambda x: x.sum() / 1000),
                        G2_COUNT_102=("G2", lambda x: (x > 0).sum())
                    )
            )



            JAMI103 = (
                JAMI_103.groupby("ADRES", as_index=False)
                    .agg(
                        G2_SUM_1000_103=("G2", lambda x: x.sum() / 1000),
                        G2_COUNT_103=("G2", lambda x: (x > 0).sum())
                    )
            )


            # 4️⃣ Endi ikkisini ADRES bo‘yicha birlashtiramiz
            JAMI_MERGED = pd.merge(JAMI103, JAMI102, on="ADRES", how="outer")

            # 5️⃣ Ko‘rinishni chiroyli qilish
            pd.options.display.float_format = '{:,.1f}'.format

            # === Sizdagi asosiy jadval ===
            # JAMI_MERGED mavjud deb olinadi

            # 1️⃣ Viloyatlar tartibi
            viloyat_tartib = [
                "Ўзбекистон Республикаси",
                "qoraqalpog'iston respublikasi",
                "viloyatlar",
                "andijon viloyati",
                "buxoro viloyati",
                "jizzax viloyati",
                "qashqadaryo viloyati",
                "navoiy viloyati",
                "namangan viloyati",
                "samarqand viloyati",
                "surxondaryo viloyati",
                "sirdaryo viloyati",
                "toshkent viloyati",
                "farg'ona viloyati",
                "xorazm viloyati",
                "toshkent shahri"
            ]

            # 2️⃣ Farq va foiz ustunlarini hisoblaymiz
            JAMI_MERGED["COUNT_DIFF"] = JAMI_MERGED["G2_COUNT_102"] - JAMI_MERGED["G2_COUNT_103"]
            JAMI_MERGED["SUM_DIFF_1000"] = JAMI_MERGED["G2_SUM_1000_102"] - JAMI_MERGED["G2_SUM_1000_103"]

            # Nolga bo‘lishni oldini olish
            JAMI_MERGED["SUM_RATIO_PERCENT"] = (
                (JAMI_MERGED["G2_SUM_1000_102"] / JAMI_MERGED["G2_SUM_1000_103"]) * 100
            ).replace([float('inf'), -float('inf')], pd.NA)

            # 3️⃣ Umumiy yig‘indi satr ("Ўзбекистон Республикаси")
            sum_row = pd.DataFrame({
                "ADRES": ["Ўзбекистон Республикаси"],
                "G2_COUNT_103": [JAMI_MERGED["G2_COUNT_103"].sum()],
                "G2_COUNT_102": [JAMI_MERGED["G2_COUNT_102"].sum()],
                "COUNT_DIFF": [JAMI_MERGED["COUNT_DIFF"].sum()],
                "G2_SUM_1000_103": [JAMI_MERGED["G2_SUM_1000_103"].sum()],
                "G2_SUM_1000_102": [JAMI_MERGED["G2_SUM_1000_102"].sum()],
                "SUM_DIFF_1000": [JAMI_MERGED["SUM_DIFF_1000"].sum()],
                "SUM_RATIO_PERCENT": [
                    (JAMI_MERGED["G2_SUM_1000_102"].sum() / JAMI_MERGED["G2_SUM_1000_103"].sum()) * 100
                ]
            })

            # 4️⃣ Barchasini birlashtiramiz
            JAMI_ub = pd.concat([sum_row, JAMI_MERGED], ignore_index=True)

            colm = ['G2_COUNT_103','G2_COUNT_102','COUNT_DIFF','G2_SUM_1000_103','G2_SUM_1000_102'
                    ,'SUM_DIFF_1000','SUM_RATIO_PERCENT']
            # JAMI_ub = JAMI_ub[~(JAMI_ub[colm] == 0).all(axis=1)]

            JAMI_ub=JAMI_ub.set_index("ADRES")
            JAMI_ub=JAMI_ub.reindex(viloyat_tartib).replace("0","-")
            JAMI_ub=JAMI_ub.replace("0","-")



            # yoz_sheetga(JAMI_ub,"обл убытка",start_row=5,start_col=2)
            print("✅ обл убытка  sheetga yozildi!")

            #  Zarar vazirlik boyicha ==========================================================================================




            KATALOH09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]

            BAZA09=BAZA09[BAZA09["SATR"]==102]

            JAMI09=pd.merge(BAZA09,KATALOH09,on="OKPO",how="left")
            JAMI09=JAMI09[["SOOGU","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format

            # 1️⃣ Ma'lumotlarni tayyorlash
            KATALOH09 = katalog09[["SOOGU", "OKPO"]].copy()
            BAZA09 = baza09[["OKPO", "G1", "G2", "SATR"]].copy()

            # 2️⃣ Kodlar formatini to‘g‘rilash
            KATALOH09["SOOGU"] = KATALOH09["SOOGU"].astype(str).str.zfill(5)


            # 4️⃣ Hisoblash funksiyasi
            def hisobla(df, satr):
                B = df[df["SATR"] == satr]
                J = pd.merge(B, KATALOH09, on="OKPO", how="left")[["SOOGU", "G2"]]

                asosiy = (
                    J[J["SOOGU"].isin(asosiy_kodlar)]
                    .groupby("SOOGU", as_index=False)
                    .agg(
                        SUM=("G2", lambda x: x.sum() / 1000),
                        COUNT=("G2", lambda x: (x > 0).sum())
                    )
                )

                boshqalar = J[~J["SOOGU"].isin(asosiy_kodlar)]
                boshqalar_agg = pd.DataFrame({
                    "SOOGU": ["99999"],
                    "SUM": [boshqalar["G2"].sum() / 1000],
                    "COUNT": [(boshqalar["G2"] > 0).sum()]
                })

                jami = pd.concat([asosiy, boshqalar_agg], ignore_index=True)

                # Айрим вазирлик ва идоралар бўйича (asosiy yig‘indisi)
                ayrim_sum_ = pd.DataFrame({
                    "SOOGU": ["00001"],
                    "SUM": [asosiy["SUM"].sum()],
                    "COUNT": [asosiy["COUNT"].sum()]
                })

                # Ўзбекистон Республикаси бўйича жами
                total = pd.DataFrame({
                    "SOOGU": ["00000"],
                    "SUM": [jami["SUM"].sum()],
                    "COUNT": [jami["COUNT"].sum()]
                })

                jami = pd.concat([total, ayrim_sum_, jami], ignore_index=True)
                return jami

            # 5️⃣ 102 va 103 uchun hisoblash
            res102 = hisobla(BAZA09, 102).rename(columns={"SUM": "G2_SUM_1000_102", "COUNT": "G2_COUNT_102"})
            res103 = hisobla(BAZA09, 103).rename(columns={"SUM": "G2_SUM_1000_103", "COUNT": "G2_COUNT_103"})

            # 6️⃣ Birlashtirish
            yakuniy_v = pd.merge(res102, res103, on="SOOGU", how="outer").fillna(0)

            # 7️⃣ Qo‘shimcha ustunlar (farq va foiz)
            yakuniy_v["COUNT_DIFF"] = yakuniy_v["G2_COUNT_102"].astype(int) - yakuniy_v["G2_COUNT_103"].astype(int)
            yakuniy_v["SUM_DIFF"] = yakuniy_v["G2_SUM_1000_102"] - yakuniy_v["G2_SUM_1000_103"]
            yakuniy_v["SUM_PCT"] = yakuniy_v.apply(
                lambda x: (x["G2_SUM_1000_102"] / x["G2_SUM_1000_103"] * 100) if x["G2_SUM_1000_103"] != 0 else 0,
                axis=1
            ).round(1)

            # 8️⃣ Tartib va nomlar
            tartib = [
                ["00000", "Ўзбекистон Республикаси"],
                ["00001", "Айрим вазирлик ва идоралар бўйича"],
                ["04403", "Ўзбекистон Республикаси Қурилиш ва уй-жой коммунал хўжалиги вазирлиги"],
                ["01354", "«Ўзавтосаноат» АЖ"],
                ["08114", "«Ўздонмаҳсулот» АК"],
                ["08654", "«Ўзкимёсаноат» АЖ"],
                ["06264", "«Ўзагротехсаноатхолдинг» АЖ"],
                ["03504", "«Ўзбекистон темир йўллари» АЖ"],
                ["01024", "«Ўзбекнефтгаз» АЖ"],
                ["01124", "«Ҳудудгазтаъминот» АЖ"],
                ["01104", "«Ўзтрансгаз» АЖ"],
                ["06224", "«Ўзпахтасаноат» АЖ"],
                ["01014", "«Ўзбекистон миллий электр тармоқлари» АЖ"],
                ["01074", "«Ҳудудий электр тармоқлари» АЖ"],
                ["01094", "«Иссиқлик электр станциялари» АЖ"],
                ["08524", "Ўзбекистон Республикаси Тоғ-кон саноати ва геология вазирлиги"],
                # ["00002", "шу жумладан:"],
                
                ["99999", "Бошқалар"]
            ]

            tartib_df = pd.DataFrame(tartib, columns=["SOOGU", "НАИМ"])
            yakuniy_v = pd.merge(tartib_df, yakuniy_v, on="SOOGU", how="left").fillna(0)

            # 9️⃣ Ustunlarni nomlash
            yakuniy_v = yakuniy_v[
                ["НАИМ",
                "G2_COUNT_103","G2_COUNT_102","COUNT_DIFF",
                "G2_SUM_1000_103","G2_SUM_1000_102","SUM_DIFF","SUM_PCT"]
            ]

            # 10️⃣ Formatlash
            yakuniy_v = yakuniy_v.round(1)
            pd.set_option('display.float_format', '{:,.1f}'.format)

            colm = ["G2_COUNT_103","G2_COUNT_102","COUNT_DIFF",
                "G2_SUM_1000_103","G2_SUM_1000_102","SUM_DIFF","SUM_PCT"]
            # yakuniy_v = yakuniy_v[~(yakuniy_v[colm] == 0).all(axis=1)]

            yakuniy_v=yakuniy_v.set_index("НАИМ")
            yakuniy_v=yakuniy_v.replace("0","-")

            # yoz_sheetga(yakuniy_v,"минс убытка",start_row=6,start_col=2)
            print("✅ минс убытка sheetga yozildi!")

            # # TAB 1 ===============================================================================================================


            KATALOG09=katalog09[["OKPO","ADRES"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]

            #BAZA09=BAZA09[BAZA09["SATR"]==210]

            KATALOG09["ADRES"]=KATALOG09["ADRES"].str.split(",").str[0]
            KATALOG09.columns=KATALOG09.columns.str.strip().str.upper()
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.replace("`", "'", regex=False).str.lower().str.strip()


            JAMI=pd.merge(BAZA09,KATALOG09,on="OKPO",how="left")
            JAMI=JAMI[["ADRES","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format
            JAMI







            BAZA09_201=BAZA09[BAZA09["SATR"]==201]
            BAZA09_202 = BAZA09[BAZA09["SATR"] == 202]
            BAZA09_203 = BAZA09[BAZA09["SATR"] == 203]
            BAZA09_204 = BAZA09[BAZA09["SATR"] == 204]
            BAZA09_205 = BAZA09[BAZA09["SATR"] == 205]

            # 2️⃣ Har ikkalasini KATALOH bilan bog‘laymiz
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.split(",").str[0]
            JAMI_201 = pd.merge(BAZA09_201, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_202 = pd.merge(BAZA09_202, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_203 = pd.merge(BAZA09_203, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_204 = pd.merge(BAZA09_204, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_205 =pd.merge(BAZA09_205, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]

            hisob_201=JAMI_201.groupby("ADRES")["G1"].sum().reset_index()
            hisob_202=JAMI_202.groupby("ADRES")["G1"].sum().reset_index()
            hisob_203=JAMI_203.groupby("ADRES")["G1"].sum().reset_index()
            hisob_204=JAMI_204.groupby("ADRES")["G1"].sum().reset_index()
            hisob_205=JAMI_205.groupby("ADRES")["G1"].sum().reset_index()




            # Misol uchun: 5 ta df bor
            dfs = [hisob_201,hisob_202,hisob_203,hisob_204,hisob_205]

            # Har bir df’da G1 ustun nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)

            # Hammasini 'ADRES' bo‘yicha birlashtirish
            yakuniy1 = reduce(lambda left, right: pd.merge(left, right, on="ADRES", how="outer"), dfs)

            # Natija
            yakuniy1["G_6"]=yakuniy1["G1_1"]-yakuniy1["G1_5"]
            yakuniy1

            yakuniy1=yakuniy1.set_index("ADRES")


            yakuniy1= yakuniy1.reindex(viloyat_tartib)
            yakuniy1

            # 1️⃣ Hammasini yig‘amiz (NaN'larni hisobdan chiqarib)
            summalar = yakuniy1.sum(numeric_only=True)

            # 2️⃣ "Ўзбекистон Республикаси" qatori uchun yig‘indini qo‘yamiz
            yakuniy1.loc["Ўзбекистон Республикаси"] = summalar
            yakuniy1/1000
            pd.options.display.float_format = '{:,.1f}'.format
            YAKUN1=yakuniy1/1000
            colm = ["G1_1","G1_2","G1_3","G1_4","G1_5","G_6"]
            # YAKUN1 = YAKUN1[~(YAKUN1[colm] == 0).all(axis=1)]

            YAKUN1=YAKUN1.replace("0","-")

            
            print("✅ таб 1 sheetga yozildi!")


            # TAB 3 ================================================================================================================




            KATALOG09=katalog09[["OKPO","ADRES"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]

            #BAZA09=BAZA09[BAZA09["SATR"]==210]

            KATALOG09["ADRES"]=KATALOG09["ADRES"].str.split(",").str[0]
            KATALOG09.columns=KATALOG09.columns.str.strip().str.upper()
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.replace("`", "'", regex=False).str.lower().str.strip()


            JAMI=pd.merge(BAZA09,KATALOG09,on="OKPO",how="left")
            JAMI=JAMI[["ADRES","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format
            JAMI





            BAZA09_210 = BAZA09[BAZA09["SATR"] == 210]
            BAZA09_211 = BAZA09[BAZA09["SATR"] == 211]
            BAZA09_212 = BAZA09[BAZA09["SATR"] == 212]
            BAZA09_213 = BAZA09[BAZA09["SATR"] == 213]
            BAZA09_214 = BAZA09[BAZA09["SATR"] == 214]
            BAZA09_215 = BAZA09[BAZA09["SATR"] == 215]
            BAZA09_216 = BAZA09[BAZA09["SATR"] == 216]
            BAZA09_218 = BAZA09[BAZA09["SATR"] == 218]


            # 2️⃣ Har ikkalasini KATALOH bilan bog‘laymiz
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.split(",").str[0]
            JAMI_210 = pd.merge(BAZA09_210, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_211 = pd.merge(BAZA09_211, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_212 = pd.merge(BAZA09_212, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_213 = pd.merge(BAZA09_213, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_214 = pd.merge(BAZA09_214, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_215 =pd.merge(BAZA09_215, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_216 = pd.merge(BAZA09_216, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_218 = pd.merge(BAZA09_218, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]

            hisob_210=JAMI_210.groupby("ADRES")["G1"].sum().reset_index()
            hisob_211=JAMI_211.groupby("ADRES")["G1"].sum().reset_index()
            hisob_212=JAMI_212.groupby("ADRES")["G1"].sum().reset_index()
            hisob_213=JAMI_213.groupby("ADRES")["G1"].sum().reset_index()
            hisob_214=JAMI_214.groupby("ADRES")["G1"].sum().reset_index()
            hisob_215=JAMI_215.groupby("ADRES")["G1"].sum().reset_index()
            hisob_216=JAMI_216.groupby("ADRES")["G1"].sum().reset_index()
            hisob_218=JAMI_218.groupby("ADRES")["G1"].sum().reset_index()



            # Misol uchun: 5 ta df bor
            dfs = [hisob_210,hisob_211,hisob_212,hisob_213,hisob_214,hisob_215,hisob_216,hisob_218]

            # Har bir df’da G1 ustun nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)

            # Hammasini 'ADRES' bo‘yicha birlashtirish
            yakuniy3 = reduce(lambda left, right: pd.merge(left, right, on="ADRES", how="outer"), dfs)

            yakuniy3
            # Natija
            yakuniy3["G_9"]=yakuniy3["G1_1"]-yakuniy3["G1_7"]
            yakuniy3

            yakuniy3=yakuniy3.set_index("ADRES")


            yakuniy3 = yakuniy3.reindex(viloyat_tartib)
            yakuniy3

            # 1️⃣ Hammasini yig‘amiz (NaN'larni hisobdan chiqarib)
            summalar = yakuniy3.sum(numeric_only=True)

            # 2️⃣ "Ўзбекистон Республикаси" qatori uchun yig‘indini qo‘yamiz
            yakuniy3.loc["Ўзбекистон Республикаси"] = summalar
            yakuniy3/1000
            pd.options.display.float_format = '{:,.1f}'.format
            YAKUN3=yakuniy3/1000

            colm = ["G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9"]
            # YAKUN3 = YAKUN3[~(YAKUN3[colm] == 0).all(axis=1)]

            YAKUN3=YAKUN3.replace("0","-")
            # yoz_sheetga(YAKUN,"таб 3",start_row=7,start_col=2)
            print("✅ таб 3 sheetga yozildi!")

            #  TAB 5 ============================================================================================================






            KATALOH09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","SATR"]]
            KATALOH09.loc[:,"SOOGU"] = KATALOH09["SOOGU"].astype(str).str.zfill(5)

            BAZA_201=BAZA09[BAZA09["SATR"]==201]
            BAZA_202=BAZA09[BAZA09["SATR"]==202]
            BAZA_203=BAZA09[BAZA09["SATR"]==203]
            BAZA_204=BAZA09[BAZA09["SATR"]==204]
            BAZA_205=BAZA09[BAZA09["SATR"]==205]

            JAMI201=pd.merge(BAZA_201,KATALOH09,on="OKPO",how="left")
            JAMI202=pd.merge(BAZA_202,KATALOH09,on="OKPO",how="left")
            JAMI203=pd.merge(BAZA_203,KATALOH09,on="OKPO",how="left")
            JAMI204=pd.merge(BAZA_204,KATALOH09,on="OKPO",how="left")
            JAMI205=pd.merge(BAZA_205,KATALOH09,on="OKPO",how="left")

            pd.options.display.float_format = '{:,.1f}'.format
            JAMI_201=JAMI201[["SOOGU","G1"]]
            JAMI_202=JAMI202[["SOOGU","G1"]]
            JAMI_203=JAMI203[["SOOGU","G1"]]
            JAMI_204=JAMI204[["SOOGU","G1"]]
            JAMI_205=JAMI205[["SOOGU","G1"]]


            import pandas as pd

            # === SOOGU nomlari ro‘yxati (faqat 1 marta yoziladi) ===
            tartib = [
                ["00000", "Республика бўйича жами"],
                ["00001", "Айрим вазирлик ва идоралар бўйича"],
                ["04403", "Ўзбекистон Республикаси Қурилиш ва уй-жой коммунал хўжалиги вазирлиги"],
                ["01354", "«Ўзавтосаноат» АЖ"],
                ["08114", "«Ўздонмаҳсулот» АК"],
                ["08654", "«Ўзкимёсаноат» АЖ"],
                ["06264", "«Ўзагротехсаноатхолдинг» АЖ"],
                ["03504", "«Ўзбекистон темир йўллари» АЖ"],
                ["01024", "«Ўзбекнефтгаз» АЖ"],
                ["01124", "«Ҳудудгазтаъминот» АЖ"],
                ["01104", "«Ўзтрансгаз» АЖ"],
                ["06224", "«Ўзпахтасаноат» АЖ"],
                ["01014", "«Ўзбекистон миллий электр тармоқлари» АЖ"],
                ["01074", "«Ҳудудий электр тармоқлари» АЖ"],
                ["01094", "«Иссиқлик электр станциялари» АЖ"],
                ["08524", "Ўзбекистон Республикаси Тоғ-кон саноати ва геология вазирлиги"],
                ["00002", "шу жумладан:"],
                ["06213", "«Олмалиқ кон-металлургия комбинати» АЖ"],
                ["01164", "«Навоий кон-металлургия комбинати» АЖ"],
                ["99999", "Бошқалар"]
            ]

            tartib_df = pd.DataFrame(tartib, columns=["SOOGU", "NAIMUZ"])

            # Idoralar ro‘yxati (har safar ishlatiladi)
            ayrm = ["04403","01354","08114","08654","06264","03504","01024","01124","01104","06224","01014","01074","01094","08524","06213","01164"
            ]
            boshqamin=["01144", "04043", "04413", "06184", "07254"]
            # === FUNKSIYA: Har bir SATR (201, 202, …) bo‘yicha hisoblash ===
            def hisobla(df):
                df = df.copy()
                df["SOOGU"] = df["SOOGU"].astype(str).str.zfill(5)

                # 1. SOOGU bo‘yicha yig‘ish
                agg = df.groupby("SOOGU", as_index=False)["G1"].sum()

                # 2. Tartib bilan birlashtirish
                merged = pd.merge(tartib_df, agg, on="SOOGU", how="left")

                # 3. “Айрим вазирлик ва идоралар бўйича”
                ayrim_sum = df.loc[df["SOOGU"].isin(ayrm), "G1"].sum()
                merged.loc[merged["SOOGU"] == "00001", "G1"] = ayrim_sum

                # 4. “Республика бўйича жами”
                total_sum = df["G1"].sum()
                merged.loc[merged["SOOGU"] == "00000", "G1"] = total_sum

                # 5. “Бошқалар”
                asosiy_kodlar = tartib_df["SOOGU"].tolist() + boshqamin
                boshqalar_sum = df.loc[~df["SOOGU"].isin(asosiy_kodlar), "G1"].sum()
                merged.loc[merged["SOOGU"] == "99999", "G1"] = boshqalar_sum

                # 6. To‘ldirish va 1000 ga bo‘lish
                merged["G1"] = merged["G1"].fillna(0).round(1) / 1000

                # 7. Yakuniy natija
                result = merged[["NAIMUZ", "G1"]].set_index("NAIMUZ")
                return result
            HISOB_201 = hisobla(JAMI_201)
            HISOB_202 = hisobla(JAMI_202)
            HISOB_203 = hisobla(JAMI_203)
            HISOB_204 = hisobla(JAMI_204)
            HISOB_205 = hisobla(JAMI_205)

            from functools import reduce

            dfs = [HISOB_201, HISOB_202, HISOB_203, HISOB_204, HISOB_205]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)
                if "NAIMUZ" in df.columns:
                    df.set_index("NAIMUZ", inplace=True)

            # Join orqali birlashtirish
            yakuniy = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash
            yakuniy = yakuniy.reindex(HISOB_201.index)

            # Yangi ustun qo‘shish
            yakuniy["G_6"] = yakuniy["G1_1"] - yakuniy["G1_5"]

            yakuniy

            colm = ["G1_1","G1_2","G1_3","G1_4","G1_5","G_6"]
            # yakuniy1 = yakuniy1[~(yakuniy1[colm] == 0).all(axis=1)]


            yakuniy1=yakuniy.replace("0","-")


            print("✅ таб 5 sheetga yozildi!")


            # TAB 7 ==============================================================================================================





            KATALOH09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","SATR"]]
            KATALOH09["SOOGU"] = KATALOH09["SOOGU"].astype(str).str.zfill(5)

            BAZA_210=BAZA09[BAZA09["SATR"]==210]
            BAZA_211=BAZA09[BAZA09["SATR"]==211]
            BAZA_212=BAZA09[BAZA09["SATR"]==212]
            BAZA_213=BAZA09[BAZA09["SATR"]==213]
            BAZA_214=BAZA09[BAZA09["SATR"]==214]
            BAZA_215=BAZA09[BAZA09["SATR"]==215]
            BAZA_216=BAZA09[BAZA09["SATR"]==216]
            BAZA_218=BAZA09[BAZA09["SATR"]==218]

            JAMI210=pd.merge(BAZA_210,KATALOH09,on="OKPO",how="left")
            JAMI211=pd.merge(BAZA_211,KATALOH09,on="OKPO",how="left")
            JAMI212=pd.merge(BAZA_212,KATALOH09,on="OKPO",how="left")
            JAMI213=pd.merge(BAZA_213,KATALOH09,on="OKPO",how="left")
            JAMI214=pd.merge(BAZA_214,KATALOH09,on="OKPO",how="left")
            JAMI215=pd.merge(BAZA_215,KATALOH09,on="OKPO",how="left")
            JAMI216=pd.merge(BAZA_216,KATALOH09,on="OKPO",how="left")
            JAMI218=pd.merge(BAZA_218,KATALOH09,on="OKPO",how="left")

            pd.options.display.float_format = '{:,.1f}'.format

            JAMI_210=JAMI210[["SOOGU","G1"]]
            JAMI_211=JAMI211[["SOOGU","G1"]]
            JAMI_212=JAMI212[["SOOGU","G1"]]
            JAMI_213=JAMI213[["SOOGU","G1"]]
            JAMI_214=JAMI214[["SOOGU","G1"]]
            JAMI_215=JAMI215[["SOOGU","G1"]]
            JAMI_216=JAMI216[["SOOGU","G1"]]
            JAMI_218=JAMI218[["SOOGU","G1"]]

            #------------------------------------------------------------------------------------------------------------------


            tartib = [
                ["00000", "Республика бўйича жами"],
                ["00001", "Айрим вазирлик ва идоралар бўйича"],
                ["04403", "Ўзбекистон Республикаси Қурилиш ва уй-жой коммунал хўжалиги вазирлиги"],
                ["01354", "«Ўзавтосаноат» АЖ"],
                ["08114", "«Ўздонмаҳсулот» АК"],
                ["08654", "«Ўзкимёсаноат» АЖ"],
                ["06264", "«Ўзагротехсаноатхолдинг» АЖ"],
                ["03504", "«Ўзбекистон темир йўллари» АЖ"],
                ["01024", "«Ўзбекнефтгаз» АЖ"],
                ["01124", "«Ҳудудгазтаъминот» АЖ"],
                ["01104", "«Ўзтрансгаз» АЖ"],
                ["06224", "«Ўзпахтасаноат» АЖ"],
                ["01014", "«Ўзбекистон миллий электр тармоқлари» АЖ"],
                ["01074", "«Ҳудудий электр тармоқлари» АЖ"],
                ["01094", "«Иссиқлик электр станциялари» АЖ"],
                ["08524", "Ўзбекистон Республикаси Тоғ-кон саноати ва геология вазирлиги"],
                ["00002", "шу жумладан:"],
                ["06213", "«Олмалиқ кон-металлургия комбинати» АЖ"],
                ["01164", "«Навоий кон-металлургия комбинати» АЖ"],
                ["99999", "Бошқалар"]
            ]

            tartib_df = pd.DataFrame(tartib, columns=["SOOGU", "NAIMUZ"])

            # Idoralar ro‘yxati (har safar ishlatiladi)
            ayrm = ["04403","01354","08114","08654","06264","03504","01024","01124","01104","06224","01014","01074","01094","08524","06213","01164"
            ]
            boshqamin=["01144", "04043", "04413", "06184", "07254"]
            #----------------------------------------------------------------
            def hisobla(df):
                df = df.copy()
                df.loc[:,"SOOGU"] = df["SOOGU"].astype(str).str.zfill(5)

                # 1. SOOGU bo‘yicha yig‘ish
                agg = df.groupby("SOOGU", as_index=False)["G1"].sum()

                # 2. Tartib bilan birlashtirish
                merged = pd.merge(tartib_df, agg, on="SOOGU", how="left")

                # 3. “Айрим вазирлик ва идоралар бўйича”
                ayrim_sum = df.loc[df["SOOGU"].isin(ayrm), "G1"].sum()
                merged.loc[merged["SOOGU"] == "00001", "G1"] = ayrim_sum

                # 4. “Республика бўйича жами”
                total_sum = df["G1"].sum()
                merged.loc[merged["SOOGU"] == "00000", "G1"] = total_sum

                # 5. “Бошқалар”
                asosiy_kodlar = tartib_df["SOOGU"].tolist() + boshqamin
                boshqalar_sum = df.loc[~df["SOOGU"].isin(asosiy_kodlar), "G1"].sum()
                merged.loc[merged["SOOGU"] == "99999", "G1"] = boshqalar_sum

                # 6. To‘ldirish va 1000 ga bo‘lish
                merged["G1"] = merged["G1"].fillna(0).round(1) / 1000

                # 7. Yakuniy natija
                result = merged[["NAIMUZ", "G1"]].set_index("NAIMUZ")
                return result

            HISOB_210=hisobla(JAMI_210)
            HISOB_211=hisobla(JAMI_211)
            HISOB_212=hisobla(JAMI_212)
            HISOB_213=hisobla(JAMI_213)
            HISOB_214=hisobla(JAMI_214)
            HISOB_215=hisobla(JAMI_215)
            HISOB_216=hisobla(JAMI_216)
            HISOB_218=hisobla(JAMI_218)


            #-------------------------------------------------------------------------------------------------------------
            from functools import reduce

            dfs = [HISOB_210,HISOB_211,HISOB_212,HISOB_213,HISOB_214,HISOB_215,HISOB_216,HISOB_218]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)
                if "NAIMUZ" in df.columns:
                    df.set_index("NAIMUZ", inplace=True)

            # Join orqali birlashtirish
            yakuniy = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash
            yakuniy = yakuniy.reindex(HISOB_210.index)

            # Yangi ustun qo‘shish
            yakuniy["G_9"] = yakuniy["G1_1"] - yakuniy["G1_7"]

            yakuniy

            colm = ["G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9"]
            # yakuniy2 = yakuniy[~(yakuniy[colm] == 0).all(axis=1)]
            yakuniy2=yakuniy.replace("0","-")

    



            # fin VILOYAT BOYICHA ==================================================================================================



            KATALOG09=katalog09[["OKPO","ADRES"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.replace("`", "'", regex=False).str.capitalize().str.strip()


            #-----------------------------------------------------------------------------------------------------------


            # 1️⃣ 102 va 103 satrlarni ajratamiz
            BAZA09_102 = BAZA09[BAZA09["SATR"] == 102]
            BAZA09_103 = BAZA09[BAZA09["SATR"] == 103]
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.split(",").str[0]
            # 2️⃣ Har ikkalasini KATALOH bilan bog‘laymiz

            #--------------------------------------------------------------------------------------------------------
            JAMI_102 = pd.merge(BAZA09_102, KATALOG09, on="OKPO", how="left")[["ADRES", "G1","G2"]]
            JAMI_103 = pd.merge(BAZA09_103, KATALOG09, on="OKPO", how="left")[["ADRES", "G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format

            jami_102=JAMI_102.groupby("ADRES")[["G1","G2"]].sum()
            jami_102["102"]=jami_102["G1"]-jami_102["G2"]
            hisob_102=jami_102["102"].reset_index()
            hisob_102

            jami_103=JAMI_103.groupby("ADRES")[["G1","G2"]].sum()
            jami_103["102"]=jami_103["G1"]-jami_103["G2"]
            hisob_103=jami_103["102"].reset_index()
            hisob_103
            H=pd.merge(hisob_103,hisob_102,on="ADRES")
            H["102_x"]=H["102_x"]/1000
            H["102_y"]=H["102_y"]/1000
            H["foiz"]=H["102_y"]/H["102_x"]*100
            H

            JAMI_102 = pd.merge(BAZA09_102, KATALOG09, on="OKPO", how="left")[["ADRES", "G2","G1"]]
            JAMI_103 = pd.merge(BAZA09_103, KATALOG09, on="OKPO", how="left")[["ADRES", "G2","G1"]]



            JAMI102_Z = (
                JAMI_102.groupby("ADRES", as_index=False)
                    .agg(
                        G2_COUNT_102=("G2", lambda x: (x > 0).sum()),
                        G2_SUM_1000_102=("G2", lambda x: x.sum() / 1000)
                    )
            )
            JAMI102_F = (
                JAMI_102.groupby("ADRES", as_index=False)
                    .agg(
                        G2_COUNT_102=("G1", lambda x: (x > 0).sum()),
                        G2_SUM_1000_102=("G1", lambda x: x.sum() / 1000)
                    )
            )





            JAMI103_Z = (
                JAMI_103.groupby("ADRES", as_index=False)
                    .agg(
                        G2_COUNT_103=("G2", lambda x: (x > 0).sum()),
                        G2_SUM_1000_103=("G2", lambda x: x.sum() / 1000)
                    )
            )
            JAMI103_F = (
                JAMI_103.groupby("ADRES", as_index=False)
                    .agg(
                        G2_COUNT_103=("G1", lambda x: (x > 0).sum()),
                        G2_SUM_1000_103=("G1", lambda x: x.sum() / 1000)
                    )
            )


            dfs = [JAMI103_F,JAMI103_Z,JAMI102_F,JAMI102_Z]

            # Har bir df’da G1 ustun nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)

            # Hammasini 'ADRES' bo‘yicha birlashtirish
            yakuniy = reduce(lambda left, right: pd.merge(left, right, on="ADRES", how="outer"), dfs)

            YAKUN=pd.merge(H,yakuniy,on="ADRES")
            YAKUN

            sum_row = pd.DataFrame({
                "ADRES": ["O'zbekiston Respublikasi"],
                "102_x": [YAKUN["102_x"].sum()],
                "102_y": [YAKUN["102_y"].sum()],
                "foiz": [(YAKUN["102_y"].sum() / YAKUN["102_x"].sum()) * 100],
                "G2_COUNT_103_x": [YAKUN["G2_COUNT_103_x"].sum()],
                "G2_SUM_1000_103_x": [YAKUN["G2_SUM_1000_103_x"].sum()],
                "G2_COUNT_103_y": [YAKUN["G2_COUNT_103_y"].sum()],
                "G2_SUM_1000_103_y": [YAKUN["G2_SUM_1000_103_y"].sum()],
                "G2_COUNT_102_x": [YAKUN["G2_COUNT_102_x"].sum()],
                "G2_SUM_1000_102_x": [YAKUN["G2_SUM_1000_102_x"].sum()],
                "G2_COUNT_102_y": [YAKUN["G2_COUNT_102_y"].sum()],
                "G2_SUM_1000_102_y": [YAKUN["G2_SUM_1000_102_y"].sum()]
            })

            # 4️⃣ Barchasini birlashtiramiz
            JAMI_FINAL = pd.concat([sum_row, YAKUN], ignore_index=True)
            JAMI_FINAL

            viloyat_tartib = [
                "O'zbekiston Respublikasi",
                "Qoraqalpog'iston respublikasi",
                "viloyatlar",
                "Andijon viloyati",
                "Buxoro viloyati",
                "Jizzax viloyati",
                "Qashqadaryo viloyati",
                "Navoiy viloyati",
                "Namangan viloyati",
                "Samarqand viloyati",
                "Surxondaryo viloyati",
                "Sirdaryo viloyati",
                "Toshkent viloyati",
                "Farg'ona viloyati",
                "Xorazm viloyati",
                "Toshkent shahri"
            ]


            JAMI_FINAL=JAMI_FINAL.set_index("ADRES")
            JAMI_FINAL=JAMI_FINAL.reindex(viloyat_tartib)
            colm = ["102_x","102_y","foiz","G2_COUNT_103_x","G2_SUM_1000_103_x","G2_COUNT_103_y",
                    "G2_SUM_1000_103_y","G2_COUNT_102_x","G2_SUM_1000_102_x","G2_COUNT_102_y","G2_SUM_1000_102_y"]
            # JAMI_FINAL1 = JAMI_FINAL[~(JAMI_FINAL[colm] == 0).all(axis=1)]

            JAMI_FINAL1=JAMI_FINAL.replace("0","-")


            print("✅ ФИН(обл) sheetga yozildi!")

            # FIN VAZIRLIK boyicha ===========================================================================================





            KATALOG09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]

            # BAZA09=BAZA09[BAZA09["SATR"]==102]

            JAMI09=pd.merge(BAZA09,KATALOG09,on="OKPO",how="left")
            JAMI09=JAMI09[["SOOGU","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format

            # 1️⃣ Ma'lumotlarni tayyorlash
            KATALOG09 = katalog09[["SOOGU", "OKPO"]].copy()
            BAZA09 = baza09[["OKPO", "G1", "G2", "SATR"]].copy()

            # 2️⃣ Kodlar formatini to‘g‘rilash
            KATALOG09["SOOGU"] = KATALOG09["SOOGU"].astype(str).str.zfill(5)

            # 3️⃣ Asosiy kodlar ro‘yxati
            asosiy_kodlar = [
                "01014","01024","01074","01094","01104","01124","01144","01164",
                "01354","03504","04043","04403","04413","06184","06213","06224",
                "06264","07254","08114","08524","08654"
            ]
            # === 5. “Айрим вазирлик ва идоралар бўйича” uchun sum ===



            # 4️⃣ Hisoblash funksiyasi
            asosiy_kodlar = [
                "01014","01024","01074","01094","01104","01124","01144","01164","01354"
            ,"03504","04043","04403","04413","06184","06213","06224","06264","07254","08114","08524","08654"

            ]

            # 👉 faqat ayrimlarini ajratamiz
            ayrim_kodlar = ["01144", "04043","01164","06213", "04413", "06184", "07254"]

            def hisobla(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, KATALOG09, on="OKPO", how="left")[["SOOGU", "G1", "G2"]]
                
                # 🔹 Asosiy kodlar bo‘yicha hisoblash
                asosiy = (
                    J[J["SOOGU"].isin(asosiy_kodlar)]
                    .groupby("SOOGU", as_index=False)
                    .agg(
                        COUNT_G1=("G1", lambda x: (x > 0).sum()),
                        SUM_G1=("G1", lambda x: x.sum() / 1000),
                        COUNT_G2=("G2", lambda x: (x > 0).sum()),
                        SUM_G2=("G2", lambda x: x.sum() / 1000)
                    )
                )

                # 🔹 Boshqalar (asosiy kodlarga kirmaganlar)
                boshqalar = J[~J["SOOGU"].isin(asosiy_kodlar)]
                boshqalar_agg = pd.DataFrame({
                    "SOOGU": ["99999"],
                    "COUNT_G1": [(boshqalar["G1"] > 0).sum()],
                    "SUM_G1": [boshqalar["G1"].sum() / 1000],
                    "COUNT_G2": [(boshqalar["G2"] > 0).sum()],
                    "SUM_G2": [boshqalar["G2"].sum() / 1000]
                })

                # 🔹 Jamlangan ro‘yxat
                jami = pd.concat([asosiy, boshqalar_agg], ignore_index=True)

                # 🔹 Ayrim vazirlik va idoralar bo‘yicha (faqat ayrim_kodlar bo‘yicha)
                ayrim = asosiy[asosiy["SOOGU"].isin(ayrim_kodlar)]
                ayrim_sum = pd.DataFrame({
                    "SOOGU": ["00001"],
                    "COUNT_G1": [ayrim["COUNT_G1"].sum()],
                    "SUM_G1": [ayrim["SUM_G1"].sum()],
                    "COUNT_G2": [ayrim["COUNT_G2"].sum()],
                    "SUM_G2": [ayrim["SUM_G2"].sum()]
                })

                # 🔹 Ўзбекистон Республикаси бўйича жами
                total = pd.DataFrame({
                    "SOOGU": ["00000"],
                    "COUNT_G1": [jami["COUNT_G1"].sum()],
                    "SUM_G1": [jami["SUM_G1"].sum()],
                    "COUNT_G2": [jami["COUNT_G2"].sum()],
                    "SUM_G2": [jami["SUM_G2"].sum()]
                })

                # 🔹 Hammasini birlashtiramiz
                j = pd.concat([total, ayrim_sum, jami], ignore_index=True)
                
                return j


                
            res102 = hisobla(BAZA09, 102)
            res103 = hisobla(BAZA09, 103)
            res102=res102.fillna(0)
            res103=res103.fillna(0)

            hisob=pd.merge(res103,res102,on="SOOGU")
            hisob["103_zf"]=hisob["SUM_G1_x"]-hisob["SUM_G2_x"]
            hisob["102_zf"]=hisob["SUM_G1_y"]-hisob["SUM_G2_y"]
            hisob

            mapping_list = [
                ["00000", "Республика бўйича жами"],
                ["00001", "Айрим вазирлик ва идоралар бўйича"],
                ["04403", "Ўзбекистон Республикаси Қурилиш ва уй-жой коммунал хўжалиги вазирлиги"],
                ["01354", "«Ўзавтосаноат» АЖ"],
                ["08114", "«Ўздонмаҳсулот» АК"],
                ["08654", "«Ўзкимёсаноат» АЖ"],
                ["06264", "«Ўзагротехсаноатхолдинг» АЖ"],
                ["03504", "«Ўзбекистон темир йўллари» АЖ"],
                ["01024", "«Ўзбекнефтгаз» АЖ"],
                ["01124", "«Ҳудудгазтаъминот» АЖ"],
                ["01104", "«Ўзтрансгаз» АЖ"],
                ["06224", "«Ўзпахтасаноат» АЖ"],
                ["01014", "«Ўзбекистон миллий электр тармоқлари» АЖ"],
                ["01074", "«Ҳудудий электр тармоқлари» АЖ"],
                ["01094", "«Иссиқлик электр станциялари» АЖ"],
                ["08524", "Ўзбекистон Республикаси Тоғ-кон саноати ва геология вазирлиги"],
                
                # ["06213", "«Олмалиқ кон-металлургия комбинати» АЖ"],
                # ["01164", "«Навоий кон-металлургия комбинати» АЖ"],
                ["99999", "Бошқалар"]
            ]
            mapping_df = pd.DataFrame(mapping_list, columns=["SOOGU", "NAIMUZ"])

            # ==============================
            # 7️⃣ Mapping bilan birlashtiramiz
            # ==============================
            hisob = mapping_df.merge(hisob, on="SOOGU", how="left").fillna(0)

            # ==============================
            # 8️⃣ Tartibni mapping_list tartibida saqlaymiz',
            # ==============================
            hisob["order"] = hisob.index
            yakuniy = hisob.sort_values("order").drop(columns="order")
            yakuniy["FOIZ"]=yakuniy["102_zf"]/yakuniy["103_zf"]*100
            yakuniy
            YAKUN2=yakuniy = yakuniy[[
                "NAIMUZ", "103_zf", "102_zf", "FOIZ",
                "COUNT_G1_x", "SUM_G1_x", "COUNT_G2_x", "SUM_G2_x",
                "COUNT_G1_y", "SUM_G1_y", "COUNT_G2_y", "SUM_G2_y"
            ]]

            colm = ["103_zf", "102_zf", "FOIZ",
                "COUNT_G1_x", "SUM_G1_x", "COUNT_G2_x", "SUM_G2_x",
                "COUNT_G1_y", "SUM_G1_y", "COUNT_G2_y", "SUM_G2_y"]
            # JAMI_FINAL1 = JAMI_FINAL1[~(JAMI_FINAL1[colm] == 0).all(axis=1)]

            YAKUN2=YAKUN2.set_index("NAIMUZ")
            YAKUN2=YAKUN2.replace("0","-")

        
            

            yoz_sheetga(df_final,wb,"дебитор-по министр",start_row=10,start_col=2)
            yoz_sheetga(df_final_2,wb,"кредитор-по министр",start_row=9,start_col=2)
            yoz_sheetga(DEB_VILOYAT,wb,"дебитор-по обл",start_row=8,start_col=2)
            yoz_sheetga(KIR_VILOYAT,wb,"кредитор-по обл",start_row=8,start_col=2)
            yoz_sheetga(JAMI_ub,wb,"обл убытка",start_row=5,start_col=2)
            yoz_sheetga(yakuniy_v,wb,"минс убытка",start_row=6,start_col=2)
            yoz_sheetga(YAKUN1,wb,"таб 1",start_row=7,start_col=2)
            yoz_sheetga(YAKUN3,wb,"таб 3",start_row=7,start_col=2)
            yoz_sheetga(yakuniy1,wb,"таб 5",start_row=7,start_col=2)
            yoz_sheetga(yakuniy2,wb,"таб 7",start_row=7,start_col=2)
            yoz_sheetga(JAMI_FINAL1,wb,"ФИН(обл)",start_row=6,start_col=2)
            yoz_sheetga(YAKUN2,wb,"ФИН (мин)",start_row=6,start_col=2)





            # =========================
            # Streamlit ilova qismi
            # =========================
            # st.title("💰 Moliyaviy Hisobot (Cloud PDF)")

            # Excel buffer yaratish (sizning wb obyekt)
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # PDF yaratish
            # pdf_buffer = excel_sheets_to_pdf(excel_buffer)

            # # Yuklab olish tugmasi
            # st.download_button(
            #     "📄 PDF yuklab olish",
            #     data=pdf_buffer,
            #     file_name="hisobot.pdf",
            #     mime="application/pdf"
            # )

            # 🔽 Yuklab olish
            st.download_button(
                "📥 Excel yuklab olish",
                data=excel_buffer,
                file_name=f"{output_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )









    #====================================================================================================
    #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================   
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================

            
    # ========================================
    # MOLIYA NATIJA
    # ========================================
    elif choice == "Moliya Natija":
        st.header("📊 Moliya Natija uchun fayllarni tanlang")

        katalog = st.file_uploader("Katalog", type=["xlsx"])
        baza = st.file_uploader("Baza", type=["xlsx"])
        oked = st.file_uploader("Oked", type=["xlsx"])
        soogu = st.file_uploader("SOOGU", type=["xlsx"])
        template = st.file_uploader("Shablon", type=["xlsx"])

        output_name = st.text_input("Natija nomi")




        def numeric_like_kivy(df, cols):
            for col in cols:
                if col in df.columns:
                    df[col] = (
                        df[col]
                        .astype(str)
                        .str.replace(",", "", regex=False)
                        .str.strip()
                    )
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
                return df
            # --- Excelga yozish funksiyasi 
        from openpyxl.styles import Alignment
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils.dataframe import dataframe_to_rows

        def yoz_sheetga(df, wb, sheet_name, start_row=4, start_col=1):
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            df = df.reset_index(drop=True)

            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=False),
                start=start_row
            ):
                for c_idx, value in enumerate(row, start=start_col):
                    cell = ws.cell(row=r_idx, column=c_idx)

                    if isinstance(cell, MergedCell):
                        continue

                    cell.value = value
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.0'


            from io import BytesIO
            import tempfile
            import os
        # def excel_bytes_to_pdf_bytes(excel_bytes):
        #         with tempfile.TemporaryDirectory() as tmpdir:
        #             excel_path = os.path.join(tmpdir, "temp.xlsx")
        #             pdf_path = os.path.join(tmpdir, "temp.pdf")

        #             with open(excel_path, "wb") as f:
        #                 f.write(excel_bytes.getbuffer())

        #             excel_to_clean_pdf(excel_path, pdf_path)

        #             pdf_buffer = BytesIO()
        #             with open(pdf_path, "rb") as f:
        #                 pdf_buffer.write(f.read())

        #             pdf_buffer.seek(0)
        #             return pdf_buffer




        if st.button("START"):
            if katalog is None:
                st.error("❌ Katalog fayli tanlanmadi")
            elif baza is None:
                st.error("❌ Baza fayli tanlanmadi")
            elif oked is None:
                st.error("❌ Oked fayli tanlanmadi")
            elif soogu is None:
                st.error("❌ SOOGU fayli tanlanmadi")
            elif template is None:
                st.error("❌ Shablon fayli tanlanmadi")
            elif not output_name.strip():
                st.error("❌ Natija nomini kiriting")
        
            else:
                st.success("✅ Hamma narsa joyida, ishni boshlaymiz")

        if st.button("🚪 Chiqish"):
            st.session_state.logged_in = False
            st.rerun()


            # 1️⃣ Excel’larni TO‘G‘RI o‘qish
            katalog09 = pd.read_excel(
                katalog,
                dtype={"OKPO": str, "SOOGU": str, "OKED": str, "ADRES": str}
            )

            baza09 = pd.read_excel(baza, dtype={"OKPO": str})
            oked_avgust = pd.read_excel(oked, dtype={"kod2": str, "OKED": str})
            SOOGU = pd.read_excel(soogu, dtype={"SOOGU": str})

            # 2️⃣ FAQAT numeric ustunlar
            baza09 = numeric_like_kivy(
                baza09, ["G1","G2","G3","G4","G5","G6","G7"]
            )

                        





        #===============================================================================================

            from openpyxl import load_workbook

            wb = load_workbook(template)

            KATALOG09=katalog09[["OKPO","ADRES"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.replace("`", "'", regex=False).str.capitalize().str.strip()


            # 1️⃣ 102 va 103 satrlarni ajratamiz
            BAZA09_102 = BAZA09[BAZA09["SATR"] == 102]
            BAZA09_103 = BAZA09[BAZA09["SATR"] == 103]
            KATALOG09["ADRES"] = KATALOG09["ADRES"].str.split(",").str[0]
            # 2️⃣ Har ikkalasini KATALOH bilan bog‘laymiz

            #--------------------------------------------------------------------------------------------------------
            JAMI_102 = pd.merge(BAZA09_102, KATALOG09, on="OKPO", how="left")[["ADRES", "G1","G2"]]
            JAMI_103 = pd.merge(BAZA09_103, KATALOG09, on="OKPO", how="left")[["ADRES", "G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format

            jami_102=JAMI_102.groupby("ADRES")[["G1","G2"]].sum()
            jami_102["102"]=jami_102["G1"]-jami_102["G2"]
            hisob_102=jami_102["102"].reset_index()
            hisob_102

            jami_103=JAMI_103.groupby("ADRES")[["G1","G2"]].sum()
            jami_103["102"]=jami_103["G1"]-jami_103["G2"]
            hisob_103=jami_103["102"].reset_index()
            hisob_103
            H=pd.merge(hisob_103,hisob_102,on="ADRES")
            H["102_x"]=H["102_x"]/1000
            H["102_y"]=H["102_y"]/1000
            H["foiz"]=H["102_y"]/H["102_x"]*100
            H

            JAMI_102 = pd.merge(BAZA09_102, KATALOG09, on="OKPO", how="left")[["ADRES", "G2","G1"]]
            JAMI_103 = pd.merge(BAZA09_103, KATALOG09, on="OKPO", how="left")[["ADRES", "G2","G1"]]



            JAMI102_Z = (
                JAMI_102.groupby("ADRES", as_index=False)
                    .agg(
                        G2_COUNT_102=("G2", lambda x: (x > 0).sum()),
                        G2_SUM_1000_102=("G2", lambda x: x.sum() / 1000)
                    )
            )
            JAMI102_F = (
                JAMI_102.groupby("ADRES", as_index=False)
                    .agg(
                        G2_COUNT_102=("G1", lambda x: (x > 0).sum()),
                        G2_SUM_1000_102=("G1", lambda x: x.sum() / 1000)
                    )
            )





            JAMI103_Z = (
                JAMI_103.groupby("ADRES", as_index=False)
                    .agg(
                        G2_COUNT_103=("G2", lambda x: (x > 0).sum()),
                        G2_SUM_1000_103=("G2", lambda x: x.sum() / 1000)
                    )
            )
            JAMI103_F = (
                JAMI_103.groupby("ADRES", as_index=False)
                    .agg(
                        G2_COUNT_103=("G1", lambda x: (x > 0).sum()),
                        G2_SUM_1000_103=("G1", lambda x: x.sum() / 1000)
                    )
            )


            dfs = [JAMI103_F,JAMI103_Z,JAMI102_F,JAMI102_Z]

            # Har bir df’da G1 ustun nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)

            # Hammasini 'ADRES' bo‘yicha birlashtirish
            yakuniy = reduce(lambda left, right: pd.merge(left, right, on="ADRES", how="outer"), dfs)

            YAKUN=pd.merge(H,yakuniy,on="ADRES")
            YAKUN

            sum_row = pd.DataFrame({
                "ADRES": ["O'zbekiston Respublikasi"],
                "102_x": [YAKUN["102_x"].sum()],
                "102_y": [YAKUN["102_y"].sum()],
                "foiz": [(YAKUN["102_y"].sum() / YAKUN["102_x"].sum()) * 100],
                "G2_COUNT_103_x": [YAKUN["G2_COUNT_103_x"].sum()],
                "G2_SUM_1000_103_x": [YAKUN["G2_SUM_1000_103_x"].sum()],
                "G2_COUNT_103_y": [YAKUN["G2_COUNT_103_y"].sum()],
                "G2_SUM_1000_103_y": [YAKUN["G2_SUM_1000_103_y"].sum()],
                "G2_COUNT_102_x": [YAKUN["G2_COUNT_102_x"].sum()],
                "G2_SUM_1000_102_x": [YAKUN["G2_SUM_1000_102_x"].sum()],
                "G2_COUNT_102_y": [YAKUN["G2_COUNT_102_y"].sum()],
                "G2_SUM_1000_102_y": [YAKUN["G2_SUM_1000_102_y"].sum()]
            })

            # 4️⃣ Barchasini birlashtiramiz
            JAMI_FINAL = pd.concat([sum_row, YAKUN], ignore_index=True)
            JAMI_FINAL

            viloyat_tartib = [
                "O'zbekiston Respublikasi",
                "Qoraqalpog'iston respublikasi",
                "viloyatlar",
                "Andijon viloyati",
                "Buxoro viloyati",
                "Jizzax viloyati",
                "Qashqadaryo viloyati",
                "Navoiy viloyati",
                "Namangan viloyati",
                "Samarqand viloyati",
                "Surxondaryo viloyati",
                "Sirdaryo viloyati",
                "Toshkent viloyati",
                "Farg'ona viloyati",
                "Xorazm viloyati",
                "Toshkent shahri"
            ]


            JAMI_FINAL=JAMI_FINAL.set_index("ADRES")
            JAMI_FINAL=JAMI_FINAL.reindex(viloyat_tartib)

            JAMI_FINAL=JAMI_FINAL[["102_y","102_x","foiz","G2_COUNT_102_x","G2_SUM_1000_102_x","G2_COUNT_102_y","G2_SUM_1000_102_y"]]
            #-------------------------------------------------------------------------------------------------------------------
            cols = ["102_y","102_x","foiz","G2_COUNT_102_x","G2_SUM_1000_102_x","G2_COUNT_102_y","G2_SUM_1000_102_y"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                JAMI_FINAL[col] = JAMI_FINAL[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                JAMI_FINAL[col] = pd.to_numeric(JAMI_FINAL[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                JAMI_FINAL[col] = JAMI_FINAL[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )
            JAMI_FINAL=JAMI_FINAL.fillna(" ").reset_index()
            JAMI_FINAL["RUSCHA"]=["Республика Узбекистан","Республика Каракалпакстан"," области:","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]
            JAMI_FINAL=JAMI_FINAL.replace("0","-")

            
            JAMI_FINAL = JAMI_FINAL[~(JAMI_FINAL[cols] == 0).all(axis=1)]

            

            #==============================================================================================
            # vazirlik




        


            KATALOG09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]



            

            def hisoblar(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, KATALOG09, on="OKPO", how="left")[["SOOGU", "G1", "G2"]]

                JAMI_KOR=J.groupby("SOOGU").agg(
                        COUNT_G1=("G1", lambda x: (x > 0).sum()),
                        SUM_G1=("G1", lambda x: x.sum() / 1000),
                        COUNT_G2=("G2", lambda x: (x > 0).sum()),
                        SUM_G2=("G2", lambda x: x.sum() / 1000)
                    )
                return JAMI_KOR
            H103=hisoblar(BAZA09,103)
            H102=hisoblar(BAZA09,102)
            H102=pd.merge(H102,SOOGU,on="SOOGU",how="left")
            H103



            H103["ZF"]=(H103["SUM_G1"]-H103["SUM_G2"])
            HH103=H103["ZF"]

            H102=pd.merge(H102,HH103,on="SOOGU")
            H102["ZF2"]=H102["SUM_G1"]-H102["SUM_G2"]


            total = pd.DataFrame({
                    "SOOGU": ["00000"],
                    "COUNT_G1": [H102["COUNT_G1"].sum()],
                    "SUM_G1": [H102["SUM_G1"].sum()],
                    "COUNT_G2": [H102["COUNT_G2"].sum()],
                    "SUM_G2": [H102["SUM_G2"].sum()],
                    "NAIM":["Всего по видам экономической деятельности"],
                    "NAIMUZ":["Jami iqtisodiy faoliyat turlari bo‘yicha"],
                    "ZF":[H102["ZF"].sum()],
                    "ZF2":[H102["ZF2"].sum()]
                })
            j = pd.concat([total, H102], ignore_index=True)
            j["foiz"]=j["ZF2"]/j["ZF"]*100

            j=j.set_index("SOOGU")
            j=j[["NAIMUZ","ZF2","ZF","foiz","COUNT_G1","SUM_G1","COUNT_G2","SUM_G2","NAIM"]]

            cols =["ZF2","ZF","foiz","COUNT_G1","SUM_G1","COUNT_G2","SUM_G2"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                j[col] = j[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                j[col] = pd.to_numeric(j[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                j[col] = j[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


            j = j[~(j[cols] == 0).all(axis=1)]
            j=j.fillna(" ").replace("0","-")
            j

            
            #=================================================================================================
            # korhona





            # oked_avgust=oked_avgust.drop_duplicates(subset=["kod2"])




            AVGUST_K=katalog09[["OKPO","OKED","SEK1"]]

            AVGUST_B=baza09[["OKPO","G1","G2","SATR"]]


            AVGUST_B_K=pd.merge(AVGUST_B,AVGUST_K,on="OKPO",how="left")

            AVGUST_B_K["kod22"]=AVGUST_B_K["OKED"].astype(str).str[:2]


                            
            jami= pd.merge(oked_avgust,AVGUST_B_K,on="OKED",how="left")
            jami=jami.drop_duplicates()
            jami

            jami_103=jami[jami["SATR"]==103]
            jami=jami[jami["SATR"]==102]

            hisob_102=jami.groupby([ "SEK1","kod2"]).agg(COUNT_G1=("G1",lambda x: (x>0).sum()),
                                                        SUM_G1=("G1",lambda x: x.sum()/1000),
                                                        COUNT_G2=("G2",lambda x: (x>0).sum()),
                                                        SUM_G2=("G2",lambda x: x.sum()/1000)
                                                        
                                                        )

            hisob_102
            hisob_102["dif"]=hisob_102["SUM_G1"]-hisob_102["SUM_G2"]
            pd.options.display.float_format = '{:,.1f}'.format
            hisob_102=hisob_102.reset_index().set_index("SEK1")
            #h=h[h["SEK1"]=="B"]
            hisob_102

            hisob_103=jami_103.groupby(["SEK1","kod2"])[["G1","G2"]].sum(["G1","G2"])
            hisob_103["dif2"]=(hisob_103["G1"]-hisob_103["G2"])/1000
            hisob_103=hisob_103[["dif2"]].reset_index()


            HISOB=pd.merge(hisob_103,hisob_102,on="kod2")



                            




        
            HISOB


            for col in ["dif2", "COUNT_G1", "SUM_G1", "COUNT_G2", "SUM_G2", "dif"]:
                HISOB[col] = HISOB[col].astype(str).str.replace(",", "").astype(float)

            import string

            # 21 ta harf (A–U)
            letters = list(string.ascii_uppercase[:21])  
            # ['A','B','C',...,'U']

            # UMUMIY satrlarni qo‘shamiz
            agg_df = HISOB.groupby("SEK1").sum(numeric_only=True).reset_index()

            # Har bir umumiy satrga tartib bo'yicha A, B, C, ...
            agg_df["kod2"] = letters[:len(agg_df)]

            # Birlashtiramiz
            result = pd.concat([agg_df, HISOB], ignore_index=True)

            # Maxsus tartiblash: harflar (A,B,C...) tepada chiqishi uchun
            result["order"] = result["kod2"].apply(lambda x: 0 if x in letters else 1)

            # Har SEK1 ichida harflar tepada
            result = result.sort_values(["SEK1", "order"]).drop(columns="order")


            result = result.set_index(["SEK1","kod2"]).drop_duplicates()
            result=result.reset_index()



            result


            oked=oked_avgust


            oked["kod2"]=oked["OKED"].str[:2]
            oked=oked[["kod2","naimuz","naim"]]
            oked

            # DF2 dan birinchi uchragan qatorni qoldiramiz
            df2_first = oked.drop_duplicates(subset=["kod2"], keep="first")

            # Endi oddiy merge
            KOR_deb = result.merge(df2_first, on="kod2", how="left").reset_index()

            KOR_deb=KOR_deb[["naimuz","dif2", "COUNT_G1", "SUM_G1", "COUNT_G2", "SUM_G2", "dif","naim"]]
            
            letters = list(string.ascii_uppercase[:21]) 
            nn=oked[oked["kod2"].isin(letters)]
            nn=nn["naimuz"].reset_index()
            nn

            # nn dagi naimuz qiymatlari ro‘yxati
            values = nn["naimuz"].unique()

            # KOR_KIR_G2 dan mos kelgan qatorlarni olish
            filtered = KOR_deb[KOR_deb["naimuz"].isin(values)]




            RES=pd.DataFrame({
                "naimuz":["Jami iqtisodiy faoliyat turlari bo‘yicha"],
                "dif2":[filtered["dif2"].sum()],
                "COUNT_G1":[filtered["COUNT_G1"].sum()],
                "SUM_G1":[filtered["SUM_G1"].sum()],
                "COUNT_G2":[filtered["COUNT_G2"].sum()],
                "SUM_G2":[filtered["SUM_G2"].sum()],
                "dif":[filtered["dif"].sum()],
                "naim":["Всего по видам экономической деятельности"]

            })


            KORXONA=pd.concat([RES,KOR_deb],ignore_index=True)
            KORXONA








            mask = (KORXONA["dif"] > 0) & (KORXONA["dif2"] > 0)
            KORXONA.loc[mask, "FOIZ"] = KORXONA.loc[mask, "dif"] / KORXONA.loc[mask, "dif2"] * 100

            KORXONA=KORXONA[["naimuz","dif", "dif2","FOIZ","COUNT_G1", "SUM_G1", "COUNT_G2", "SUM_G2","naim"]].fillna(" ")




            KORXONA = KORXONA.drop_duplicates()






            cols = ["dif2", "COUNT_G1", "SUM_G1", "COUNT_G2", "SUM_G2", "dif", "FOIZ"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                KORXONA[col] = KORXONA[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                KORXONA[col] = pd.to_numeric(KORXONA[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                KORXONA[col] = KORXONA[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )

        
            KORXONA = KORXONA[~(KORXONA[cols] == 0).all(axis=1)]
            KORXONA=KORXONA.fillna(" ").replace("0","-")






            yoz_sheetga(JAMI_FINAL,wb,"viloyat")
            yoz_sheetga(j,wb,"vazirlik")
            yoz_sheetga(KORXONA,wb,"korxona")


            from io import BytesIO
            import tempfile
            import os

            # 🔹 Excel → BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # 🔹 PDF
            # pdf_buffer = excel_bytes_to_pdf_bytes(excel_buffer)

            # 🔽 Yuklab olish
            st.download_button(
                "📥 Excel yuklab olish",
                data=excel_buffer,
                file_name=f"{output_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # st.download_button(
            #     "📄 PDF yuklab olish",
            #     data=pdf_buffer,
            #     file_name=f"{output_name}.pdf",
            #     mime="application/pdf"
            # )



            # output_excel = os.path.join(output_folder, f"{output_name}.xlsx")
            # wb.save(output_excel)

            # output_pdf = os.path.join(output_folder, f"{output_name}.pdf")
            # excel_to_clean_pdf(output_excel, output_pdf)

            # st.success("✅ Moliya Natija tayyor!")
            # st.download_button("📥 Excelni yuklab olish", data=open(output_excel, "rb"), file_name=f"{output_name}.xlsx")
            # st.download_button("📥 PDFni yuklab olish", data=open(output_pdf, "rb"), file_name=f"{output_name}.pdf")

    # ========================================
    # HISOBOT
    # ========================================

    #====================================================================================================
    #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================
    # #====================================================================================================   
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================
    #====================================================================================================


    elif choice == "Hisobot":
        from functools import reduce
        st.header("📊 Hisobot natija uchun fayllarni tanlang")

        katalog = st.file_uploader("Katalog", type=["xlsx"])
        baza = st.file_uploader("Baza", type=["xlsx"])
        oked = st.file_uploader("Oked", type=["xlsx"])
        soogu = st.file_uploader("SOOGU", type=["xlsx"])
        template = st.file_uploader("Shablon", type=["xlsx"])

        output_name = st.text_input("Natija nomi")
        # output_folder = st.text_input("Papka yo'li (masalan C:/Users/User/Desktop)")

        def numeric_like_kivy(df, cols):
            for col in cols:
                if col in df.columns:
                    df[col] = (
                        df[col]
                        .astype(str)
                        .str.replace(",", "", regex=False)
                        .str.strip()
                    )
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
                return df
            # --- Excelga yozish funksiyasi 
                # ---from openpyxl.styles import Alignment
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils.dataframe import dataframe_to_rows

        def yoz_sheetga(df, wb, sheet_name, start_row=4, start_col=1):
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            df = df.reset_index(drop=True)

            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=False),
                start=start_row
            ):
                for c_idx, value in enumerate(row, start=start_col):
                    cell = ws.cell(row=r_idx, column=c_idx)

                    if isinstance(cell, MergedCell):
                        continue

                    cell.value = value
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.0'


    
        # def excel_bytes_to_pdf_bytes(excel_bytes):
        #         with tempfile.TemporaryDirectory() as tmpdir:
        #             excel_path = os.path.join(tmpdir, "temp.xlsx")
        #             pdf_path = os.path.join(tmpdir, "temp.pdf")

        #             with open(excel_path, "wb") as f:
        #                 f.write(excel_bytes.getbuffer())

        #             excel_to_clean_pdf(excel_path, pdf_path)

        #             pdf_buffer = BytesIO()
        #             with open(pdf_path, "rb") as f:
        #                 pdf_buffer.write(f.read())

        #             pdf_buffer.seek(0)
        #             return pdf_buffer



        if st.button("START"):
            if katalog is None:
                st.error("❌ Katalog fayli tanlanmadi")
            elif baza is None:
                st.error("❌ Baza fayli tanlanmadi")
            elif oked is None:
                st.error("❌ Oked fayli tanlanmadi")
            elif soogu is None:
                st.error("❌ SOOGU fayli tanlanmadi")
            elif template is None:
                st.error("❌ Shablon fayli tanlanmadi")
            elif not output_name.strip():
                st.error("❌ Natija nomini kiriting")
        
            else:
                st.success("✅ Hamma narsa joyida, ishni boshlaymiz")

        if st.button("🚪 Chiqish"):
            st.session_state.logged_in = False
            st.rerun()


            # 1️⃣ Excel’larni TO‘G‘RI o‘qish
            katalog09 = pd.read_excel(
                katalog,
                dtype={"OKPO": str, "SOOGU": str, "OKED": str, "ADRES": str}
            )

            baza09 = pd.read_excel(baza, dtype={"OKPO": str})
            oked_avgust = pd.read_excel(oked, dtype={"kod2": str, "OKED": str})
            SOOGU = pd.read_excel(soogu, dtype={"SOOGU": str})

            # 2️⃣ FAQAT numeric ustunlar
            baza09 = numeric_like_kivy(
                baza09, ["G1","G2","G3","G4","G5","G6","G7"]
            )






            # 3️⃣ KATALOG tozalash
            KATALOG09 = katalog09[["OKPO","ADRES","SOOGU"]].copy()
            KATALOG09["ADRES"] = (
                KATALOG09["ADRES"]
                .fillna("")
                .astype(str)
                .str.split(",").str[0]
                .str.replace("`", "'", regex=False)
                .str.capitalize()
                .str.strip()
            )
            KATALOG09.columns = KATALOG09.columns.str.upper()

            BAZA09 = baza09[["OKPO","G1","G2","G3","G4","G5","G6","G7","SATR"]].copy()

            # 4️⃣ SATR bo‘yicha ajratish
            jami = {}
            for satr in [201,202,203,204,205]:
                df = BAZA09[BAZA09["SATR"] == satr]
                m = pd.merge(df, KATALOG09, on="OKPO", how="left")[["ADRES","G1"]]
                jami[satr] = m.groupby("ADRES")["G1"].sum().reset_index()

            dfs = []
            for i, satr in enumerate([201,202,203,204,205], start=1):
                df = jami[satr].rename(columns={"G1": f"G1_{i}"})
                dfs.append(df)

            yakuniy = reduce(lambda l, r: pd.merge(l, r, on="ADRES", how="outer"), dfs)
            yakuniy["G_6"] = yakuniy["G1_1"] - yakuniy["G1_5"]

            yakuniy = yakuniy.set_index("ADRES")

            viloyat_tartib = [
                "O'zbekiston Respublikasi",
                "Qoraqalpog'iston respublikasi",
                "viloyatlar",
                "Andijon viloyati","Buxoro viloyati","Jizzax viloyati",
                "Qashqadaryo viloyati","Navoiy viloyati","Namangan viloyati",
                "Samarqand viloyati","Surxondaryo viloyati","Sirdaryo viloyati",
                "Toshkent viloyati","Farg'ona viloyati","Xorazm viloyati",
                "Toshkent shahri"
            ]

            yakuniy = yakuniy.reindex(viloyat_tartib)
            yakuniy.loc["O'zbekiston Respublikasi"] = yakuniy.sum(numeric_only=True)

            YAKUN_DEB = (yakuniy / 1000).reset_index().fillna(" ")


            cols = ["G1_1","G1_2","G1_3","G1_4","G1_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                YAKUN_DEB[col] = YAKUN_DEB[col].astype(str).str.replace(",", "", regex=False).str.strip().copy()

                # 2) Floatga o'zgartirish
                YAKUN_DEB[col] = pd.to_numeric(YAKUN_DEB[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                YAKUN_DEB[col] = YAKUN_DEB[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )
            YAKUN_DEB=YAKUN_DEB.fillna(" ")

            YAKUN_DEB["RUSCHA"]=["Республика Узбекистан","Республика Каракалпакстан"," области:","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]

            YAKUN_DEB = YAKUN_DEB[~(YAKUN_DEB[cols] == 0).all(axis=1)]

            YAKUN_DEB=YAKUN_DEB.replace("0","-")
            YAKUN_DEB= YAKUN_DEB.reset_index(drop=True)


            # 5️⃣ Excel yozish
            wb = load_workbook(template)

            




            KATALOG09.loc[:,"ADRES"]=KATALOG09["ADRES"].str.split(",").str[0]
            KATALOG09.columns=KATALOG09.columns.str.strip().str.upper()
            KATALOG09.loc[:,"ADRES"] = KATALOG09["ADRES"].str.replace("`", "'", regex=False).str.capitalize().str.strip()


            JAMI=pd.merge(BAZA09,KATALOG09,on="OKPO",how="left")
            JAMI=JAMI[["ADRES","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format
            JAMI





            BAZA09_210 = BAZA09[BAZA09["SATR"] == 210]
            BAZA09_211 = BAZA09[BAZA09["SATR"] == 211]
            BAZA09_212 = BAZA09[BAZA09["SATR"] == 212]
            BAZA09_213 = BAZA09[BAZA09["SATR"] == 213]
            BAZA09_214 = BAZA09[BAZA09["SATR"] == 214]
            BAZA09_215 = BAZA09[BAZA09["SATR"] == 215]
            BAZA09_216 = BAZA09[BAZA09["SATR"] == 216]
            BAZA09_218 = BAZA09[BAZA09["SATR"] == 218]


            # 2️⃣ Har ikkalasini KATALOH bilan bog‘laymiz
            KATALOG09.loc[:,"ADRES"] = KATALOG09["ADRES"].str.split(",").str[0]
            JAMI_210 = pd.merge(BAZA09_210, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_211 = pd.merge(BAZA09_211, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_212 = pd.merge(BAZA09_212, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_213 = pd.merge(BAZA09_213, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_214 = pd.merge(BAZA09_214, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_215 =pd.merge(BAZA09_215, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_216 = pd.merge(BAZA09_216, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]
            JAMI_218 = pd.merge(BAZA09_218, KATALOG09, on="OKPO", how="left")[["ADRES", "G1"]]

            hisob_210=JAMI_210.groupby("ADRES")["G1"].sum().reset_index()
            hisob_211=JAMI_211.groupby("ADRES")["G1"].sum().reset_index()
            hisob_212=JAMI_212.groupby("ADRES")["G1"].sum().reset_index()
            hisob_213=JAMI_213.groupby("ADRES")["G1"].sum().reset_index()
            hisob_214=JAMI_214.groupby("ADRES")["G1"].sum().reset_index()
            hisob_215=JAMI_215.groupby("ADRES")["G1"].sum().reset_index()
            hisob_216=JAMI_216.groupby("ADRES")["G1"].sum().reset_index()
            hisob_218=JAMI_218.groupby("ADRES")["G1"].sum().reset_index()



            # Misol uchun: 5 ta df bor
            dfs_1 = [hisob_210,hisob_211,hisob_212,hisob_213,hisob_214,hisob_215,hisob_216,hisob_218]

            # Har bir df’da G1 ustun nomini noyob qilish
            for i, df in enumerate(dfs_1, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)

            # Hammasini 'ADRES' bo‘yicha birlashtirish
            yakuniy_1 = reduce(lambda left, right: pd.merge(left, right, on="ADRES", how="outer"), dfs_1)


            # Natija
            yakuniy_1["G_9"]=yakuniy_1["G1_1"]-yakuniy_1["G1_7"]


            yakuniy_1=yakuniy_1.set_index("ADRES")
            viloyat_tartib = [
                "O'zbekiston Respublikasi",
                "Qoraqalpog'iston respublikasi",
                "viloyatlar",
                "Andijon viloyati",
                "Buxoro viloyati",
                "Jizzax viloyati",
                "Qashqadaryo viloyati",
                "Navoiy viloyati",
                "Namangan viloyati",
                "Samarqand viloyati",
                "Surxondaryo viloyati",
                "Sirdaryo viloyati",
                "Toshkent viloyati",
                "Farg'ona viloyati",
                "Xorazm viloyati",
                "Toshkent shahri"
            ]

            yakuniy_1 = yakuniy_1.reindex(viloyat_tartib)


            # 1️⃣ Hammasini yig‘amiz (NaN'larni hisobdan chiqarib)
            summalar_1 = yakuniy_1.sum(numeric_only=True)

            # 2️⃣ "Ўзбекистон Республикаси" qatori uchun yig‘indini qo‘yamiz
            yakuniy_1.loc["O'zbekiston Respublikasi"] = summalar_1
            yakuniy_1/1000
            pd.options.display.float_format = '{:,.1f}'.format
            YAKUN_KIR=yakuniy_1/1000

            cols = ["G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                YAKUN_KIR[col] = YAKUN_KIR[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                YAKUN_KIR[col] = pd.to_numeric(YAKUN_KIR[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                YAKUN_KIR[col] = YAKUN_KIR[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )

            YAKUN_KIR = YAKUN_KIR[~(YAKUN_KIR[cols] == 0).all(axis=1)]

            YAKUN_KIR=YAKUN_KIR.fillna(" ").reset_index().replace("0","-")

            YAKUN_KIR["RUSCHA"]=["Республика Узбекистан","Республика Каракалпакстан"," области:","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]
            YAKUN_KIR


    #========================================================================================



            KATALOG09=katalog09[["OKPO","ADRES"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]

            #BAZA09=BAZA09[BAZA09["SATR"]==210]

            KATALOG09.loc[:,"ADRES"]=KATALOG09["ADRES"].str.split(",").str[0]
            KATALOG09.columns=KATALOG09.columns.str.strip().str.upper()
            KATALOG09.loc[:,"ADRES"] = KATALOG09["ADRES"].str.replace("`", "'", regex=False).str.capitalize().str.strip()


            JAMI=pd.merge(BAZA09,KATALOG09,on="OKPO",how="left")
            JAMI=JAMI[["ADRES","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format
            JAMI







            BAZA09_201=BAZA09[BAZA09["SATR"]==201]
            BAZA09_202 = BAZA09[BAZA09["SATR"] == 202]
            BAZA09_203 = BAZA09[BAZA09["SATR"] == 203]
            BAZA09_204 = BAZA09[BAZA09["SATR"] == 204]
            BAZA09_205 = BAZA09[BAZA09["SATR"] == 205]

            # 2️⃣ Har ikkalasini KATALOH bilan bog‘laymiz
            KATALOG09.loc[:,"ADRES"] = KATALOG09["ADRES"].str.split(",").str[0]
            JAMI_201_G2 = pd.merge(BAZA09_201, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_202_G2 = pd.merge(BAZA09_202, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_203_G2 = pd.merge(BAZA09_203, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_204_G2 = pd.merge(BAZA09_204, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_205_G2 =pd.merge(BAZA09_205, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]

            hisob_201_G2=JAMI_201_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_202_G2=JAMI_202_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_203_G2=JAMI_203_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_204_G2=JAMI_204_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_205_G2=JAMI_205_G2.groupby("ADRES")["G2"].sum().reset_index()




            # Misol uchun: 5 ta df bor
            dfs_G2 = [hisob_201_G2,hisob_202_G2,hisob_203_G2,hisob_204_G2,hisob_205_G2]

            # Har bir df’da G1 ustun nomini noyob qilish
            for i, df in enumerate(dfs_G2, start=1):
                df.rename(columns={"G2": f"G2_{i}"}, inplace=True)

            # Hammasini 'ADRES' bo‘yicha birlashtirish
            yakuniy_G2 = reduce(lambda left, right: pd.merge(left, right, on="ADRES", how="outer"), dfs_G2)

            # Natija
            yakuniy_G2["G_6"]=yakuniy_G2["G2_1"]-yakuniy_G2["G2_5"]
            yakuniy

            yakuniy_G2=yakuniy_G2.set_index("ADRES")
            viloyat_tartib = [
                "O'zbekiston Respublikasi",
                "Qoraqalpog'iston respublikasi",
                "viloyatlar",
                "Andijon viloyati",
                "Buxoro viloyati",
                "Jizzax viloyati",
                "Qashqadaryo viloyati",
                "Navoiy viloyati",
                "Namangan viloyati",
                "Samarqand viloyati",
                "Surxondaryo viloyati",
                "Sirdaryo viloyati",
                "Toshkent viloyati",
                "Farg'ona viloyati",
                "Xorazm viloyati",
                "Toshkent shahri"
            ]

            yakuniy_G2 = yakuniy_G2.reindex(viloyat_tartib)
            yakuniy_G2

            # 1️⃣ Hammasini yig‘amiz (NaN'larni hisobdan chiqarib)
            summalar_G2 = yakuniy_G2.sum(numeric_only=True)

            # 2️⃣ "Ўзбекистон Республикаси" qatori uchun yig‘indini qo‘yamiz
            yakuniy_G2.loc["O'zbekiston Respublikasi"] = summalar_G2
            yakuniy/1000
            pd.options.display.float_format = '{:,.1f}'.format
            YAKUN_DEB_G2=yakuniy_G2/1000


            cols = ["G2_1","G2_2","G2_3","G2_4","G2_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                YAKUN_DEB_G2[col] = YAKUN_DEB_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                YAKUN_DEB_G2[col] = pd.to_numeric(YAKUN_DEB_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                YAKUN_DEB_G2[col] = YAKUN_DEB_G2[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )
            YAKUN_DEB_G2 = YAKUN_DEB_G2[~(YAKUN_DEB_G2[cols] == 0).all(axis=1)]

        
            YAKUN_DEB_G2=YAKUN_DEB_G2.fillna(" ").reset_index().replace("0","-")



            YAKUN_DEB_G2["RUSCHA"]=["Республика Узбекистан","Республика Каракалпакстан"," области:","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]
            YAKUN_DEB_G2=YAKUN_DEB_G2[["ADRES","G2_1","G2_2","G2_3","G2_5","G_6","RUSCHA"]]
            YAKUN_DEB_G2
            #========================================================================================================



            KATALOG09=katalog09[["OKPO","ADRES"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]

            #BAZA09=BAZA09[BAZA09["SATR"]==210]

            KATALOG09.loc[:,"ADRES"]=KATALOG09["ADRES"].str.split(",").str[0]
            KATALOG09.columns=KATALOG09.columns.str.strip().str.upper()
            KATALOG09.loc[:,"ADRES"] = KATALOG09["ADRES"].str.replace("`", "'", regex=False).str.capitalize().str.strip()


            JAMI=pd.merge(BAZA09,KATALOG09,on="OKPO",how="left")
            JAMI=JAMI[["ADRES","G1","G2"]]
            pd.options.display.float_format = '{:,.1f}'.format
            JAMI





            BAZA09_210 = BAZA09[BAZA09["SATR"] == 210]
            BAZA09_211 = BAZA09[BAZA09["SATR"] == 211]
            BAZA09_212 = BAZA09[BAZA09["SATR"] == 212]
            BAZA09_213 = BAZA09[BAZA09["SATR"] == 213]
            BAZA09_214 = BAZA09[BAZA09["SATR"] == 214]
            BAZA09_215 = BAZA09[BAZA09["SATR"] == 215]
            BAZA09_216 = BAZA09[BAZA09["SATR"] == 216]
            #BAZA09_218 = BAZA09[BAZA09["SATR"] == 218]


            # 2️⃣ Har ikkalasini KATALOH bilan bog‘laymiz
            KATALOG09.loc[:,"ADRES"] = KATALOG09["ADRES"].str.split(",").str[0]
            JAMI_210_G2 = pd.merge(BAZA09_210, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_211_G2 = pd.merge(BAZA09_211, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_212_G2 = pd.merge(BAZA09_212, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_213_G2 = pd.merge(BAZA09_213, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_214_G2 = pd.merge(BAZA09_214, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_215_G2 =pd.merge(BAZA09_215, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            JAMI_216_G2 = pd.merge(BAZA09_216, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]
            #JAMI_218 = pd.merge(BAZA09_218, KATALOG09, on="OKPO", how="left")[["ADRES", "G2"]]

            hisob_210_G2=JAMI_210_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_211_G2=JAMI_211_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_212_G2=JAMI_212_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_213_G2=JAMI_213_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_214_G2=JAMI_214_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_215_G2=JAMI_215_G2.groupby("ADRES")["G2"].sum().reset_index()
            hisob_216_G2=JAMI_216_G2.groupby("ADRES")["G2"].sum().reset_index()
            #hisob_218=JAMI_218.groupby("ADRES")["G2"].sum().reset_index()



            # Misol uchun: 5 ta df bor
            dfs_1_G2 = [hisob_210_G2,hisob_211_G2,hisob_212_G2,hisob_213_G2,hisob_214_G2,hisob_215_G2,hisob_216_G2]

            # Har bir df’da G1 ustun nomini noyob qilish
            for i, df in enumerate(dfs_1_G2, start=1):
                df.rename(columns={"G2": f"G2_{i}"}, inplace=True)

            # Hammasini 'ADRES' bo‘yicha birlashtirish
            yakuniy_1_G2 = reduce(lambda left, right: pd.merge(left, right, on="ADRES", how="outer"), dfs_1_G2)


            # Natija
            yakuniy_1_G2["G_9"]=yakuniy_1_G2["G2_1"]-yakuniy_1_G2["G2_7"]
            yakuniy

            yakuniy_1_G2=yakuniy_1_G2.set_index("ADRES")
            viloyat_tartib = [
                "O'zbekiston Respublikasi",
                
                "Andijon viloyati",
                "Buxoro viloyati",
                "Jizzax viloyati",
                "Qashqadaryo viloyati",
                "Navoiy viloyati",
                "Namangan viloyati",
                "Samarqand viloyati",
                "Surxondaryo viloyati",
                "Sirdaryo viloyati",
                "Toshkent viloyati",
                "Farg'ona viloyati",
                "Xorazm viloyati",
                "Toshkent shahri"
            ]

            yakuniy_1_G2 = yakuniy_1_G2.reindex(viloyat_tartib)


            # 1️⃣ Hammasini yig‘amiz (NaN'larni hisobdan chiqarib)
            summalar_1_G2 = yakuniy_1_G2.sum(numeric_only=True)

            # 2️⃣ "Ўзбекистон Республикаси" qatori uchun yig‘indini qo‘yamiz
            yakuniy_1_G2.loc["O'zbekiston Respublikasi"] = summalar_1_G2
            yakuniy_1_G2/1000
            pd.options.display.float_format = '{:,.1f}'.format
            YAKUN_KIR_G2=yakuniy_1_G2/1000

            cols = ["G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                YAKUN_KIR_G2[col] = YAKUN_KIR_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                YAKUN_KIR_G2[col] = pd.to_numeric(YAKUN_KIR_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                YAKUN_KIR_G2[col] = YAKUN_KIR_G2[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )
            YAKUN_KIR_G2 = YAKUN_KIR_G2[~(YAKUN_KIR_G2[cols] == 0).all(axis=1)]
            
            YAKUN_KIR_G2=YAKUN_KIR_G2.fillna(" ").reset_index().replace("0","-")

            YAKUN_KIR_G2["RUSCHA"]=["Республика Узбекистан","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]
            YAKUN_KIR_G2

            from functools import reduce





            KATALOG09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]




            def hisoblar(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, KATALOG09, on="OKPO", how="left")[["SOOGU", "G1", "G2"]]

                JAMI_KOR=J.groupby("SOOGU").agg(
                    
                        SUM_G1=("G1", lambda x: x.sum() / 1000)
                        
                    )
                return JAMI_KOR
            H201=hisoblar(BAZA09,201)
            H202=hisoblar(BAZA09,202)
            H203=hisoblar(BAZA09,203)
            H204=hisoblar(BAZA09,204)
            H205=hisoblar(BAZA09,205)

            H201




            dfs_kor_deb = [H201, H202, H203, H204, H205]


            for i, df in enumerate(dfs_kor_deb, start=1):
                df.rename(columns={"SUM_G1": f"G1_{i}"}, inplace=True)
                if "SOOGU" in df.index:
                    df.set_index("SOOGU", inplace=True)
            # Join orqali birlashtirish
            yakuniy_kor_deb = reduce(lambda left, right: left.join(right, how="outer"), dfs_kor_deb)
            yakuniy_kor_deb["G_6"] = yakuniy_kor_deb["G1_1"] - yakuniy_kor_deb["G1_5"]
            yakuniy_kor_deb=yakuniy_kor_deb.reset_index()



            HAM_kor_deb=pd.merge(yakuniy_kor_deb,SOOGU,on="SOOGU",how="left")


            total = pd.DataFrame({
                    "SOOGU": ["00000"],
                    "G1_1": [H201["G1_1"].sum()],
                    "G1_2": [H202["G1_2"].sum()],
                    "G1_3": [H203["G1_3"].sum()],
                    "G1_4": [H204["G1_4"].sum()],
                    "G1_5": [H205["G1_5"].sum()],
                    "G_6": [(yakuniy_kor_deb["G1_1"] - yakuniy_kor_deb["G1_5"]).sum()],
                    "NAIM":["Всего по видам экономической деятельности"],
                    "NAIMUZ":["Jami iqtisodiy faoliyat turlari bo‘yicha"]

                
                })
            VAZIR_DEB = pd.concat([total, HAM_kor_deb], ignore_index=True)





            VAZIR_DEB=VAZIR_DEB.set_index("SOOGU")

            # BARCHA USTUNLAR O BOLGANI OCHIRISH 

            colm = ["G1_1","G1_2","G1_3","G1_4","G1_5","G_6"]

            VAZIR_DEB = VAZIR_DEB[~(VAZIR_DEB[colm] == 0).all(axis=1)]
            #------------------------------------------------------------------------------------------------

            VAZIR_DEB=VAZIR_DEB[["NAIMUZ","G1_1","G1_2","G1_3","G1_4","G1_5","G_6","NAIM"]]

            cols =["G1_1","G1_2","G1_3","G1_4","G1_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                VAZIR_DEB[col] = VAZIR_DEB[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                VAZIR_DEB[col] = pd.to_numeric(VAZIR_DEB[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                VAZIR_DEB[col] = VAZIR_DEB[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


            VAZIR_DEB=VAZIR_DEB.fillna(" ").replace("0","-")
            VAZIR_DEB_DEB=VAZIR_DEB
            #===================================================================================================

            def hisoblar(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, KATALOG09, on="OKPO", how="left")[["SOOGU", "G1", "G2"]]

                JAMI_KOR=J.groupby("SOOGU").agg(
                    
                        SUM_G1=("G1", lambda x: x.sum() / 1000)
                        
                    )
                return JAMI_KOR
            H210=hisoblar(BAZA09,210)
            H211=hisoblar(BAZA09,211)
            H212=hisoblar(BAZA09,212)
            H213=hisoblar(BAZA09,213)
            H214=hisoblar(BAZA09,214)
            H215=hisoblar(BAZA09,215)
            H216=hisoblar(BAZA09,216)
            H218=hisoblar(BAZA09,218)



            dfs_kor_deb_1 = [H210, H211, H212, H213, H214, H215, H216, H218]


            for i, df in enumerate(dfs_kor_deb_1, start=1):
                df.rename(columns={"SUM_G1": f"G1_{i}"}, inplace=True)
                if "SOOGU" in df.index:
                    df.set_index("SOOGU", inplace=True)
            # Join orqali birlashtirish
            yakuniy_kor_deb_1 = reduce(lambda left, right: left.join(right, how="outer"), dfs_kor_deb_1)
            yakuniy_kor_deb_1["G_9"] = yakuniy_kor_deb_1["G1_1"] - yakuniy_kor_deb_1["G1_7"]
            yakuniy_kor_deb_1=yakuniy_kor_deb_1.reset_index()



            HAM_kor_deb_1=pd.merge(yakuniy_kor_deb_1,SOOGU,on="SOOGU",how="left")


            total = pd.DataFrame({
                    "SOOGU": ["00000"],
                    "G1_1": [H210["G1_1"].sum()],
                    "G1_2": [H211["G1_2"].sum()],
                    "G1_3": [H212["G1_3"].sum()],
                    "G1_4": [H213["G1_4"].sum()],
                    "G1_5": [H214["G1_5"].sum()],
                    "G1_6": [H215["G1_6"].sum()],
                    "G1_7": [H216["G1_7"].sum()],
                    "G1_8": [H218["G1_8"].sum()],
                    "G_9": [(yakuniy_kor_deb_1["G1_1"] - yakuniy_kor_deb_1["G1_7"]).sum()],
                    "NAIM":["Всего по видам экономической деятельности"],
                    "NAIMUZ":["Jami iqtisodiy faoliyat turlari bo‘yicha"]

                
                })
            VAZIR_KIR= pd.concat([total, HAM_kor_deb_1], ignore_index=True)





            VAZIR_KIR=VAZIR_KIR.set_index("SOOGU")

            # BARCHA USTUNLAR O BOLGANI OCHIRISH 

            colm = ["G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9"]

            VAZIR_KIR = VAZIR_KIR[~(VAZIR_KIR[colm] == 0).all(axis=1)]
            #------------------------------------------------------------------------------------------------

            VAZIR_KIR=VAZIR_KIR[["NAIMUZ","G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9","NAIM"]]

            cols =["G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                VAZIR_KIR[col] = VAZIR_KIR[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                VAZIR_KIR[col] = pd.to_numeric(VAZIR_KIR[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                VAZIR_KIR[col] = VAZIR_KIR[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


            VAZIR_KIR=VAZIR_KIR.fillna(" ").replace("0","-")
            VAZIR_KIR
            #=============================================================================================================================================


            KATALOG09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]


            def hisoblar(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, KATALOG09, on="OKPO", how="left")[["SOOGU", "G1", "G2"]]

                JAMI_KOR=J.groupby("SOOGU").agg(
                    
                        SUM_G2=("G2", lambda x: x.sum() / 1000)
                        
                    )
                return JAMI_KOR
            H201_G2=hisoblar(BAZA09,201)
            H202_G2=hisoblar(BAZA09,202)
            H203_G2=hisoblar(BAZA09,203)
            H204_G2=hisoblar(BAZA09,204)
            H205_G2=hisoblar(BAZA09,205)






            dfs_kor_kir = [H201_G2, H202_G2, H203_G2, H204_G2, H205_G2]


            for i, df in enumerate(dfs_kor_kir, start=1):
                df.rename(columns={"SUM_G2": f"G2_{i}"}, inplace=True)
                if "SOOGU" in df.index:
                    df.set_index("SOOGU", inplace=True)
            # Join orqali birlashtirish
            yakuniy_kor_kir = reduce(lambda left, right: left.join(right, how="outer"), dfs_kor_kir)

            yakuniy_kor_kir["G_6"] = yakuniy_kor_kir["G2_1"] - yakuniy_kor_kir["G2_5"]
            yakuniy_kor_kir=yakuniy_kor_kir.reset_index()
            yakuniy_kor_kir


            HAM_kor_kir=pd.merge(yakuniy_kor_kir,SOOGU,on="SOOGU",how="left")


            total_1 = pd.DataFrame({
                    "SOOGU": ["00000"],
                    "G2_1": [H201_G2["G2_1"].sum()],
                    "G2_2": [H202_G2["G2_2"].sum()],
                    "G2_3": [H203_G2["G2_3"].sum()],
                    "G2_4": [H204_G2["G2_4"].sum()],
                    "G2_5": [H205_G2["G2_5"].sum()],
                    "G_6": [(yakuniy_kor_kir["G2_1"] - yakuniy_kor_kir["G2_5"]).sum()],
                    "NAIM":["Всего по видам экономической деятельности"],
                    "NAIMUZ":["Jami iqtisodiy faoliyat turlari bo‘yicha"]

                
                })
            VAZIR_DEB_G2 = pd.concat([total_1, HAM_kor_kir], ignore_index=True)




            # BARCHA USTUNLAR O BOLGANI OCHIRISH 

            colm = ["G2_1","G2_2","G2_3","G2_4","G2_5","G_6"]

            VAZIR_DEB_G2 = VAZIR_DEB_G2[~(VAZIR_DEB_G2[colm] == 0).all(axis=1)]
            #------------------------------------------------------------------------------------------------

            VAZIR_DEB_G2=VAZIR_DEB_G2[["NAIMUZ","G2_1","G2_2","G2_3","G2_5","G_6","NAIM"]]

            cols =["G2_1","G2_2","G2_3","G2_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                VAZIR_DEB_G2[col] = VAZIR_DEB_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                VAZIR_DEB_G2[col] = pd.to_numeric(VAZIR_DEB_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                VAZIR_DEB_G2[col] = VAZIR_DEB_G2[col].apply(
                lambda x: (
                        f"{float(x):.1f}".rstrip('0').rstrip('.') 
                        if pd.notnull(x) else x
                    )
                    )


            VAZIR_DEB_G2=VAZIR_DEB_G2.fillna(" ").replace("0","-")
            VAZIR_DEB_G2


            #============================================================================================================================================



            def hisoblar(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, KATALOG09, on="OKPO", how="left")[["SOOGU", "G1", "G2"]]

                JAMI_KOR=J.groupby("SOOGU").agg(
                    
                        SUM_G2=("G2", lambda x: x.sum() / 1000)
                        
                    )
                return JAMI_KOR
            H210_G2=hisoblar(BAZA09,210)
            H211_G2=hisoblar(BAZA09,211)
            H212_G2=hisoblar(BAZA09,212)
            H213_G2=hisoblar(BAZA09,213)
            H214_G2=hisoblar(BAZA09,214)
            H215_G2=hisoblar(BAZA09,215)
            H216_G2=hisoblar(BAZA09,216)
            H218_G2=hisoblar(BAZA09,218)



            dfs_kor_kir_1 = [H210_G2, H211_G2, H212_G2, H213_G2, H214_G2, H215_G2, H216_G2, H218_G2]


            for i, df in enumerate(dfs_kor_kir_1, start=1):
                df.rename(columns={"SUM_G2": f"G2_{i}"}, inplace=True)
                if "SOOGU" in df.index:
                    df.set_index("SOOGU", inplace=True)
            # Join orqali birlashtirish
            yakuniy_kor_kir_1 = reduce(lambda left, right: left.join(right, how="outer"), dfs_kor_kir_1)
            yakuniy_kor_kir_1["G_9"] = yakuniy_kor_kir_1["G2_1"] - yakuniy_kor_kir_1["G2_7"]
            yakuniy_kor_kir_1=yakuniy_kor_kir_1.reset_index()



            HAM_kor_kir_1=pd.merge(yakuniy_kor_kir_1,SOOGU,on="SOOGU",how="left")


            total = pd.DataFrame({
                    "SOOGU": ["00000"],
                    "G2_1": [H210_G2["G2_1"].sum()],
                    "G2_2": [H211_G2["G2_2"].sum()],
                    "G2_3": [H212_G2["G2_3"].sum()],
                    "G2_4": [H213_G2["G2_4"].sum()],
                    "G2_5": [H214_G2["G2_5"].sum()],
                    "G2_6": [H215_G2["G2_6"].sum()],
                    "G2_7": [H216_G2["G2_7"].sum()],
                    "G2_8": [H218_G2["G2_8"].sum()],
                    "G_9": [(yakuniy_kor_kir_1["G2_1"] - yakuniy_kor_kir_1["G2_7"]).sum()],
                    "NAIM":["Всего по видам экономической деятельности"],
                    "NAIMUZ":["Jami iqtisodiy faoliyat turlari bo‘yicha"]

                
                })
            VAZIR_KIR_G2= pd.concat([total, HAM_kor_kir_1], ignore_index=True)




            VAZIR_KIR_G2=VAZIR_KIR_G2.set_index("SOOGU")


            # BARCHA USTUNLAR O BOLGANI OCHIRISH 

            colm = ["G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G2_8","G_9"]

            VAZIR_KIR_G2 = VAZIR_KIR_G2[~(VAZIR_KIR_G2[colm] == 0).all(axis=1)]
            #------------------------------------------------------------------------------------------------

            VAZIR_KIR_G2=VAZIR_KIR_G2[["NAIMUZ","G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G_9","NAIM"]]

            cols =["G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                VAZIR_KIR_G2[col] = VAZIR_KIR_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                VAZIR_KIR_G2[col] = pd.to_numeric(VAZIR_KIR_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                VAZIR_KIR_G2[col] = VAZIR_KIR_G2[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )

            VAZIR_KIR_G2 = VAZIR_KIR_G2[~(VAZIR_KIR_G2[cols] == 0).all(axis=1)]
            VAZIR_KIR_G2=VAZIR_KIR_G2.fillna(" ").replace("0","-")
            VAZIR_KIR_G2


            VAZIR_KIR_G2

            #=====================================================================================================
            #===================================================================================================================================

            AVGUST_K=katalog09[["OKPO","OKED","SEK1"]]

            AVGUST_B=baza09[["OKPO","G1","G2","SATR"]]


            AVGUST_B_K=pd.merge(AVGUST_B,AVGUST_K,on="OKPO",how="left")


            oked_avgust
            oked=oked_avgust



            def BAJAR(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, oked_avgust, on="OKED", how="left")[["SATR","naimuz","naim","G1","G2","kod2","SEK1"]]

                JAM=J.groupby(["SEK1","kod2"]).agg(
                    
                        SUM_G1=("G1", lambda x: x.sum() / 1000)
                        
                    )
                return JAM
            B201=BAJAR(AVGUST_B_K,201)
            B202=BAJAR(AVGUST_B_K,202)
            B203=BAJAR(AVGUST_B_K,203)
            B204=BAJAR(AVGUST_B_K,204)
            B205=BAJAR(AVGUST_B_K,205)

            dfs_B_deb = [B201, B202, B203, B204, B205]


            for i, df in enumerate(dfs_B_deb, start=1):
                df.rename(columns={"SUM_G1": f"G1_{i}"}, inplace=True)
                if "kod2" in df.index:
                    df.set_index("kod2", inplace=True)
            # Join orqali birlashtirish
            yakuniy_B_deb = reduce(lambda left, right: left.join(right, how="outer"), dfs_B_deb)
            yakuniy_B_deb["G_6"]=yakuniy_B_deb["G1_1"]-yakuniy_B_deb["G1_5"]
            yakuniy_B_deb=yakuniy_B_deb.reset_index().drop_duplicates()
            yakuniy_B_deb



            for col in ["G1_1", "G1_2","G1_3","G1_4","G1_5","G_6",]:
                yakuniy_B_deb[col] = yakuniy_B_deb[col].astype(str).str.replace(",", "").astype(float)







            # 21 ta harf (A–U)
            letters = list(string.ascii_uppercase[:21])  
            # ['A','B','C',...,'U']

            # UMUMIY satrlarni qo‘shamiz
            agg_df = yakuniy_B_deb.groupby("SEK1").sum(numeric_only=True).reset_index()

            # Har bir umumiy satrga tartib bo'yicha A, B, C, ...
            agg_df["kod2"] = letters[:len(agg_df)]

            # Birlashtiramiz
            result = pd.concat([agg_df, yakuniy_B_deb], ignore_index=True)

            # Maxsus tartiblash: harflar (A,B,C...) tepada chiqishi uchun
            result["order"] = result["kod2"].apply(lambda x: 0 if x in letters else 1)

            # Har SEK1 ichida harflar tepada
            result = result.sort_values(["SEK1", "order"]).drop(columns="order")


            result = result.set_index(["SEK1","kod2"]).drop_duplicates()
            result=result.reset_index()


            oked=oked_avgust


            oked["kod2"]=oked["OKED"].str[:2]
            oked=oked[["kod2","naimuz","naim"]]
            oked

            # DF2 dan birinchi uchragan qatorni qoldiramiz
            df2_first = oked.drop_duplicates(subset=["kod2"], keep="first")

            # Endi oddiy merge
            KOR_DEB = result.merge(df2_first, on="kod2", how="left")

            KOR_DEB=KOR_DEB[["naimuz","G1_1","G1_2","G1_3","G1_4","G1_5","G_6","naim"]]




            letters = list(string.ascii_uppercase[:21]) 
            nn=oked[oked["kod2"].isin(letters)]
            nn=nn["naimuz"].reset_index()
            nn

            # nn dagi naimuz qiymatlari ro‘yxati
            values = nn["naimuz"].unique()

            # KOR_KIR_G2 dan mos kelgan qatorlarni olish
            filtered = KOR_DEB[KOR_DEB["naimuz"].isin(values)]

            # Endi G ustunlarini sum qilish
            RES = pd.DataFrame({
                "naimuz": ["Jami iqtisodiy faoliyat turlari bo‘yicha"],

                "G1_1": [filtered["G1_1"].sum()],
                "G1_2": [filtered["G1_2"].sum()],
                "G1_3": [filtered["G1_3"].sum()],
                "G1_4": [filtered["G1_4"].sum()],
                "G1_5": [filtered["G1_5"].sum()],
                "G_6": [filtered["G_6"].sum()],
                "naim": ["Всего по видам экономической деятельности"]
            })



            KOR_DEB=pd.concat([RES,KOR_DEB],ignore_index=True)
            KOR_DEB

            cols =["G1_1","G1_2","G1_3","G1_4","G1_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                KOR_DEB[col] = KOR_DEB[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                KOR_DEB[col] = pd.to_numeric(KOR_DEB[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                KOR_DEB[col] = KOR_DEB[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )

            KOR_DEB = KOR_DEB[~(KOR_DEB[cols] == 0).all(axis=1)]

            KOR_DEB=KOR_DEB.fillna(" ").replace("0","-")
            KOR_KOR=KOR_DEB

            #===========================================================================================================================================
            #===========================================================================================================================================



            AVGUST_K=katalog09[["OKPO","OKED","SEK1"]]
            AVGUST_B=baza09[["OKPO","G1","G2","SATR"]]

            AVGUST_B_K=pd.merge(AVGUST_B,AVGUST_K,on="OKPO",how="left")


            oked_avgust
            # TAM=pd.merge(oked_avgust,AVGUST_B_K,on="OKED",how="inner")
            # TAM=TAM[["SATR","naimuz","naim","G1","G2","kod2_x","SEK1"]]
            # TAM

            def BAJAR(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, oked_avgust, on="OKED", how="left")[["SATR","G1","G2","kod2","SEK1"]]

                JAM=J.groupby(["SEK1","kod2"]).agg(
                    
                        SUM_G1=("G1", lambda x: x.sum() / 1000)
                        
                    )
                return JAM
            B210=BAJAR(AVGUST_B_K,210)
            B211=BAJAR(AVGUST_B_K,211)
            B212=BAJAR(AVGUST_B_K,212)
            B213=BAJAR(AVGUST_B_K,213)
            B214=BAJAR(AVGUST_B_K,214)
            B215=BAJAR(AVGUST_B_K,215)
            B216=BAJAR(AVGUST_B_K,216)
            B218=BAJAR(AVGUST_B_K,218)


            dfs_B_kir = [B210,B211,B212,B213,B214,B215,B216,B218,]


            for i, df in enumerate(dfs_B_kir, start=1):
                df.rename(columns={"SUM_G1": f"G1_{i}"}, inplace=True)
                if "kod2" in df.index:
                    df.set_index("kod2", inplace=True)
            # Join orqali birlashtirish
            yakuniy_B_kir = reduce(lambda left, right: left.join(right, how="outer"), dfs_B_kir)
            yakuniy_B_kir["G_9"]=yakuniy_B_kir["G1_1"]-yakuniy_B_kir["G1_7"]
            yakuniy_B_kir=yakuniy_B_kir.reset_index().drop_duplicates()
            yakuniy_B_kir






            for col in ["G1_1", "G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9"]:
                yakuniy_B_kir[col] = yakuniy_B_kir[col].astype(str).str.replace(",", "").astype(float)







            # 21 ta harf (A–U)
            letters = list(string.ascii_uppercase[:21])  
            # ['A','B','C',...,'U']

            # UMUMIY satrlarni qo‘shamiz
            agg_df_kir = yakuniy_B_kir.groupby("SEK1").sum(numeric_only=True).reset_index()

            # Har bir umumiy satrga tartib bo'yicha A, B, C, ...
            agg_df_kir["kod2"] = letters[:len(agg_df)]

            # Birlashtiramiz
            result_kir = pd.concat([agg_df_kir, yakuniy_B_kir], ignore_index=True)

            # Maxsus tartiblash: harflar (A,B,C...) tepada chiqishi uchun
            result_kir["order"] = result_kir["kod2"].apply(lambda x: 0 if x in letters else 1)

            # Har SEK1 ichida harflar tepada
            result_kir = result_kir.sort_values(["SEK1", "order"]).drop(columns="order")



            result_kir = result_kir.set_index(["SEK1","kod2"]).drop_duplicates()
            result_kir=result_kir.reset_index()

            oked=oked_avgust


            oked["kod2"]=oked["OKED"].str[:2]
            oked=oked[["kod2","naimuz","naim"]]
            oked

            # DF2 dan birinchi uchragan qatorni qoldiramiz
            df2_first = oked.drop_duplicates(subset=["kod2"], keep="first")

            # Endi oddiy merge
            KOR_KIR = result_kir.merge(df2_first, on="kod2", how="left")

            KOR_KIR=KOR_KIR[["naimuz","G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9","naim"]]
            KOR_KIR


            # nn dagi naimuz qiymatlari ro‘yxati
            values = nn["naimuz"].unique()

            # KOR_KIR_G2 dan mos kelgan qatorlarni olish
            filtered = KOR_KIR[KOR_KIR["naimuz"].isin(values)]


            RES=pd.DataFrame({
                "naimuz":["Jami iqtisodiy faoliyat turlari bo‘yicha"],
                "G1_1":[filtered["G1_1"].sum()],
                "G1_2":[filtered["G1_2"].sum()],
                "G1_3":[filtered["G1_3"].sum()],
                "G1_4":[filtered["G1_4"].sum()],
                "G1_5":[filtered["G1_5"].sum()],
                "G1_6":[filtered["G1_6"].sum()],
                "G1_7":[filtered["G1_7"].sum()],
                "G1_8":[filtered["G1_8"].sum()],
                "G_9":[ filtered["G_9"].sum()],
                "naim":["Всего по видам экономической деятельности"]

            })


            KOR_KIR=pd.concat([RES,KOR_KIR],ignore_index=True)
            KOR_KIR

            cols =["G1_1","G1_2","G1_3","G1_4","G1_5","G1_5","G1_6","G1_7","G1_8","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                KOR_KIR[col] = KOR_KIR[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                KOR_KIR[col] = pd.to_numeric(KOR_KIR[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                KOR_KIR[col] = KOR_KIR[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


            KOR_KIR = KOR_KIR[~(KOR_KIR[cols] == 0).all(axis=1)]
            KOR_KIR=KOR_KIR.fillna(" ").replace("0","-")
            KOR_KIR

            #===================================================================================================
            #===================================================================================================================================

            AVGUST_K=katalog09[["OKPO","OKED","SEK1"]]

            AVGUST_B=baza09[["OKPO","G1","G2","SATR"]]


            AVGUST_B_K=pd.merge(AVGUST_B,AVGUST_K,on="OKPO",how="left")


            oked_avgust


            def BAJAR(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, oked_avgust, on="OKED", how="left")[["SATR","G1","G2","kod2","SEK1"]]

                JAM=J.groupby(["SEK1","kod2"]).agg(
                    
                        SUM_G2=("G2", lambda x: x.sum() / 1000)
                        
                    )
                return JAM
            B201_G2=BAJAR(AVGUST_B_K,201)
            B202_G2=BAJAR(AVGUST_B_K,202)
            B203_G2=BAJAR(AVGUST_B_K,203)
            B204_G2=BAJAR(AVGUST_B_K,204)
            B205_G2=BAJAR(AVGUST_B_K,205)

            dfs_B_deb_G2 = [B201_G2, B202_G2, B203_G2, B204_G2, B205_G2]


            for i, df in enumerate(dfs_B_deb_G2, start=1):
                df.rename(columns={"SUM_G2": f"G2_{i}"}, inplace=True)
                if "kod2" in df.index:
                    df.set_index("kod2", inplace=True)
            # Join orqali birlashtirish
            yakuniy_B_deb_G2 = reduce(lambda left, right: left.join(right, how="outer"), dfs_B_deb_G2)
            yakuniy_B_deb_G2["G_6"]=yakuniy_B_deb_G2["G2_1"]-yakuniy_B_deb_G2["G2_5"]
            yakuniy_B_deb_G2=yakuniy_B_deb_G2.reset_index().drop_duplicates()
            yakuniy_B_deb_G2



            for col in ["G2_1", "G2_2","G2_3","G2_4","G2_5","G_6",]:
                yakuniy_B_deb_G2[col] = yakuniy_B_deb_G2[col].astype(str).str.replace(",", "").astype(float)







            # 21 ta harf (A–U)
            letters = list(string.ascii_uppercase[:21])  
            # ['A','B','C',...,'U']

            # UMUMIY satrlarni qo‘shamiz
            agg_df_G2 = yakuniy_B_deb_G2.groupby("SEK1").sum(numeric_only=True).reset_index()

            # Har bir umumiy satrga tartib bo'yicha A, B, C, ...
            agg_df_G2["kod2"] = letters[:len(agg_df_G2)]

            # Birlashtiramiz
            result_G2 = pd.concat([agg_df_G2, yakuniy_B_deb_G2], ignore_index=True)

            # Maxsus tartiblash: harflar (A,B,C...) tepada chiqishi uchun
            result_G2["order"] = result_G2["kod2"].apply(lambda x: 0 if x in letters else 1)

            # Har SEK1 ichida harflar tepada
            result_G2 = result_G2.sort_values(["SEK1", "order"]).drop(columns="order")


            result_G2 = result_G2.set_index(["SEK1","kod2"]).drop_duplicates()
            result_G2=result_G2.reset_index()


            oked=oked_avgust


            oked["kod2"]=oked["OKED"].str[:2]
            oked=oked[["kod2","naimuz","naim"]]
            oked

            # DF2 dan birinchi uchragan qatorni qoldiramiz
            df2_first = oked.drop_duplicates(subset=["kod2"], keep="first")

            # Endi oddiy merge
            KOR_DEB_G2 = result_G2.merge(df2_first, on="kod2", how="left")

            KOR_DEB_G2=KOR_DEB_G2[["naimuz","G2_1","G2_2","G2_3","G2_5","G_6","naim"]]


            values=nn["naimuz"].unique()

            filtered=KOR_DEB_G2[KOR_DEB_G2["naimuz"].isin(values)]


            RES=pd.DataFrame({
                "naimuz":["Jami iqtisodiy faoliyat turlari bo‘yicha"],
                "G2_1":[filtered["G2_1"].sum()],
                "G2_2":[filtered["G2_2"].sum()],
                "G2_3":[filtered["G2_3"].sum()],
                "G2_5":[filtered["G2_5"].sum()],
                "G_6":[ filtered["G_6"].sum()],
                "naim":["Всего по видам экономической деятельности"]

            })


            KOR_DEB_G2=pd.concat([RES,KOR_DEB_G2],ignore_index=True)
            KOR_DEB_G2

            cols =["G2_1","G2_2","G2_3","G2_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                KOR_DEB_G2[col] = KOR_DEB_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                KOR_DEB_G2[col] = pd.to_numeric(KOR_DEB_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                KOR_DEB_G2[col] = KOR_DEB_G2[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )

            KOR_DEB_G2 = KOR_DEB_G2[~(KOR_DEB_G2[cols] == 0).all(axis=1)]
            KOR_DEB_G2=KOR_DEB_G2.fillna(" ").replace("0","-")
            KOR_DEB_G2
            #==================================================================================================

            #===================================================================================================================================



            AVGUST_K=katalog09[["OKPO","OKED","SEK1"]]
            AVGUST_B=baza09[["OKPO","G1","G2","SATR"]]


            AVGUST_B_K=pd.merge(AVGUST_B,AVGUST_K,on="OKPO",how="left")





            def BAJAR(df, satr): 
                B = df[df["SATR"] == satr]
                J = pd.merge(B, oked_avgust, on="OKED", how="left")[["SATR","G1","G2","kod2","SEK1"]]

                JAM=J.groupby(["SEK1","kod2"]).agg(
                    
                        SUM_G2=("G2", lambda x: x.sum() / 1000)
                        
                    )
                return JAM
            B210_G2=BAJAR(AVGUST_B_K,210)
            B211_G2=BAJAR(AVGUST_B_K,211)
            B212_G2=BAJAR(AVGUST_B_K,212)
            B213_G2=BAJAR(AVGUST_B_K,213)
            B214_G2=BAJAR(AVGUST_B_K,214)
            B215_G2=BAJAR(AVGUST_B_K,215)
            B216_G2=BAJAR(AVGUST_B_K,216)
            B218_G2=BAJAR(AVGUST_B_K,218)


            dfs_B_kir_G2 = [B210_G2,B211_G2,B212_G2,B213_G2,B214_G2,B215_G2,B216_G2,B218_G2,]


            for i, df in enumerate(dfs_B_kir_G2, start=1):
                df.rename(columns={"SUM_G2": f"G2_{i}"}, inplace=True)
                if "kod2" in df.index:
                    df.set_index("kod2", inplace=True)
            # Join orqali birlashtirish
            yakuniy_B_kir_G2 = reduce(lambda left, right: left.join(right, how="outer"), dfs_B_kir_G2)
            yakuniy_B_kir_G2["G_9"]=yakuniy_B_kir_G2["G2_1"]-yakuniy_B_kir_G2["G2_7"]
            yakuniy_B_kir_G2=yakuniy_B_kir_G2.reset_index().drop_duplicates()
            yakuniy_B_kir_G2






            for col in ["G2_1", "G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G2_8","G_9"]:
                yakuniy_B_kir_G2[col] = yakuniy_B_kir_G2[col].astype(str).str.replace(",", "").astype(float)







            # 21 ta harf (A–U)
            letters = list(string.ascii_uppercase[:21])  
            # ['A','B','C',...,'U']

            # UMUMIY satrlarni qo‘shamiz
            agg_df_kir_G2 = yakuniy_B_kir_G2.groupby("SEK1").sum(numeric_only=True).reset_index()

            # Har bir umumiy satrga tartib bo'yicha A, B, C, ...
            agg_df_kir_G2["kod2"] = letters[:len(agg_df_kir_G2)]

            # Birlashtiramiz
            result_kir_G2 = pd.concat([agg_df_kir_G2, yakuniy_B_kir_G2], ignore_index=True)

            # Maxsus tartiblash: harflar (A,B,C...) tepada chiqishi uchun
            result_kir_G2["order"] = result_kir_G2["kod2"].apply(lambda x: 0 if x in letters else 1)

            # Har SEK1 ichida harflar tepada
            result_kir_G2 = result_kir_G2.sort_values(["SEK1", "order"]).drop(columns="order")



            result_kir_G2 = result_kir_G2.set_index(["SEK1","kod2"]).drop_duplicates()
            result_kir_G2=result_kir_G2.reset_index()

            oked=oked_avgust


            oked["kod2"]=oked["OKED"].str[:2]
            oked=oked[["kod2","naimuz","naim"]]
            oked

            # DF2 dan birinchi uchragan qatorni qoldiramiz
            df2_first = oked.drop_duplicates(subset=["kod2"], keep="first")

            # Endi oddiy merge
            KOR_KIR_G2 = result_kir_G2.merge(df2_first, on="kod2", how="left")

            KOR_KIR_G2=KOR_KIR_G2[["naimuz","G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G_9","naim"]]
            KOR_KIR_G2


            # nn dagi naimuz qiymatlari ro‘yxati
            values = nn["naimuz"].unique()

            # KOR_KIR_G2 dan mos kelgan qatorlarni olish
            filtered = KOR_KIR_G2[KOR_KIR_G2["naimuz"].isin(values)]

            # Endi G ustunlarini sum qilish
            RES = pd.DataFrame({
                "naimuz": ["Jami iqtisodiy faoliyat turlari bo‘yicha"],

                "G2_1": [filtered["G2_1"].sum()],
                "G2_2": [filtered["G2_2"].sum()],
                "G2_3": [filtered["G2_3"].sum()],
                "G2_4": [filtered["G2_4"].sum()],
                "G2_5": [filtered["G2_5"].sum()],
                "G2_6": [filtered["G2_6"].sum()],
                "G2_7": [filtered["G2_7"].sum()],
                "G_9":  [filtered["G_9"].sum()],

                "naim": ["Всего по видам экономической деятельности"]
            })


            KOR_KIR_G2=pd.concat([RES,KOR_KIR_G2],ignore_index=True)
            KOR_KIR_G2

            cols =["G2_1","G2_2","G2_3","G2_4","G2_5","G2_5","G2_6","G2_7","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                KOR_KIR_G2[col] = KOR_KIR_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                KOR_KIR_G2[col] = pd.to_numeric(KOR_KIR_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                KOR_KIR_G2[col] = KOR_KIR_G2[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


            KOR_KIR_G2=KOR_KIR_G2.fillna(" ").replace("0","-")
            KOR_KIR_G2




            KATALOH09=katalog09[["SOOGU","OKPO","ADRES","SOATO"]]
            BAZA09=baza09[["OKPO","G1","SATR","G2"]]
            KATALOH09.loc[:,"SOOGU"] = KATALOH09["SOOGU"].astype(str).str.zfill(5)

            KATALOH09.loc[:,"ADRES"]=KATALOH09["ADRES"].str.split(",").str[0]
            # KATALOH09.columns=KATALOH09.columns.str.strip().str.capitalize()
            KATALOH09.loc[:,"ADRES"] = KATALOH09["ADRES"].str.replace("`", "'", regex=False).str.capitalize().str.strip()

            # Idoralar ro‘yxati (doimiy)
            ayrm = [
                "04403","01354","08114","08654","06264","03504","01024","01124",
                "01104","06224","01014","01074","01094","08524","06213","01164"
            ]

            pd.options.display.float_format = '{:,.1f}'.format


            def hisobla_satr(BAZA09, satr):
                # 1. SATR bo‘yicha filtrlash
                df = BAZA09[BAZA09["SATR"] == satr]

                # 2. Katalog bilan ulash
                df = pd.merge(df, KATALOH09, on="OKPO", how="left")

                # 3. Kerakli ustunlar
                df = df[["SOOGU", "G1", "ADRES"]]

                # 4. Guruhlash
                df = df.groupby(["ADRES", "SOOGU"], as_index=False)["G1"].sum()

                # === Respublika bo‘yicha jami ===
                RES = pd.DataFrame({
                    "ADRES": ["O'zbekiston Respublikasi"],
                    "G1": [df["G1"].sum() / 1000]
                }).set_index("ADRES")

                # === Ayrim vazirlik va idoralar ===
                df_ayr = df[df["SOOGU"].isin(ayrm)]

                RR = df_ayr.groupby("ADRES", as_index=False)["G1"].sum()
                RR["G1"] = RR["G1"] / 1000
                RR = RR.set_index("ADRES")

                AYR = pd.DataFrame({
                    "ADRES": ["Ayrim vazirlik va idoralar bo‘yicha"],
                    "G1": [RR["G1"].sum()]
                }).set_index("ADRES")

                # 5. Yakuniy jadval
                jami = pd.concat([RES, AYR, RR]).reset_index()

                return jami

            HISOB_201_D_M = hisobla_satr(BAZA09, 201)
            HISOB_202_D_M = hisobla_satr(BAZA09, 202)
            HISOB_203_D_M = hisobla_satr(BAZA09, 203)
            HISOB_204_D_M = hisobla_satr(BAZA09, 204)
            HISOB_205_D_M = hisobla_satr(BAZA09, 205)


            from functools import reduce

            dfs = [HISOB_201_D_M, HISOB_202_D_M, HISOB_203_D_M, HISOB_204_D_M, HISOB_205_D_M]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)
                if "ADRES" in df.columns:
                    df.set_index("ADRES", inplace=True)

            # Join orqali birlashtirish
            YAKUN_D_M = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash


            # Yangi ustun qo‘shish
            YAKUN_D_M["G_6"] = YAKUN_D_M["G1_1"] - YAKUN_D_M["G1_5"]

            viloyat_tartib = [
                        "O'zbekiston Respublikasi",
                        "shu jumladan",
                        "Ayrim vazirlik va idoralar bo‘yicha",
                        "shu jumladan",
                        "Qoraqalpog'iston respublikasi",
                        "viloyatlar",
                        "Andijon viloyati",
                        "Buxoro viloyati",
                        "Jizzax viloyati",
                        "Qashqadaryo viloyati",
                        "Navoiy viloyati",
                        "Namangan viloyati",
                        "Samarqand viloyati",
                        "Surxondaryo viloyati",
                        "Sirdaryo viloyati",
                        "Toshkent viloyati",
                        "Farg'ona viloyati",
                        "Xorazm viloyati",
                        "Toshkent shahri"
                    ]
            YAKUN_D_M=YAKUN_D_M.reindex(viloyat_tartib)
            YAKUN_D_M["RUSCHA"]=["Республика Узбекистан","в том числе:","По отдельным министерствам и ведомствам",
                                " в том числе:","Республика Каракалпакстан"," области:","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]
            YAKUN_D_M=YAKUN_D_M.reset_index().replace("0","-")

            cols =["G1_1","G1_2","G1_3","G1_4","G1_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                YAKUN_D_M[col] = YAKUN_D_M[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                YAKUN_D_M[col] = pd.to_numeric(YAKUN_D_M[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                YAKUN_D_M[col] = YAKUN_D_M[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


                YAKUN_D_M = YAKUN_D_M[~(YAKUN_D_M[cols] == 0).all(axis=1)]
                YAKUN_D_M=YAKUN_D_M.fillna(" ").replace("0","-")
            YAKUN_D_M
            #=============================================================================================================


            HISOB_210_K_M = hisobla_satr(BAZA09, 210)
            HISOB_211_K_M = hisobla_satr(BAZA09, 211)
            HISOB_212_K_M = hisobla_satr(BAZA09, 212)
            HISOB_213_K_M = hisobla_satr(BAZA09, 213)
            HISOB_214_K_M = hisobla_satr(BAZA09, 214)
            HISOB_215_K_M = hisobla_satr(BAZA09, 215)
            HISOB_216_K_M = hisobla_satr(BAZA09, 216)
            HISOB_218_K_M = hisobla_satr(BAZA09, 218)


            from functools import reduce

            dfs = [HISOB_210_K_M,HISOB_211_K_M,HISOB_212_K_M,HISOB_213_K_M,HISOB_214_K_M,HISOB_215_K_M,HISOB_216_K_M,HISOB_218_K_M]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)
                if "ADRES" in df.columns:
                    df.set_index("ADRES", inplace=True)

            # Join orqali birlashtirish
            YAKUN_K_M = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash


            # Yangi ustun qo‘shish
            YAKUN_K_M["G_9"] = YAKUN_K_M["G1_1"] - YAKUN_K_M["G1_7"]

            YAKUN_K_M=YAKUN_K_M.reindex(viloyat_tartib)
            YAKUN_K_M["RUSCHA"]=["Республика Узбекистан","в том числе:","По отдельным министерствам и ведомствам",
                                " в том числе:","Республика Каракалпакстан"," области:","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]
            YAKUN_K_M=YAKUN_K_M.reset_index()
            cols =["G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_6","G1_7","G1_8","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                YAKUN_K_M[col] = YAKUN_K_M[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                YAKUN_K_M[col] = pd.to_numeric(YAKUN_K_M[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                YAKUN_K_M[col] = YAKUN_K_M[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


                YAKUN_K_M = YAKUN_K_M[~(YAKUN_K_M[cols] == 0).all(axis=1)]
                YAKUN_K_M=YAKUN_K_M.fillna(" ").replace("0","-")
            YAKUN_K_M

            #============================================================================================================


            def hisobla_satr_g2(BAZA09, satr):
                # 1. SATR bo‘yicha filtrlash
                df = BAZA09[BAZA09["SATR"] == satr]

                # 2. Katalog bilan ulash
                df = pd.merge(df, KATALOH09, on="OKPO", how="left")

                # 3. Kerakli ustunlar
                df = df[["SOOGU", "G2", "ADRES"]]

                # 4. Guruhlash
                df = df.groupby(["ADRES", "SOOGU"], as_index=False)["G2"].sum()

                # === Respublika bo‘yicha jami ===
                RES = pd.DataFrame({
                    "ADRES": ["O'zbekiston Respublikasi"],
                    "G2": [df["G2"].sum() / 1000]
                }).set_index("ADRES")

                # === Ayrim vazirlik va idoralar ===
                df_ayr = df[df["SOOGU"].isin(ayrm)]

                RR = df_ayr.groupby("ADRES", as_index=False)["G2"].sum()
                RR["G2"] = RR["G2"] / 1000
                RR = RR.set_index("ADRES")

                AYR = pd.DataFrame({
                    "ADRES": ["Ayrim vazirlik va idoralar bo‘yicha"],
                    "G2": [RR["G2"].sum()]
                }).set_index("ADRES")

                # 5. Yakuniy jadval
                jami = pd.concat([RES, AYR, RR]).reset_index()

                return jami


            HISOB_201_D_M_G2 = hisobla_satr_g2(BAZA09, 201)
            HISOB_202_D_M_G2 = hisobla_satr_g2(BAZA09, 202)
            HISOB_203_D_M_G2 = hisobla_satr_g2(BAZA09, 203)
            HISOB_204_D_M_G2 = hisobla_satr_g2(BAZA09, 204)
            HISOB_205_D_M_G2 = hisobla_satr_g2(BAZA09, 205)


            from functools import reduce

            dfs = [HISOB_201_D_M_G2, HISOB_202_D_M_G2 ,HISOB_203_D_M_G2, HISOB_204_D_M_G2, HISOB_205_D_M_G2]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G2": f"G2_{i}"}, inplace=True)
                if "ADRES" in df.columns:
                    df.set_index("ADRES", inplace=True)

            # Join orqali birlashtirish
            YAKUN_D_M_G2 = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash


            # Yangi ustun qo‘shish
            YAKUN_D_M_G2["G_6"] = YAKUN_D_M_G2["G2_1"] - YAKUN_D_M_G2["G2_5"]

                    
            YAKUN_D_M_G2=YAKUN_D_M_G2.reindex(viloyat_tartib)
            colm = ["G2_1","G2_2","G2_3","G2_4","G2_5","G_6"]
            YAKUN_D_M_G2["RUSCHA"]=["Республика Узбекистан","в том числе:","По отдельным министерствам и ведомствам",
                                " в том числе:","Республика Каракалпакстан"," области:","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]
            YAKUN_D_M_G2 = YAKUN_D_M_G2[~(YAKUN_D_M_G2[colm] == 0).all(axis=1)].reset_index()
            YAKUN_D_M_G2=YAKUN_D_M_G2[["ADRES","G2_1","G2_2","G2_3","G2_5","G_6","RUSCHA"]]

            cols =["G2_1","G2_2","G2_3","G2_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                YAKUN_D_M_G2[col] = YAKUN_D_M_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                YAKUN_D_M_G2[col] = pd.to_numeric(YAKUN_D_M_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                YAKUN_D_M_G2[col] = YAKUN_D_M_G2[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


                YAKUN_D_M_G2 = YAKUN_D_M_G2[~(YAKUN_D_M_G2[cols] == 0).all(axis=1)]
                YAKUN_D_M_G2=YAKUN_D_M_G2.fillna(" ").replace("0","-")
            YAKUN_D_M_G2
            #=================================================================================================================

            HISOB_210_K_M_G2 = hisobla_satr_g2(BAZA09, 210)
            HISOB_211_K_M_G2 = hisobla_satr_g2(BAZA09, 211)
            HISOB_212_K_M_G2 = hisobla_satr_g2(BAZA09, 212)
            HISOB_213_K_M_G2 = hisobla_satr_g2(BAZA09, 213)
            HISOB_214_K_M_G2 = hisobla_satr_g2(BAZA09, 214)
            HISOB_215_K_M_G2 = hisobla_satr_g2(BAZA09, 215)
            HISOB_216_K_M_G2 = hisobla_satr_g2(BAZA09, 216)
            HISOB_218_K_M_G2 = hisobla_satr_g2(BAZA09, 218)


            from functools import reduce

            dfs = [HISOB_210_K_M_G2,HISOB_211_K_M_G2,HISOB_212_K_M_G2,HISOB_213_K_M_G2,
                HISOB_214_K_M_G2,HISOB_215_K_M_G2,HISOB_216_K_M_G2,HISOB_218_K_M_G2]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G2": f"G2_{i}"}, inplace=True)
                if "ADRES" in df.columns:
                    df.set_index("ADRES", inplace=True)

            # Join orqali birlashtirish
            YAKUN_K_M_G2 = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash


            # Yangi ustun qo‘shish
            YAKUN_K_M_G2["G_9"] = YAKUN_K_M_G2["G2_1"] - YAKUN_K_M_G2["G2_7"]

            YAKUN_K_M_G2=YAKUN_K_M_G2.reindex(viloyat_tartib)
            YAKUN_K_M_G2["RUSCHA"]=["Республика Узбекистан","в том числе:","По отдельным министерствам и ведомствам",
                                " в том числе:","Республика Каракалпакстан"," области:","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]

            colm = ["G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G2_8","G_9"]
            YAKUN_K_M_G2=YAKUN_K_M_G2[~(YAKUN_K_M_G2[colm]==0).all(axis=1)].reset_index()
            YAKUN_K_M_G2


            YAKUN_K_M_G2=YAKUN_K_M_G2[["ADRES","G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G_9","RUSCHA"]]

            cols =["G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                YAKUN_K_M_G2[col] = YAKUN_K_M_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                YAKUN_K_M_G2[col] = pd.to_numeric(YAKUN_K_M_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                YAKUN_K_M_G2[col] = YAKUN_K_M_G2[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )

                YAKUN_K_M_G2 = YAKUN_K_M_G2[~(YAKUN_K_M_G2[cols] == 0).all(axis=1)]
                YAKUN_K_M_G2=YAKUN_K_M_G2.fillna(" ").replace("0","-")
            YAKUN_K_M_G2
            #==================================================================================================







            KATALOH09=katalog09[["SOOGU","OKPO"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR"]]
            KATALOH09.loc[:,"SOOGU"] = KATALOH09["SOOGU"].astype(str).str.zfill(5)

            BAZA_201=BAZA09[BAZA09["SATR"]==201]
            BAZA_202=BAZA09[BAZA09["SATR"]==202]
            BAZA_203=BAZA09[BAZA09["SATR"]==203]
            BAZA_204=BAZA09[BAZA09["SATR"]==204]
            BAZA_205=BAZA09[BAZA09["SATR"]==205]

            JAMI201=pd.merge(BAZA_201,KATALOH09,on="OKPO",how="left")
            JAMI202=pd.merge(BAZA_202,KATALOH09,on="OKPO",how="left")
            JAMI203=pd.merge(BAZA_203,KATALOH09,on="OKPO",how="left")
            JAMI204=pd.merge(BAZA_204,KATALOH09,on="OKPO",how="left")
            JAMI205=pd.merge(BAZA_205,KATALOH09,on="OKPO",how="left")

            pd.options.display.float_format = '{:,.1f}'.format
            JAMI_201=JAMI201[["SOOGU","G1"]]
            JAMI_202=JAMI202[["SOOGU","G1"]]
            JAMI_203=JAMI203[["SOOGU","G1"]]
            JAMI_204=JAMI204[["SOOGU","G1"]]
            JAMI_205=JAMI205[["SOOGU","G1"]]


            import pandas as pd

            tartib = [
                ["00000", "Respublika boʻyicha jami", "Всего по Республике"],
                ["00020", "shu jumladan", "в том числе:"],
                ["00001", "Ayrim vazirlik va idoralar boʻyicha", "По отдельным министерствам и ведомствам"],

                ["04403", "O‘zbekiston Respublikasi Qurilish va uy-joy kommunal xo‘jaligi vazirligi",
                "Министерство строительства и жилищно-коммунального хозяйства Республики Узбекистан"],

                ["01354", "“O‘zavtosanoat” aksiyadorlik jamiyati",
                "Акционерное общество «Узавтосаноат»"],

                ["08114", "“O‘zdonmahsulot” aksiyadorlik kompaniyasi",
                "Акционерная компания «Уздонмахсулот»"],

                ["08654", "“O‘zkimyosanoat” aksiyadorlik jamiyati",
                "Акционерное общество «Узкимесаноат»"],

                ["06264", "“O‘zagrotexsanoatxolding” aksiyadorlik jamiyati",
                "Акционерное общество «Узагротехсаноатхолдинг»"],

                ["03504", "“O‘zbekiston temir yo‘llari” aksiyadorlik jamiyati",
                "Акционерное общество «Узбекистон темир йуллари»"],

                ["01024", "“O‘zbekneftgaz” aksiyadorlik jamiyati",
                "Акционерное общество «Узбекнефтегаз»"],

                ["01124", "“Hududgazta’minot” aksiyadorlik jamiyati",
                "Акционерное общество «Худудгазтаъминот»"],

                ["01104", "“O‘ztransgaz” aksiyadorlik jamiyati",
                "Акционерное общество «Узтрансгаз»"],

                ["06224", "“O‘zpaxtasanoat” aksiyadorlik jamiyati",
                "Акционерное общество «Узпахтасаноат»"],

                ["01014", "“O‘zbekiston milliy elektr tarmoqlari” aksiyadorlik jamiyati",
                "Акционерное общество «Национальные электрические сети Узбекистана»"],

                ["01074", "“Hududiy elektr tarmoqlari” aksiyadorlik jamiyati",
                "Акционерное общество «Региональные электрические сети»"],

                ["01094", "“Issiqlik elektr stansiyalari” aksiyadorlik jamiyati",
                "Акционерное общество «Тепловые электрические станции»"],

                ["08524", "O‘zbekiston Respublikasi Tog‘-kon sanoati va geologiya vazirligi",
                "Министерство горнодобывающей промышленности и геологии Республики Узбекистан"],

                ["99999", "Boshqalar", "Другие"]
            ]

            tartib_df = pd.DataFrame(tartib, columns=["SOOGU", "NAIMUZ", "NAIM"])


            # Idoralar ro‘yxati (har safar ishlatiladi)
            ayrm = ["04403","01354","08114","08654","06264","03504","01024","01124","01104","06224","01014","01074","01094","08524","06213","01164"
            ]
            boshqamin=["01144", "04043", "04413", "06184", "07254"]
            # === FUNKSIYA: Har bir SATR (201, 202, …) bo‘yicha hisoblash ===
            def hisobla(df):
                df = df.copy()
                df["SOOGU"] = df["SOOGU"].astype(str).str.zfill(5)

                # 1. SOOGU bo‘yicha yig‘ish
                agg = df.groupby("SOOGU", as_index=False)["G1"].sum()

                # 2. Tartib bilan birlashtirish
                merged = pd.merge(tartib_df, agg, on="SOOGU", how="left")

                # 3. “Айрим вазирлик ва идоралар бўйича”
                ayrim_sum = df.loc[df["SOOGU"].isin(ayrm), "G1"].sum()
                merged.loc[merged["SOOGU"] == "00001", "G1"] = ayrim_sum

                # 4. “Республика бўйича жами”
                total_sum = df["G1"].sum()
                merged.loc[merged["SOOGU"] == "00000", "G1"] = total_sum

                # 5. “Бошқалар”
                asosiy_kodlar = tartib_df["SOOGU"].tolist() + boshqamin
                boshqalar_sum = df.loc[~df["SOOGU"].isin(asosiy_kodlar), "G1"].sum()
                merged.loc[merged["SOOGU"] == "99999", "G1"] = boshqalar_sum

                # 6. To‘ldirish va 1000 ga bo‘lish
                merged["G1"] = merged["G1"].fillna(0).round(1) / 1000

                # 7. Yakuniy natija
                result = merged[["NAIMUZ","NAIM", "G1"]].set_index(["NAIMUZ","NAIM"])
                return result
            HISOB_201 = hisobla(JAMI_201)
            HISOB_202 = hisobla(JAMI_202)
            HISOB_203 = hisobla(JAMI_203)
            HISOB_204 = hisobla(JAMI_204)
            HISOB_205 = hisobla(JAMI_205)

            from functools import reduce

            dfs = [HISOB_201, HISOB_202, HISOB_203, HISOB_204, HISOB_205]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)
                if "NAIMUZ" in df.columns:
                    df.set_index("NAIMUZ", inplace=True)

            # Join orqali birlashtirish
            yakuniy = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash
            yakuniy = yakuniy.reindex(HISOB_201.index)

            # Yangi ustun qo‘shish
            yakuniy["G_6"] = yakuniy["G1_1"] - yakuniy["G1_5"]

            y=yakuniy.reset_index()
            y=y[["NAIMUZ","G1_1","G1_2","G1_3","G1_4","G1_5","G_6","NAIM"]]
            MINS_DEB=y


            cols =["G1_1","G1_2","G1_3","G1_4","G1_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                MINS_DEB[col] = MINS_DEB[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                MINS_DEB[col] = pd.to_numeric(MINS_DEB[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                MINS_DEB[col] = MINS_DEB[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )

                MINS_DEB = MINS_DEB[~(MINS_DEB[cols] == 0).all(axis=1)]
                MINS_DEB=MINS_DEB.fillna(" ").replace("0","-")
            MINS_DEB

            MINS_DEB
            #===================================================================================================================





            KATALOH09=katalog09[["SOOGU","OKPO","ADRES"]]
            BAZA09=baza09[["OKPO","G1","G2","SATR",'G3', 'G4', 'G5', 'G6', 'G7']]
            KATALOH09.loc[:,"SOOGU"] = KATALOH09["SOOGU"].astype(str).str.zfill(5)

            BAZA_210=BAZA09[BAZA09["SATR"]==210]
            BAZA_211=BAZA09[BAZA09["SATR"]==211]
            BAZA_212=BAZA09[BAZA09["SATR"]==212]
            BAZA_213=BAZA09[BAZA09["SATR"]==213]
            BAZA_214=BAZA09[BAZA09["SATR"]==214]
            BAZA_215=BAZA09[BAZA09["SATR"]==215]
            BAZA_216=BAZA09[BAZA09["SATR"]==216]
            BAZA_218=BAZA09[BAZA09["SATR"]==218]

            JAMI210=pd.merge(BAZA_210,KATALOH09,on="OKPO",how="left")
            JAMI211=pd.merge(BAZA_211,KATALOH09,on="OKPO",how="left")
            JAMI212=pd.merge(BAZA_212,KATALOH09,on="OKPO",how="left")
            JAMI213=pd.merge(BAZA_213,KATALOH09,on="OKPO",how="left")
            JAMI214=pd.merge(BAZA_214,KATALOH09,on="OKPO",how="left")
            JAMI215=pd.merge(BAZA_215,KATALOH09,on="OKPO",how="left")
            JAMI216=pd.merge(BAZA_216,KATALOH09,on="OKPO",how="left")
            JAMI218=pd.merge(BAZA_218,KATALOH09,on="OKPO",how="left")

            pd.options.display.float_format = '{:,.1f}'.format

            JAMI_210=JAMI210[["SOOGU","G1"]]
            JAMI_211=JAMI211[["SOOGU","G1"]]
            JAMI_212=JAMI212[["SOOGU","G1"]]
            JAMI_213=JAMI213[["SOOGU","G1"]]
            JAMI_214=JAMI214[["SOOGU","G1"]]
            JAMI_215=JAMI215[["SOOGU","G1"]]
            JAMI_216=JAMI216[["SOOGU","G1"]]
            JAMI_218=JAMI218[["SOOGU","G1"]]


            #----------------------------------------------------------------

            HISOB_210=hisobla(JAMI_210)
            HISOB_211=hisobla(JAMI_211)
            HISOB_212=hisobla(JAMI_212)
            HISOB_213=hisobla(JAMI_213)
            HISOB_214=hisobla(JAMI_214)
            HISOB_215=hisobla(JAMI_215)
            HISOB_216=hisobla(JAMI_216)
            HISOB_218=hisobla(JAMI_218)


            #-------------------------------------------------------------------------------------------------------------
            from functools import reduce

            dfs = [HISOB_210,HISOB_211,HISOB_212,HISOB_213,HISOB_214,HISOB_215,HISOB_216,HISOB_218]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G1": f"G1_{i}"}, inplace=True)
                if "NAIMUZ" in df.columns:
                    df.set_index("NAIMUZ", inplace=True)

            # Join orqali birlashtirish
            yakuniy = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash
            yakuniy = yakuniy.reindex(HISOB_210.index)

            # Yangi ustun qo‘shish
            yakuniy["G_9"] = yakuniy["G1_1"] - yakuniy["G1_7"]
            yakuniy.reset_index()
            yakuniy=yakuniy.reset_index()
            yakuniy=yakuniy[["NAIMUZ","G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9","NAIM"]]
            MINS_KIR=yakuniy

            cols =["G1_1","G1_2","G1_3","G1_4","G1_5","G1_6","G1_7","G1_8","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                MINS_KIR[col] = MINS_KIR[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                MINS_KIR[col] = pd.to_numeric(MINS_KIR[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                MINS_KIR[col] = MINS_KIR[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )

                MINS_KIR = MINS_KIR[~(MINS_KIR[cols] == 0).all(axis=1)]
                MINS_KIR=MINS_KIR.fillna(" ").replace("0","-")


            MINS_KIR
            #=============================================================================================================


            pd.options.display.float_format = '{:,.1f}'.format
            JAMI_201_G2=JAMI201[["SOOGU","G2"]]
            JAMI_202_G2=JAMI202[["SOOGU","G2"]]
            JAMI_203_G2=JAMI203[["SOOGU","G2"]]
            JAMI_204_G2=JAMI204[["SOOGU","G2"]]
            JAMI_205_G2=JAMI205[["SOOGU","G2"]]







            # === FUNKSIYA: Har bir SATR (201, 202, …) bo‘yicha hisoblash ===
            def hisobla_g2(df):
                df = df.copy()
                df["SOOGU"] = df["SOOGU"].astype(str).str.zfill(5)

                # 1. SOOGU bo‘yicha yig‘ish
                agg = df.groupby("SOOGU", as_index=False)["G2"].sum()

                # 2. Tartib bilan birlashtirish
                merged = pd.merge(tartib_df, agg, on="SOOGU", how="left")

                # 3. “Айрим вазирлик ва идоралар бўйича”
                ayrim_sum = df.loc[df["SOOGU"].isin(ayrm), "G2"].sum()
                merged.loc[merged["SOOGU"] == "00001", "G2"] = ayrim_sum

                # 4. “Республика бўйича жами”
                total_sum = df["G2"].sum()
                merged.loc[merged["SOOGU"] == "00000", "G2"] = total_sum

                # 5. “Бошқалар”
                asosiy_kodlar = tartib_df["SOOGU"].tolist() + boshqamin
                boshqalar_sum = df.loc[~df["SOOGU"].isin(asosiy_kodlar), "G2"].sum()
                merged.loc[merged["SOOGU"] == "99999", "G2"] = boshqalar_sum

                # 6. To‘ldirish va 1000 ga bo‘lish
                merged["G2"] = merged["G2"].fillna(0).round(1) / 1000

                # 7. Yakuniy natija
                result = merged[["NAIMUZ","NAIM", "G2"]].set_index(["NAIMUZ","NAIM"])
                return result
            HISOB_201_G2 = hisobla_g2(JAMI_201_G2)
            HISOB_202_G2 = hisobla_g2(JAMI_202_G2)
            HISOB_203_G2 = hisobla_g2(JAMI_203_G2)
            HISOB_204_G2 = hisobla_g2(JAMI_204_G2)
            HISOB_205_G2= hisobla_g2(JAMI_205_G2)

            from functools import reduce

            dfs = [HISOB_201_G2, HISOB_202_G2, HISOB_203_G2, HISOB_204_G2, HISOB_205_G2]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G2": f"G2_{i}"}, inplace=True)
                if "NAIMUZ" in df.columns:
                    df.set_index("NAIMUZ", inplace=True)

            # Join orqali birlashtirish
            yakuniy_G2 = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash
            yakuniy_G2 = yakuniy_G2.reindex(HISOB_201_G2.index)

            # Yangi ustun qo‘shish
            yakuniy_G2["G_6"] = yakuniy_G2["G2_1"] - yakuniy_G2["G2_5"]

            y=yakuniy_G2.reset_index()
            y=y[["NAIMUZ","G2_1","G2_2","G2_3","G2_5","G_6","NAIM"]]
            MINS_DEB_G2=y
            colm = ["G2_1","G2_2","G2_3","G2_5","G_6"]

            MINS_DEB_G2 = MINS_DEB_G2[~(MINS_DEB_G2[colm] == 0).all(axis=1)]
            MINS_DEB_G2


            cols =["G2_1","G2_2","G2_3","G2_5","G_6"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                MINS_DEB_G2[col] = MINS_DEB_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                MINS_DEB_G2.loc[:,col] = pd.to_numeric(MINS_DEB_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                MINS_DEB_G2.loc[:,col]= MINS_DEB_G2.loc[:,col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


                MINS_DEB_G2=MINS_DEB_G2.fillna(" ").replace("0","-")


            MINS_DEB_G2

            #=======================================================================================================

            JAMI_210_G2=JAMI210[["SOOGU","G2"]]
            JAMI_211_G2=JAMI211[["SOOGU","G2"]]
            JAMI_212_G2=JAMI212[["SOOGU","G2"]]
            JAMI_213_G2=JAMI213[["SOOGU","G2"]]
            JAMI_214_G2=JAMI214[["SOOGU","G2"]]
            JAMI_215_G2=JAMI215[["SOOGU","G2"]]
            JAMI_216_G2=JAMI216[["SOOGU","G2"]]
            JAMI_218_G2=JAMI218[["SOOGU","G2"]]

            #---------------------------------------------------------------------------------------------------------------



            HISOB_210_G2=hisobla_g2(JAMI_210_G2)
            HISOB_211_G2=hisobla_g2(JAMI_211_G2)
            HISOB_212_G2=hisobla_g2(JAMI_212_G2)
            HISOB_213_G2=hisobla_g2(JAMI_213_G2)
            HISOB_214_G2=hisobla_g2(JAMI_214_G2)
            HISOB_215_G2=hisobla_g2(JAMI_215_G2)
            HISOB_216_G2=hisobla_g2(JAMI_216_G2)
            HISOB_218_G2=hisobla_g2(JAMI_218_G2)


            #-------------------------------------------------------------------------------------------------------------
            from functools import reduce

            dfs = [HISOB_210_G2,HISOB_211_G2,HISOB_212_G2,HISOB_213_G2,HISOB_214_G2,HISOB_215_G2,HISOB_216_G2,HISOB_218_G2]

            # Har bir df’da G1 nomini noyob qilish
            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"G2": f"G2_{i}"}, inplace=True)
                if "NAIMUZ" in df.columns:
                    df.set_index("NAIMUZ", inplace=True)

            # Join orqali birlashtirish
            yakuniy = reduce(lambda left, right: left.join(right, how="outer"), dfs)

            # 🔥 Eng muhim qadam — tartibni tiklash
            yakuniy = yakuniy.reindex(HISOB_210.index)

            # Yangi ustun qo‘shish
            yakuniy["G_9"] = yakuniy["G2_1"] - yakuniy["G2_7"]

            yakuniy=yakuniy.reset_index()
            yakuniy=yakuniy[["NAIMUZ","G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G_9","NAIM"]]
            MINS_KIR_G2=yakuniy
            colm = ["G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G_9"]
            MINS_KIR_G2 = MINS_KIR_G2[~(MINS_KIR_G2[colm] == 0).all(axis=1)]
            MINS_KIR_G2


            cols =["G2_1","G2_2","G2_3","G2_4","G2_5","G2_6","G2_7","G_9"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                MINS_KIR_G2[col] = MINS_KIR_G2[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                MINS_KIR_G2.loc[:,col] = pd.to_numeric(MINS_KIR_G2[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                MINS_KIR_G2.loc[:,col] = MINS_KIR_G2[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


                MINS_KIR_G2=MINS_KIR_G2.fillna(" ").replace("0","-")


            MINS_KIR_G2


    #========================================================================================================
            cols = ["OKPO","G1","G2","G3","G4","G5","G6","G7","SATR"]
            existing_cols = [c for c in cols if c in baza09.columns]
            BAZA09 = baza09[existing_cols].copy()
            
            cols1 = ["OKPO","ADRES","SOOGU"]
            existing_cols1 = [c for c in cols1 if c in katalog09.columns]
            KATALOG09 = katalog09[existing_cols1].copy()
                    
            

            KATALOG09.loc[:,"ADRES"]=KATALOG09["ADRES"].str.split(",").str[0]
            KATALOG09.columns=KATALOG09.columns.str.strip().str.upper()
            KATALOG09.loc[:, "ADRES"] = (
                KATALOG09["ADRES"]
                    .str.replace("`", "'", regex=False)
                    .str.capitalize()
                    .str.strip()
            )


            def hisobla_baza(satr, qiymat_ustun):
                """
                SATR bo‘yicha filtrlaydi,
                KATALOG09 bilan OKPO orqali bog‘laydi,
                ADRES bo‘yicha group qilib yig‘indini qaytaradi
                """

                # ADRES ni tozalash
                katalog = KATALOG09.copy()
                katalog["ADRES"] = katalog["ADRES"].str.split(",").str[0]

                # SATR bo‘yicha filtr
                baza_filtr = BAZA09[BAZA09["SATR"] == satr].copy()

                # Merge
                jami = pd.merge(
                    baza_filtr,
                    katalog,
                    on="OKPO",
                    how="left"
                )[["ADRES", qiymat_ustun]]

                # Group by ADRES
                natija = (
                    jami
                    .groupby("ADRES", as_index=False)[qiymat_ustun]
                    .sum()
                )

                summalar = natija.select_dtypes(include="number").sum()

                jami_qator = {"ADRES": "O'zbekiston Respublikasi"}
                jami_qator.update(summalar.to_dict())

                natija = pd.concat([natija, pd.DataFrame([jami_qator])], ignore_index=True)



                return natija

            OY221_G1=hisobla_baza(221,"G1" )
            OY221_G2=hisobla_baza(221,"G2" )
            OY221_G3=hisobla_baza(221,"G3" )
            OY221_G4=hisobla_baza(221,"G4" )
            OY221_G5=hisobla_baza(221,"G5" )
            OY222_G1=hisobla_baza(222,"G1" )
            OY222_G2=hisobla_baza(222,"G2" )
            OY222_G3=hisobla_baza(222,"G3" )
            OY222_G4=hisobla_baza(222,"G4" )
            OY222_G5=hisobla_baza(222,"G5" )

            # Misol uchun: 5 ta df bor
            dfs = [OY221_G1,OY221_G2,OY221_G3,OY221_G4,OY221_G5,OY222_G1,OY222_G2,OY222_G3,OY222_G4,OY222_G5]

            # Har bir df’da G1 ustun nomini noyob qilish
            # for i, df in enumerate(dfs, start=1):
            #     df.rename(columns={"G1": f"G1_{i}"}, inplace=True)

            # Hammasini 'ADRES' bo‘yicha birlashtirish
            yakuniy = reduce(lambda left, right: pd.merge(left, right, on="ADRES", how="outer"), dfs)

            # Natija

            yakuniy=yakuniy.set_index("ADRES")
            viloyat_tartib = [
                "O'zbekiston Respublikasi",
                "Qoraqalpog'iston respublikasi",
                "viloyatlar",
                "Andijon viloyati",
                "Buxoro viloyati",
                "Jizzax viloyati",
                "Qashqadaryo viloyati",
                "Navoiy viloyati",
                "Namangan viloyati",
                "Samarqand viloyati",
                "Surxondaryo viloyati",
                "Sirdaryo viloyati",
                "Toshkent viloyati",
                "Farg'ona viloyati",
                "Xorazm viloyati",
                "Toshkent shahri"
            ]

            yakuniy = yakuniy.reindex(viloyat_tartib).reset_index()


            yakuniy["RUSCHA"]=["Республика Узбекистан","Республика Каракалпакстан"," области:","Андижанская","Бухарская",
                                "Джизакская","Кашкадарьинская","Навоийская","Наманганская","Самаркандская","Сурхандарьинская",
                                "Сырдарьинская","Ташкентская","Ферганская","Хорезмская","г. Ташкент"]
            yakuniy=yakuniy.set_index(["ADRES","RUSCHA"]).div(1000).reset_index()
            yakuniy=yakuniy[["ADRES","G1_x","G1_y","G2_x","G2_y","G3_x","G3_y","G4_x","G4_y","G5_x","G5_y","RUSCHA"]]
            yakuniy

            cols =["G1_x","G1_y","G2_x","G2_y","G3_x","G3_y","G4_x","G4_y","G5_x","G5_y"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                yakuniy[col] = yakuniy[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                yakuniy[col] = pd.to_numeric(yakuniy[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                yakuniy[col] = yakuniy[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )

            yakuniy = yakuniy[~(yakuniy[cols] == 0).all(axis=1)]
            yakuniy=yakuniy.fillna(" ").replace("0","-")
            OY_VILOYAT=yakuniy
            OY_VILOYAT

    #================================================================================================

            def hisobla_baza(satr, qiymat_ustun):
                """
                SATR bo‘yicha filtrlaydi,
                KATALOG09 bilan OKPO orqali bog‘laydi,
                ADRES bo‘yicha group qilib yig‘indini qaytaradi
                """

                # ADRES ni tozalash
                katalog = KATALOG09.copy()
            

                # SATR bo‘yicha filtr
                baza_filtr = BAZA09[BAZA09["SATR"] == satr].copy()

                # Merge
                jami = pd.merge(
                    baza_filtr,
                    katalog,
                    on="OKPO",
                    how="left"
                )[["SOOGU", qiymat_ustun]]

                # Group by ADRES
                natija = (
                    jami
                    .groupby("SOOGU", as_index=False)[qiymat_ustun]
                    .sum()
                )

                # summalar = natija.select_dtypes(include="number").sum()

                # jami_qator = {"SOOGU": "O'zbekiston Respublikasi"}
                # jami_qator.update(summalar.to_dict())

                # natija = pd.concat([natija, pd.DataFrame([jami_qator])], ignore_index=True)



                return natija


            OY221_G1=hisobla_baza(221,"G1" )
            OY221_G2=hisobla_baza(221,"G2" )
            OY221_G3=hisobla_baza(221,"G3" )
            OY221_G4=hisobla_baza(221,"G4" )
            OY221_G5=hisobla_baza(221,"G5" )
            OY222_G1=hisobla_baza(222,"G1" )
            OY222_G2=hisobla_baza(222,"G2" )
            OY222_G3=hisobla_baza(222,"G3" )
            OY222_G4=hisobla_baza(222,"G4" )
            OY222_G5=hisobla_baza(222,"G5" )

            # Misol uchun: 5 ta df bor
            dfs = [OY221_G1,OY221_G2,OY221_G3,OY221_G4,OY221_G5,OY222_G1,OY222_G2,OY222_G3,OY222_G4,OY222_G5]
            yakuniy = reduce(lambda left, right: pd.merge(left, right, on="SOOGU", how="outer"), dfs)
            pd.options.display.float_format = '{:,.1f}'.format
            yakuniy
            HAM_kor_deb=pd.merge(yakuniy,SOOGU,on="SOOGU",how="left")
            HAM_kor_deb


            total = pd.DataFrame({
                    "SOOGU": ["00000"],
                    "G1_x": [HAM_kor_deb["G1_x"].sum()],
                    "G2_x": [HAM_kor_deb["G2_x"].sum()],
                    "G3_x": [HAM_kor_deb["G3_x"].sum()],
                    "G4_x": [HAM_kor_deb["G4_x"].sum()],
                    "G5_x": [HAM_kor_deb["G5_x"].sum()],
                    "G1_y": [HAM_kor_deb["G1_y"].sum()],
                    "G2_y": [HAM_kor_deb["G2_y"].sum()],
                    "G3_y": [HAM_kor_deb["G3_y"].sum()],
                    "G4_y": [HAM_kor_deb["G4_y"].sum()],
                    "G5_y": [HAM_kor_deb["G5_y"].sum()],
                    "NAIM":["Всего по видам экономической деятельности"],
                    "NAIMUZ":["Jami iqtisodiy faoliyat turlari bo‘yicha"]

                })
            VAZIR_DEB = pd.concat([total, HAM_kor_deb], ignore_index=True)
            colm = ["G1_x","G2_x","G3_x","G4_x","G5_x","G1_y","G2_y","G3_y","G4_y","G5_y"]

            VAZIR_DEB = VAZIR_DEB[~(VAZIR_DEB[colm] == 0).all(axis=1)]
            VAZIR_DEB
            VAZIR_DEB=VAZIR_DEB.set_index(["NAIMUZ","NAIM"]).astype(int).div(1000).reset_index()
            VAZIR_DEB=VAZIR_DEB[["NAIMUZ","G1_x","G2_x","G3_x","G4_x","G5_x","G1_y","G2_y","G3_y","G4_y","G5_y","NAIM"]]

            cols =["G1_x","G2_x","G3_x","G4_x","G5_x","G1_y","G2_y","G3_y","G4_y","G5_y"]

            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                VAZIR_DEB[col] = VAZIR_DEB[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                VAZIR_DEB[col] = pd.to_numeric(VAZIR_DEB[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                VAZIR_DEB[col] = VAZIR_DEB[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


            VAZIR_DEB=VAZIR_DEB.fillna(" ").replace("0","-")
            OY_VAZIR_DEB=VAZIR_DEB
            #================================================================================================

            from functools import reduce




            #===================================================================================================================================

            AVGUST_K=katalog09[["OKPO","OKED","SEK1"]].copy()

            AVGUST_B=baza09[["OKPO","G1","G3","G4","G5","G6","G2","SATR"]].copy()


            AVGUST_B_K=pd.merge(AVGUST_B,AVGUST_K,on="OKPO",how="left")

            AVGUST_B_K["kod2"]=AVGUST_B_K["OKED"].astype(str).str[:2]
            AVGUST_B_K
            oked_avgust
            oked=oked_avgust



            def BAJAR(satr, ustn):
                # 1. SATR bo‘yicha filtrlash
                B = AVGUST_B_K[AVGUST_B_K["SATR"] == satr]

                # 2. OKED nomlari bilan bog‘lash
                J = pd.merge(B, oked_avgust, on="OKED", how="left")

                # 3. TO‘G‘RI groupby (SEK1 YO‘QOLMAYDI)
                JAM = (
                    J.groupby(["SEK1", "kod2_x"], as_index=False)
                    .agg(SUM=(ustn, "sum"))
                )

                # 4. Mingga bo‘lish
                JAM["SUM"] = JAM["SUM"] / 1000

                return JAM

            dfs = [
                BAJAR(221, "G1"),
                BAJAR(221, "G2"),
                BAJAR(221, "G3"),
                BAJAR(221, "G4"),
                BAJAR(221, "G5"),
                BAJAR(222, "G1"),
                BAJAR(222, "G2"),
                BAJAR(222, "G3"),
                BAJAR(222, "G4"),
                BAJAR(222, "G5"),
            ]


            for i, df in enumerate(dfs, start=1):
                df.rename(columns={"SUM": f"G_{i}"}, inplace=True)

            yakuniy_B_deb = reduce(
                lambda left, right: pd.merge(
                    left,
                    right,
                    on=["SEK1", "kod2_x"],
                    how="outer"
                ),
                dfs
            )

            yakuniy_B_deb


            for col in ["G_1", "G_2","G_3","G_4","G_5","G_6","G_7","G_8","G_9","G_10"]:
                yakuniy_B_deb[col] = yakuniy_B_deb[col].astype(str).str.replace(",", "").astype(float)








            # 21 ta harf (A–U)
            letters = list(string.ascii_uppercase[:21])  
            # ['A','B','C',...,'U']

            # UMUMIY satrlarni qo‘shamiz
            agg_df = yakuniy_B_deb.groupby("SEK1").sum(numeric_only=True).reset_index()

            # Har bir umumiy satrga tartib bo'yicha A, B, C, ...
            agg_df["kod2_x"] = letters[:len(agg_df)]

            # Birlashtiramiz
            result = pd.concat([agg_df, yakuniy_B_deb], ignore_index=True)

            # Maxsus tartiblash: harflar (A,B,C...) tepada chiqishi uchun
            result["order"] = result["kod2_x"].apply(lambda x: 0 if x in letters else 1)

            # Har SEK1 ichida harflar tepada
            result = result.sort_values(["SEK1", "order"]).drop(columns="order")


            result = result.set_index(["SEK1","kod2_x"]).drop_duplicates()
            result=result.reset_index()


            oked=oked_avgust


            oked["kod2_x"]=oked["OKED"].str[:2]
            oked=oked[["kod2_x","naimuz","naim"]]
            oked

            # DF2 dan birinchi uchragan qatorni qoldiramiz
            df2_first = oked.drop_duplicates(subset=["kod2_x"], keep="first")

            # Endi oddiy merge
            KOR_DEB = result.merge(df2_first, on="kod2_x", how="left")

            KOR_DEB=KOR_DEB[["naimuz","G_1", "G_2","G_3","G_4","G_5","G_6","G_7","G_8","G_9","G_10","naim"]]




            letters = list(string.ascii_uppercase[:21]) 
            nn=oked[oked["kod2_x"].isin(letters)]
            nn=nn["naimuz"].reset_index()
            nn

            # nn dagi naimuz qiymatlari ro‘yxati
            values = nn["naimuz"].unique()

            # KOR_KIR_G2 dan mos kelgan qatorlarni olish
            filtered = KOR_DEB[KOR_DEB["naimuz"].isin(values)]

            # Endi G ustunlarini sum qilish
            RES = pd.DataFrame({
                "naimuz": ["Jami iqtisodiy faoliyat turlari bo‘yicha"],

                "G_1": [filtered["G_1"].sum()],
                "G_2": [filtered["G_2"].sum()],
                "G_3": [filtered["G_3"].sum()],
                "G_4": [filtered["G_4"].sum()],
                "G_5": [filtered["G_5"].sum()],
                "G_6": [filtered["G_6"].sum()],
                "G_7": [filtered["G_7"].sum()],
                "G_8": [filtered["G_8"].sum()],
                "G_9": [filtered["G_9"].sum()],
                "G_10": [filtered["G_10"].sum()],
            
                "naim": ["Всего по видам экономической деятельности"]
            })



            KOR_DEB=pd.concat([RES,KOR_DEB],ignore_index=True)
            KOR_DEB

            cols =["G_1", "G_2","G_3","G_4","G_5","G_6","G_7","G_8","G_9","G_10"]
            KOR_DEB = KOR_DEB[~(KOR_DEB[cols] == 0).all(axis=1)]
            for col in cols:
                # 1) Minglik vergullarni olib tashlash
                KOR_DEB[col] = KOR_DEB[col].astype(str).str.replace(",", "", regex=False).str.strip()

                # 2) Floatga o'zgartirish
                KOR_DEB[col] = pd.to_numeric(KOR_DEB[col], errors="coerce")

                # 3) Formatlash: butun bo‘lsa butun, kasr bo‘lsa kasr
                KOR_DEB[col] = KOR_DEB[col].apply(
                lambda x: (
                    f"{float(x):.1f}".rstrip('0').rstrip('.') 
                    if pd.notnull(x) else x
                )
                )


            KOR_DEB=KOR_DEB.fillna(" ").replace("0","-")
            OY_KOR_DEB=KOR_DEB



            yoz_sheetga(YAKUN_DEB,wb,"YAKUN_DEB")
            yoz_sheetga(YAKUN_KIR,wb,"YAKUN_KIR")
            yoz_sheetga(YAKUN_DEB_G2,wb,"YAKUN_DEB_G2")
            yoz_sheetga(YAKUN_KIR_G2,wb,"YAKUN_KIR_G2")
            yoz_sheetga(VAZIR_DEB_DEB,wb,"VAZIR_DEB")
            yoz_sheetga(VAZIR_KIR,wb,"VAZIR_KIR")
            yoz_sheetga(VAZIR_DEB_G2,wb,"VAZIR_DEB_G2")
            yoz_sheetga(VAZIR_KIR_G2,wb,"VAZIR_KIR_G2")
            yoz_sheetga(KOR_KOR,wb,"KOR_DEB")
            yoz_sheetga(KOR_KIR,wb,"KOR_KIR")
            yoz_sheetga(KOR_DEB_G2,wb,"KOR_DEB_G2")
            yoz_sheetga(KOR_KIR_G2,wb,"KOR_KIR_G2")

            yoz_sheetga(YAKUN_D_M,wb,"YAKUN_D_M")
            yoz_sheetga(YAKUN_K_M,wb,"YAKUN_K_M")
            yoz_sheetga(YAKUN_D_M_G2,wb,"YAKUN_D_M_G2")
            yoz_sheetga(YAKUN_K_M_G2,wb,"YAKUN_K_M_G2")
            yoz_sheetga(MINS_DEB,wb,"MINS_DEB")
            yoz_sheetga(MINS_KIR,wb,"MINS_KIR")
            yoz_sheetga(MINS_DEB_G2,wb,"MINS_DEB_G2")
            yoz_sheetga(MINS_KIR_G2,wb,"MINS_KIR_G2")
            yoz_sheetga(OY_VILOYAT,wb,"OY_VILOYAT",start_row=5)
            yoz_sheetga(OY_VAZIR_DEB,wb,"OY_VAZIR_DEB",start_row=5)
            yoz_sheetga(OY_KOR_DEB,wb,"OY_KOR_DEB",start_row=5)

        


            # 🔹 Excel → BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # 🔹 PDF
            # pdf_buffer = excel_bytes_to_pdf_bytes(excel_buffer)

            # 🔽 Yuklab olish
            st.download_button(
                "📥 Excel yuklab olish",
                data=excel_buffer,
                file_name=f"{output_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # st.download_button(
        #     "📄 PDF yuklab olish",
        #     data=pdf_buffer,
        #     file_name=f"{output_name}.pdf",
        #     mime="application/pdf"
        # )










# cd "C:\Users\hp\Desktop\python\12 MOLIYA.APP"
# streamlit run  app.py




